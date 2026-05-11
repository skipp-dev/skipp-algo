from __future__ import annotations

from io import BytesIO
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

import databento_volatility_screener as dvs
import scripts.databento_production_export as export_mod
import scripts.databento_production_workbook as workbook_mod
from scripts.databento_production_workbook import WorkbookWriteResult, write_databento_production_workbook_from_frames


def test_shared_workbook_writer_is_deterministic_for_fixed_input(tmp_path: Path) -> None:
    summary = pd.DataFrame(
        [
            {
                "trade_date": "2026-03-06",
                "symbol": "AAPL",
                "window_range_pct": 2.1,
                "realized_vol_pct": 1.2,
                "window_return_pct": 0.5,
                "prev_close_to_premarket_pct": 1.1,
                "premarket_to_open_pct": 0.4,
                "open_to_current_pct": 0.2,
            }
        ]
    )
    additional = {"daily_bars": pd.DataFrame([{"trade_date": "2026-03-06", "symbol": "AAPL", "close": 180.0}])}

    left = write_databento_production_workbook_from_frames(
        summary=summary,
        output_path=tmp_path / "left.xlsx",
        generated_at=1_700_000_000.0,
        additional_sheets=additional,
    )
    # Sleep across a full second to force openpyxl's internal datetime.utcnow()
    # to differ between the two writes. Pre-fix this caused docProps/core.xml
    # <dcterms:created>/<dcterms:modified> to drift and the byte comparison
    # failed on CI (PR #31). Post-fix, generated_at pins both timestamps so
    # the workbooks remain byte-identical regardless of wall-clock skew.
    import time as _time

    _time.sleep(1.1)
    right = write_databento_production_workbook_from_frames(
        summary=summary,
        output_path=tmp_path / "right.xlsx",
        generated_at=1_700_000_000.0,
        additional_sheets=additional,
    )

    assert left.generated_at == right.generated_at
    assert left.row_counts == right.row_counts
    assert left.output_path.read_bytes() == right.output_path.read_bytes()


def test_streamlit_create_excel_workbook_reuses_shared_helper(monkeypatch) -> None:
    calls: list[dict[str, object]] = []

    def _fake_create_excel_workbook_bytes(
        summary: pd.DataFrame,
        *,
        minute_detail: pd.DataFrame | None = None,
        second_detail: pd.DataFrame | None = None,
        additional_sheets: dict[str, pd.DataFrame] | None = None,
    ) -> bytes:
        calls.append(
            {
                "summary_rows": len(summary),
                "minute_rows": 0 if minute_detail is None else len(minute_detail),
                "second_rows": 0 if second_detail is None else len(second_detail),
                "sheet_count": 0 if additional_sheets is None else len(additional_sheets),
            }
        )
        return b"xlsx-bytes"

    monkeypatch.setattr(dvs, "create_excel_workbook_bytes", _fake_create_excel_workbook_bytes)

    payload = dvs.create_excel_workbook(pd.DataFrame([{"symbol": "AAPL"}]))
    assert payload == b"xlsx-bytes"
    assert len(calls) == 1
    assert calls[0]["summary_rows"] == 1


def test_create_excel_workbook_bytes_splits_oversized_sheets(monkeypatch) -> None:
    monkeypatch.setattr(workbook_mod, "EXCEL_MAX_ROWS_PER_SHEET", 4)

    summary = pd.DataFrame(
        [
            {
                "trade_date": "2026-03-06",
                "symbol": f"SYM{idx}",
                "window_range_pct": 1.0,
                "realized_vol_pct": 1.0,
                "window_return_pct": 1.0,
                "prev_close_to_premarket_pct": 1.0,
                "premarket_to_open_pct": 1.0,
                "open_to_current_pct": 1.0,
            }
            for idx in range(4)
        ]
    )
    second_detail = pd.DataFrame({"symbol": ["A", "B", "C", "D", "E"], "value": [1, 2, 3, 4, 5]})
    payload = workbook_mod.create_excel_workbook_bytes(summary=summary, second_detail=second_detail)

    workbook = load_workbook(filename=BytesIO(payload))
    assert "summary" in workbook.sheetnames
    assert "summary_002" in workbook.sheetnames
    assert "second_detail" in workbook.sheetnames
    assert "second_detail_002" in workbook.sheetnames


def test_create_excel_workbook_bytes_uses_header_aware_row_limit(monkeypatch) -> None:
    monkeypatch.setattr(workbook_mod, "EXCEL_MAX_ROWS_PER_SHEET", 4)

    rows_fit_with_header = pd.DataFrame(
        {
            "trade_date": ["2026-03-06"] * 3,
            "symbol": ["A", "B", "C"],
            "window_range_pct": [1.0, 1.0, 1.0],
            "realized_vol_pct": [1.0, 1.0, 1.0],
            "window_return_pct": [1.0, 1.0, 1.0],
            "prev_close_to_premarket_pct": [1.0, 1.0, 1.0],
            "premarket_to_open_pct": [1.0, 1.0, 1.0],
            "open_to_current_pct": [1.0, 1.0, 1.0],
        }
    )
    payload_fit = workbook_mod.create_excel_workbook_bytes(summary=rows_fit_with_header)
    workbook_fit = load_workbook(filename=BytesIO(payload_fit))
    assert "summary" in workbook_fit.sheetnames
    assert "summary_002" not in workbook_fit.sheetnames

    rows_overflow_by_one = pd.concat(
        [rows_fit_with_header, rows_fit_with_header.iloc[[0]].assign(symbol="D")],
        ignore_index=True,
    )
    payload_split = workbook_mod.create_excel_workbook_bytes(summary=rows_overflow_by_one)
    workbook_split = load_workbook(filename=BytesIO(payload_split))
    assert "summary" in workbook_split.sheetnames
    assert "summary_002" in workbook_split.sheetnames


def test_production_pipeline_canonical_workbook_helper_invokes_shared_writer(monkeypatch, tmp_path: Path) -> None:
    recorded: dict[str, object] = {}

    def _fake_writer(**kwargs):
        recorded.update(kwargs)
        out_path = Path(kwargs["output_path"])
        out_path.parent.mkdir(parents=True, exist_ok=True)
        out_path.write_bytes(b"ok")
        return WorkbookWriteResult(
            output_path=out_path,
            generated_at="2026-03-26T00:00:00+00:00",
            row_counts={"summary": len(kwargs["summary"])},
            sheet_names=["summary"],
            canonical_upstream_artifact="databento_production_export_bundle",
        )

    monkeypatch.setattr(export_mod, "write_databento_production_workbook_from_frames", _fake_writer)

    path = export_mod._write_canonical_production_workbook(
        export_dir=tmp_path,
        summary=pd.DataFrame([{"symbol": "AAPL"}]),
        minute_detail=pd.DataFrame(),
        second_detail=pd.DataFrame(),
        manifest={"dataset": "DBEQ.BASIC"},
        raw_universe=pd.DataFrame(),
        daily_bars=pd.DataFrame(),
        intraday=pd.DataFrame(),
        ranked=pd.DataFrame(),
        daily_symbol_features_full_universe=pd.DataFrame(),
        full_universe_second_detail_open=pd.DataFrame(),
        full_universe_second_detail_close=pd.DataFrame(),
        full_universe_close_trade_detail=pd.DataFrame(),
        full_universe_close_outcome_minute=pd.DataFrame(),
        close_imbalance_features_full_universe=pd.DataFrame(),
        close_imbalance_outcomes_full_universe=pd.DataFrame(),
        premarket_features_full_universe=pd.DataFrame(),
        premarket_window_features_full_universe=pd.DataFrame(),
        symbol_day_diagnostics=pd.DataFrame(),
        research_event_flags_full_universe=pd.DataFrame(),
        research_event_flag_coverage=pd.DataFrame(),
        research_event_flag_trade_date_distribution=pd.DataFrame(),
        research_event_flag_outcome_slices=pd.DataFrame(),
        research_news_flags_full_universe=pd.DataFrame(),
        research_news_flag_coverage=pd.DataFrame(),
        research_news_flag_trade_date_distribution=pd.DataFrame(),
        research_news_flag_outcome_slices=pd.DataFrame(),
        core_vs_benzinga_news_side_by_side=pd.DataFrame(),
        core_vs_benzinga_news_overlap_stats=pd.DataFrame(),
        quality_window_status=pd.DataFrame(),
        batl_debug={},
        output_summary={},
    )

    assert path.exists()
    assert recorded["output_path"] == tmp_path / "databento_volatility_production_workbook.xlsx"


def test_production_pipeline_canonical_workbook_helper_slims_base_only_sheet_set(monkeypatch, tmp_path: Path) -> None:
    recorded: dict[str, object] = {}

    def _fake_writer(**kwargs):
        recorded.update(kwargs)
        out_path = Path(kwargs["output_path"])
        out_path.parent.mkdir(parents=True, exist_ok=True)
        out_path.write_bytes(b"ok")
        return WorkbookWriteResult(
            output_path=out_path,
            generated_at="2026-03-26T00:00:00+00:00",
            row_counts={"summary": len(kwargs["summary"])},
            sheet_names=["summary"],
            canonical_upstream_artifact="databento_production_export_bundle",
        )

    monkeypatch.setattr(export_mod, "write_databento_production_workbook_from_frames", _fake_writer)

    path = export_mod._write_canonical_production_workbook(
        export_dir=tmp_path,
        summary=pd.DataFrame([{"symbol": "AAPL"}]),
        minute_detail=pd.DataFrame([{"timestamp": "2026-03-26T13:30:00Z", "price": 1.0}]),
        second_detail=pd.DataFrame([{"timestamp": "2026-03-26T13:30:00Z", "price": 1.0}]),
        manifest={"dataset": "DBEQ.BASIC"},
        raw_universe=pd.DataFrame([{"symbol": "AAPL"}]),
        daily_bars=pd.DataFrame([{"trade_date": "2026-03-26", "symbol": "AAPL", "close": 1.0}]),
        intraday=pd.DataFrame([{"trade_date": "2026-03-26", "symbol": "AAPL"}]),
        ranked=pd.DataFrame([{"trade_date": "2026-03-26", "symbol": "AAPL"}]),
        daily_symbol_features_full_universe=pd.DataFrame([{"trade_date": "2026-03-26", "symbol": "AAPL"}]),
        full_universe_second_detail_open=pd.DataFrame([{"trade_date": "2026-03-26", "symbol": "AAPL"}]),
        full_universe_second_detail_close=pd.DataFrame([{"trade_date": "2026-03-26", "symbol": "AAPL"}]),
        full_universe_close_trade_detail=pd.DataFrame([{"trade_date": "2026-03-26", "symbol": "AAPL"}]),
        full_universe_close_outcome_minute=pd.DataFrame([{"trade_date": "2026-03-26", "symbol": "AAPL"}]),
        close_imbalance_features_full_universe=pd.DataFrame([{"trade_date": "2026-03-26", "symbol": "AAPL"}]),
        close_imbalance_outcomes_full_universe=pd.DataFrame([{"trade_date": "2026-03-26", "symbol": "AAPL"}]),
        premarket_features_full_universe=pd.DataFrame([{"trade_date": "2026-03-26", "symbol": "AAPL"}]),
        premarket_window_features_full_universe=pd.DataFrame([{"trade_date": "2026-03-26", "symbol": "AAPL"}]),
        symbol_day_diagnostics=pd.DataFrame([{"trade_date": "2026-03-26", "symbol": "AAPL"}]),
        research_event_flags_full_universe=pd.DataFrame([{"trade_date": "2026-03-26", "symbol": "AAPL"}]),
        research_event_flag_coverage=pd.DataFrame([{"trade_date": "2026-03-26"}]),
        research_event_flag_trade_date_distribution=pd.DataFrame([{"trade_date": "2026-03-26"}]),
        research_event_flag_outcome_slices=pd.DataFrame([{"bucket": "all"}]),
        research_news_flags_full_universe=pd.DataFrame([{"trade_date": "2026-03-26", "symbol": "AAPL"}]),
        research_news_flag_coverage=pd.DataFrame([{"trade_date": "2026-03-26"}]),
        research_news_flag_trade_date_distribution=pd.DataFrame([{"trade_date": "2026-03-26"}]),
        research_news_flag_outcome_slices=pd.DataFrame([{"bucket": "all"}]),
        core_vs_benzinga_news_side_by_side=pd.DataFrame([{"trade_date": "2026-03-26", "symbol": "AAPL"}]),
        core_vs_benzinga_news_overlap_stats=pd.DataFrame([{"bucket": "all"}]),
        quality_window_status=pd.DataFrame([{"symbol": "AAPL"}]),
        batl_debug={"ok": True},
        output_summary={"rows": 1},
        smc_base_only=True,
    )

    assert path.exists()
    assert isinstance(recorded["minute_detail"], pd.DataFrame)
    assert recorded["minute_detail"].empty
    assert isinstance(recorded["second_detail"], pd.DataFrame)
    assert recorded["second_detail"].empty
    assert isinstance(recorded["additional_sheets"], dict)
    assert set(recorded["additional_sheets"].keys()) == {
        "manifest",
        "daily_bars",
        "batl_debug",
        "output_checks",
    }


def test_create_excel_workbook_bytes_emits_progress_per_sheet() -> None:
    """Q1 obs(workbook): per-sheet progress markers are emitted into a callback.

    Without a callback the writer must remain silent (no behavior change).
    With a callback every sheet emits begin+done markers, plus pre-/post-
    serialization markers around the openpyxl context exit and the xlsx
    zip-timestamp normalization.
    """
    summary = pd.DataFrame(
        [
            {
                "trade_date": "2026-03-06",
                "symbol": "AAPL",
                "window_range_pct": 2.1,
                "realized_vol_pct": 1.2,
                "window_return_pct": 0.5,
                "prev_close_to_premarket_pct": 1.1,
                "premarket_to_open_pct": 0.4,
                "open_to_current_pct": 0.2,
            }
        ]
    )
    additional = {
        "daily_bars": pd.DataFrame(
            [{"trade_date": "2026-03-06", "symbol": "AAPL", "close": 180.0}]
        ),
        "manifest": pd.DataFrame([{"key": "v", "value": "1"}]),
    }

    # Default behavior: no callback → no progress emitted, payload still valid.
    payload_silent = workbook_mod.create_excel_workbook_bytes(
        summary=summary,
        additional_sheets=additional,
        generated_at=1_700_000_000.0,
    )
    assert isinstance(payload_silent, bytes) and len(payload_silent) > 0

    msgs: list[str] = []
    payload = workbook_mod.create_excel_workbook_bytes(
        summary=summary,
        additional_sheets=additional,
        generated_at=1_700_000_000.0,
        progress_callback=msgs.append,
    )
    assert isinstance(payload, bytes) and len(payload) > 0

    # All progress markers must use the agreed "workbook: " prefix so the
    # producer's existing _progress closure can route them uniformly.
    assert msgs, "expected at least one progress marker when callback is supplied"
    assert all(m.startswith("workbook: ") for m in msgs), msgs

    joined = "\n".join(msgs)
    # Per-sheet begin/done markers cover summary + every additional sheet.
    assert "'summary' rows=1 begin" in joined, joined
    assert "'summary' rows=1 done" in joined, joined
    assert "'daily_bars' rows=1 begin" in joined, joined
    assert "'manifest' rows=1 begin" in joined, joined
    # Per-chunk heartbeats emitted by _write_chunked_sheet (diagnostic for
    # GHA no-output-watchdog at lookback=30 — see runs 25568632083 et al.).
    assert "chunk 1 rows=0-1" in joined, joined
    assert "chunk 1 to_excel done in" in joined, joined
    # Styling-pass markers with header/col_width split timing.
    assert "styling ws 1/3 'summary'" in joined, joined
    assert "styling pass total" in joined, joined
    # Conditional formatting phase timing.
    assert "conditional formatting applied in" in joined, joined
    # __exit__ boundary markers (XML serialization timing).
    assert "openpyxl writer __exit__ begin (XML serialization)" in joined, joined
    assert "openpyxl writer __exit__ complete in" in joined, joined
    assert "normalize_xlsx_zip_timestamps begin" in joined, joined
    assert "normalize_xlsx_zip_timestamps done" in joined, joined


def test_write_databento_production_workbook_from_frames_threads_progress(tmp_path: Path) -> None:
    """The public writer entry point forwards the callback to the bytes writer
    and additionally emits its own write_bytes begin/done markers."""
    summary = pd.DataFrame(
        [
            {
                "trade_date": "2026-03-06",
                "symbol": "AAPL",
                "window_range_pct": 1.0,
                "realized_vol_pct": 1.0,
                "window_return_pct": 1.0,
                "prev_close_to_premarket_pct": 1.0,
                "premarket_to_open_pct": 1.0,
                "open_to_current_pct": 1.0,
            }
        ]
    )
    msgs: list[str] = []
    out = tmp_path / "out.xlsx"
    result = write_databento_production_workbook_from_frames(
        summary=summary,
        output_path=out,
        generated_at=1_700_000_000.0,
        progress_callback=msgs.append,
    )
    assert result.output_path == out
    assert out.exists() and out.stat().st_size > 0
    joined = "\n".join(msgs)
    assert "workbook: write_bytes begin" in joined, joined
    assert "workbook: write_bytes done" in joined, joined
    # All progress lines (including write_bytes markers) carry the elapsed-time suffix.
    assert all("(t+" in m for m in msgs), msgs
    # Inner per-sheet markers must also be present (callback was forwarded).
    assert "'summary' rows=1 begin" in joined, joined

