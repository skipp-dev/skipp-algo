"""Shared Databento production workbook builder.

Canonical upstream artifact policy:
- Canonical artifact: Databento production export bundle (manifest + parquet frames).
- Derived artifact: production workbook (.xlsx) generated from canonical frames.

This module centralizes workbook generation so daily/base production paths and UI paths
use the same producer logic.
"""

from __future__ import annotations

import zipfile
from dataclasses import dataclass
from datetime import UTC, datetime
from io import BytesIO
from pathlib import Path
from time import perf_counter
from collections.abc import Callable
from typing import Any

import pandas as pd
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pandas.api.types import is_datetime64_any_dtype

from scripts.load_databento_export_bundle import load_export_bundle

DEFAULT_PRODUCTION_EXPORT_DIR = Path("artifacts") / "smc_microstructure_exports"
CANONICAL_PRODUCTION_WORKBOOK_NAME = "databento_volatility_production_workbook.xlsx"
LEGACY_WORKBOOK_GLOB = "databento_volatility_production_*.xlsx"
EXCEL_MAX_ROWS_PER_SHEET = 1_048_576


@dataclass(frozen=True)
class WorkbookWriteResult:
    output_path: Path
    generated_at: str
    row_counts: dict[str, int]
    sheet_names: list[str]
    canonical_upstream_artifact: str


def canonical_production_workbook_path(*, export_dir: Path | None = None) -> Path:
    root = export_dir if export_dir is not None else DEFAULT_PRODUCTION_EXPORT_DIR
    return root / CANONICAL_PRODUCTION_WORKBOOK_NAME


def resolve_production_workbook_path(
    workbook: str | Path | None = None,
    *,
    export_dir: Path | None = None,
    repo_root: Path | None = None,
) -> Path:
    """Resolve workbook path with canonical-first fallback order.

    Freshness model:
    - Canonical deterministic workbook path is per daily export run and overwritten.
    - Timestamped workbook exports are still supported as fallback.
    """

    base_dir = export_dir if export_dir is not None else DEFAULT_PRODUCTION_EXPORT_DIR
    root = repo_root if repo_root is not None else Path.cwd()

    if workbook is not None:
        explicit = Path(workbook).expanduser()
        if explicit.exists():
            return explicit
        raise FileNotFoundError(f"workbook not found: {explicit}")

    canonical = canonical_production_workbook_path(export_dir=base_dir)
    if canonical.exists():
        return canonical

    from scripts.smc_artifact_resolver import latest_by_filename_iso

    timestamped = latest_by_filename_iso(base_dir.glob(LEGACY_WORKBOOK_GLOB))
    if timestamped is not None:
        return timestamped

    legacy_root = latest_by_filename_iso(root.glob(LEGACY_WORKBOOK_GLOB))
    if legacy_root is not None:
        return legacy_root

    raise FileNotFoundError(
        "No Databento production workbook found. Expected canonical path "
        f"{canonical} or timestamped exports matching {LEGACY_WORKBOOK_GLOB}."
    )


def prepare_frame_for_excel(frame: pd.DataFrame) -> pd.DataFrame:
    sanitized = frame.copy()
    for column in sanitized.columns:
        series = sanitized[column]
        if isinstance(series.dtype, pd.DatetimeTZDtype):
            sanitized[column] = series.dt.tz_localize(None)
        elif is_datetime64_any_dtype(series):
            continue
        elif series.dtype == object:
            sanitized[column] = series.map(
                lambda value: (
                    value.tz_localize(None)
                    if isinstance(value, pd.Timestamp) and value.tzinfo is not None
                    else value.replace(tzinfo=None)
                    if isinstance(value, datetime) and value.tzinfo is not None
                    else value
                )
            )
    return sanitized


def create_excel_workbook_bytes(
    summary: pd.DataFrame,
    *,
    minute_detail: pd.DataFrame | None = None,
    second_detail: pd.DataFrame | None = None,
    additional_sheets: dict[str, pd.DataFrame | None] | None = None,
    generated_at: float | None = None,
    progress_callback: Callable[[str], None] | None = None,
) -> bytes:
    # Per-sheet observability for Step 10/10c (was 10/10b before Q5b reorder).
    # Default callback is a no-op so
    # behavior is unchanged for callers (and existing tests). When wired from
    # the producer's `_progress` closure, each sheet emits an elapsed-time
    # marker so cap-hits inside this single-shot openpyxl write can be
    # attributed to a specific sheet rather than appearing as a silent gap.
    _t0 = perf_counter()

    def _emit(msg: str) -> None:
        if progress_callback is None:
            return
        progress_callback(f"workbook: {msg} (t+{perf_counter() - _t0:.1f}s)")

    # Pre-compute the planned sheet list so per-sheet markers can include
    # `idx/total` for monotonic progress tracking.
    planned_sheets: list[tuple[str, pd.DataFrame]] = [("summary", summary)]
    if additional_sheets:
        for sheet_name, frame in additional_sheets.items():
            if frame is None or frame.empty:
                continue
            planned_sheets.append((str(sheet_name)[:31], frame))
    if minute_detail is not None and not minute_detail.empty:
        planned_sheets.append(("minute_detail", minute_detail))
    if second_detail is not None and not second_detail.empty:
        planned_sheets.append(("second_detail", second_detail))
    total_sheets = len(planned_sheets)
    _emit(f"begin openpyxl write, sheets={total_sheets}")

    def _write_chunked_sheet(
        writer: pd.ExcelWriter,
        frame: pd.DataFrame,
        *,
        sheet_name: str,
        sheet_index: int,
    ) -> None:
        prepared = prepare_frame_for_excel(frame)
        rows = len(prepared)
        _emit(f"sheet {sheet_index}/{total_sheets} '{sheet_name}' rows={rows} begin")
        if prepared.empty:
            prepared.to_excel(writer, sheet_name=sheet_name, index=False)
            _emit(f"sheet {sheet_index}/{total_sheets} '{sheet_name}' rows=0 done")
            return
        # Pandas writes a header row, so data rows must stay below the worksheet max.
        max_data_rows_per_sheet = max(1, EXCEL_MAX_ROWS_PER_SHEET - 1)
        base_name = str(sheet_name)[:31]
        for chunk_index, start_row in enumerate(range(0, len(prepared), max_data_rows_per_sheet), start=1):
            end_row = start_row + max_data_rows_per_sheet
            if chunk_index == 1:
                chunk_sheet_name = base_name
            else:
                suffix = f"_{chunk_index:03d}"
                chunk_sheet_name = f"{base_name[: max(0, 31 - len(suffix))]}{suffix}"
            chunk_rows = min(end_row, rows) - start_row
            _emit(
                f"sheet {sheet_index}/{total_sheets} '{sheet_name}' chunk {chunk_index} "
                f"rows={start_row}-{min(end_row, rows)} ({chunk_rows} rows) to_excel begin"
            )
            chunk_t0 = perf_counter()
            prepared.iloc[start_row:end_row].to_excel(writer, sheet_name=chunk_sheet_name, index=False)
            _emit(
                f"sheet {sheet_index}/{total_sheets} '{sheet_name}' chunk {chunk_index} "
                f"to_excel done in {perf_counter() - chunk_t0:.2f}s"
            )
        _emit(f"sheet {sheet_index}/{total_sheets} '{sheet_name}' rows={rows} done")

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for idx, (name, frame) in enumerate(planned_sheets, start=1):
            _write_chunked_sheet(writer, frame, sheet_name=name, sheet_index=idx)

        workbook = writer.book
        total_ws = len(workbook.worksheets)
        _emit(f"all sheets written, applying header styling (worksheets={total_ws})")
        styling_t0 = perf_counter()
        for ws_idx, worksheet in enumerate(workbook.worksheets, start=1):
            ws_t0 = perf_counter()
            rows_ct = worksheet.max_row
            cols_ct = worksheet.max_column
            _emit(
                f"styling ws {ws_idx}/{total_ws} '{worksheet.title}' "
                f"rows={rows_ct} cols={cols_ct} begin"
            )
            worksheet.freeze_panes = "A2"
            worksheet.auto_filter.ref = worksheet.dimensions
            header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
            header_font = Font(color="FFFFFF", bold=True)
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            header_done = perf_counter()
            for column_cells in worksheet.columns:
                max_length = max(len(str(cell.value or "")) for cell in column_cells)
                worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_length + 2, 12), 28)
            width_done = perf_counter()
            _emit(
                f"styling ws {ws_idx}/{total_ws} '{worksheet.title}' done "
                f"(header={header_done - ws_t0:.2f}s col_width={width_done - header_done:.2f}s "
                f"total={width_done - ws_t0:.2f}s)"
            )
        _emit(f"styling pass total {perf_counter() - styling_t0:.2f}s")

        cf_t0 = perf_counter()
        summary_sheet = workbook["summary"]
        headers = {cell.value: idx + 1 for idx, cell in enumerate(summary_sheet[1])}
        heat_columns = [
            "window_range_pct",
            "realized_vol_pct",
            "window_return_pct",
            "prev_close_to_premarket_pct",
            "premarket_to_open_pct",
            "open_to_current_pct",
        ]
        for col_name in heat_columns:
            col_idx = headers.get(col_name)
            if col_idx is None or summary_sheet.max_row < 2:
                continue
            letter = get_column_letter(col_idx)
            summary_sheet.conditional_formatting.add(
                f"{letter}2:{letter}{summary_sheet.max_row}",
                ColorScaleRule(
                    start_type="num",
                    start_value=-10,
                    start_color="C00000",
                    mid_type="num",
                    mid_value=0,
                    mid_color="FFF2CC",
                    end_type="num",
                    end_value=10,
                    end_color="63BE7B",
                ),
            )
        _emit(f"conditional formatting applied in {perf_counter() - cf_t0:.2f}s")

        # Pin docProps/core.xml timestamps so byte output is deterministic for a
        # fixed generated_at. Without this, openpyxl writes datetime.utcnow()
        # into <dcterms:created>/<dcterms:modified>, making two calls with the
        # same input produce different bytes (CI flake on PR #31).
        if generated_at is not None:
            pinned = datetime.fromtimestamp(float(generated_at), tz=UTC).replace(tzinfo=None)
            workbook.properties.created = pinned
            workbook.properties.modified = pinned
        _emit("openpyxl writer __exit__ begin (XML serialization)")
        _exit_started = perf_counter()
    _emit(f"openpyxl writer __exit__ complete in {perf_counter() - _exit_started:.2f}s")
    raw = output.getvalue()
    if generated_at is not None:
        _emit(f"normalize_xlsx_zip_timestamps begin (raw_bytes={len(raw)})")
        raw = _normalize_xlsx_zip_timestamps(raw, generated_at=float(generated_at))
        _emit(f"normalize_xlsx_zip_timestamps done (final_bytes={len(raw)})")
    return raw


def _normalize_xlsx_zip_timestamps(raw: bytes, *, generated_at: float) -> bytes:
    """Rewrite an xlsx zip so every entry has a deterministic mtime.

    Python's :mod:`zipfile` stamps each :class:`zipfile.ZipInfo` with the
    wall-clock time at write. That bleeds into the central directory and the
    per-entry local headers, so two byte-identical workbook payloads written a
    second apart produce different bytes. By round-tripping the archive with a
    pinned ``date_time`` we guarantee byte-stability for a fixed input.
    """
    import re

    pinned_dt = datetime.fromtimestamp(generated_at, tz=UTC)
    pinned_tuple = (
        pinned_dt.year,
        pinned_dt.month,
        pinned_dt.day,
        pinned_dt.hour,
        pinned_dt.minute,
        pinned_dt.second,
    )
    pinned_iso = pinned_dt.strftime("%Y-%m-%dT%H:%M:%SZ").encode("ascii")
    # openpyxl stamps docProps/core.xml's <dcterms:created>/<dcterms:modified>
    # with datetime.utcnow() at save time, even if workbook.properties.modified
    # was set earlier — pandas's writer reassigns it on context exit. Rewrite
    # the embedded XML so a fixed generated_at produces byte-stable output.
    _DCTERMS_RE = re.compile(
        rb'(<dcterms:(?:created|modified)[^>]*>)[^<]*(</dcterms:(?:created|modified)>)'
    )

    def _pin_core_xml(blob: bytes) -> bytes:
        return _DCTERMS_RE.sub(rb"\1" + pinned_iso + rb"\2", blob)

    src = BytesIO(raw)
    dst = BytesIO()
    with zipfile.ZipFile(src, "r") as src_zip, zipfile.ZipFile(
        dst, "w", compression=zipfile.ZIP_DEFLATED
    ) as dst_zip:
        # Sort entries by filename so the central directory order is stable
        # across runs. openpyxl writes via a temp staging area whose iteration
        # order is not guaranteed deterministic.
        for info in sorted(src_zip.infolist(), key=lambda i: i.filename):
            data = src_zip.read(info.filename)
            if info.filename == "docProps/core.xml":
                data = _pin_core_xml(data)
            new_info = zipfile.ZipInfo(filename=info.filename, date_time=pinned_tuple)
            new_info.compress_type = info.compress_type
            new_info.external_attr = info.external_attr
            new_info.create_system = info.create_system
            dst_zip.writestr(new_info, data)
    return dst.getvalue()


def write_databento_production_workbook_from_frames(
    *,
    summary: pd.DataFrame,
    output_path: str | Path,
    generated_at: float | None = None,
    minute_detail: pd.DataFrame | None = None,
    second_detail: pd.DataFrame | None = None,
    additional_sheets: dict[str, pd.DataFrame | None] | None = None,
    progress_callback: Callable[[str], None] | None = None,
) -> WorkbookWriteResult:
    path = Path(output_path).expanduser()
    path.parent.mkdir(parents=True, exist_ok=True)
    payload = create_excel_workbook_bytes(
        summary,
        minute_detail=minute_detail,
        second_detail=second_detail,
        additional_sheets=additional_sheets,
        generated_at=generated_at,
        progress_callback=progress_callback,
    )
    if progress_callback is not None:
        _t0 = perf_counter()
        progress_callback(
            f"workbook: write_bytes begin (bytes={len(payload)}) (t+0.0s)"
        )
    path.write_bytes(payload)
    if progress_callback is not None:
        progress_callback(
            f"workbook: write_bytes done (t+{perf_counter() - _t0:.1f}s)"
        )

    rows = {"summary": len(summary)}
    if minute_detail is not None:
        rows["minute_detail"] = len(minute_detail)
    if second_detail is not None:
        rows["second_detail"] = len(second_detail)
    if additional_sheets:
        for name, frame in additional_sheets.items():
            if frame is None:
                continue
            rows[str(name)] = len(frame)

    return WorkbookWriteResult(
        output_path=path,
        generated_at=(
            datetime.fromtimestamp(float(generated_at), tz=UTC).isoformat(timespec="seconds")
            if generated_at is not None
            else datetime.now(UTC).isoformat(timespec="seconds")
        ),
        row_counts=rows,
        sheet_names=list(rows.keys()),
        canonical_upstream_artifact="databento_production_export_bundle",
    )


def write_databento_production_workbook(
    *,
    export_bundle_path: str | Path | None = None,
    output_path: str | Path,
    generated_at: float | None = None,
    include_sheets: list[str] | None = None,
) -> dict[str, Any]:
    """Write production workbook from canonical export bundle frames."""

    bundle_ref = export_bundle_path if export_bundle_path is not None else DEFAULT_PRODUCTION_EXPORT_DIR
    payload = load_export_bundle(bundle_ref, manifest_prefix="databento_volatility_production_")
    frames = payload["frames"]

    summary = frames.get("summary", pd.DataFrame())
    minute_detail = frames.get("minute_detail")
    second_detail = frames.get("second_detail")

    selected = set(include_sheets or [])
    additional_sheets: dict[str, pd.DataFrame] = {}
    for name in [
        "manifest",
        "universe",
        "daily_bars",
        "intraday",
        "ranked",
        "daily_symbol_features_full_universe",
        "premarket_window_features_full_universe",
        "premarket_features_full_universe",
        "symbol_day_diagnostics",
    ]:
        frame = frames.get(name)
        if not isinstance(frame, pd.DataFrame) or frame.empty:
            continue
        if selected and name not in selected:
            continue
        additional_sheets[name] = frame

    result = write_databento_production_workbook_from_frames(
        summary=summary,
        output_path=output_path,
        generated_at=generated_at,
        minute_detail=minute_detail if isinstance(minute_detail, pd.DataFrame) else None,
        second_detail=second_detail if isinstance(second_detail, pd.DataFrame) else None,
        additional_sheets=additional_sheets,
    )

    return {
        "output_path": str(result.output_path),
        "generated_at": result.generated_at,
        "row_counts": result.row_counts,
        "sheet_names": result.sheet_names,
        "canonical_upstream_artifact": result.canonical_upstream_artifact,
        "bundle_manifest_path": str(payload["manifest_path"]),
    }
