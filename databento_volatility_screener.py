from __future__ import annotations

import hashlib
import json
import logging
import math
import os
import re
import sys
import tempfile
import time as time_module
import warnings
import asyncio
import gc
from dataclasses import dataclass
from datetime import UTC, date, datetime, time, timedelta, tzinfo
from io import BytesIO
from pathlib import Path
from typing import Any, Callable
from urllib.request import Request, urlopen
from zoneinfo import ZoneInfo

import certifi

import numpy as np
import pandas as pd
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pandas.api.types import is_datetime64_any_dtype

from open_prep.macro import FMPClient
from strategy_config import (
    LONG_DIP_ENTRY_EARLY_DIP_MAX_SECONDS,
    LONG_DIP_ENTRY_EARLY_DIP_MIN_PCT,
    LONG_DIP_ENTRY_OPEN30_VOLUME_MIN,
    LONG_DIP_ENTRY_RECLAIM_MAX_SECONDS,
    LONG_DIP_MIN_GAP_PCT,
    LONG_DIP_MIN_PREMARKET_DOLLAR_VOLUME,
)

logger = logging.getLogger(__name__)

_API_KEY_REDACTION_PATTERNS = (
    re.compile(r"(api[_-]?key=)([^&\s]+)", flags=re.IGNORECASE),
    re.compile(r"(token=)([^&\s]+)", flags=re.IGNORECASE),
    re.compile(r"(Authorization:\s*Bearer\s+)([^\s]+)", flags=re.IGNORECASE),
)

US_EASTERN_TZ = ZoneInfo("America/New_York")
DEFAULT_DISPLAY_TZ = "Europe/Berlin"
CACHE_VERSION = "v1"
CACHE_VERSION_BY_CATEGORY = {
    "daily_bars": "v2",
    "symbol_support": "v2",
    "full_universe_open_second_detail": "v2",
    "full_universe_close_trade_detail": "v1",
    "full_universe_close_outcome_minute_detail": "v1",
    "intraday_summary": "v2",
    "symbol_detail_second": "v2",
    "symbol_detail_minute": "v2",
}
CACHE_ROOT = Path(__file__).resolve().parent / "artifacts" / "databento_volatility_cache"
WATCHLIST_SNAPSHOT_FILE = "watchlist_rank_history.parquet"
EXACT_NAMED_EXPORT_STATE_FILE = "databento_exact_named_state.json"
UI_RUNTIME_STATE_KEY = "ui_runtime"
FULL_UNIVERSE_OPTIONAL_FEATURE_COLUMNS = (
    "earnings_date",
    "earnings_time",
    "earnings_surprise_pct",
    "news_score",
    "news_category",
    "analyst_rating",
    "analyst_target_price",
    "filing_date",
    "filing_type",
    "mna_flag",
    "mna_side",
    "float_shares",
    "shares_outstanding",
    "short_interest",
    "short_interest_ratio",
    "short_float_pct",
)
SUPPORTED_DISPLAY_TZ = {
    "America/New_York": ZoneInfo("America/New_York"),
    "Europe/Berlin": ZoneInfo("Europe/Berlin"),
}
PREFERRED_DATABENTO_DATASETS = (
    "XNAS.ITCH",
    "XNYS.PILLAR",
    "DBEQ.BASIC",
    "XNAS.BASIC",
)
# ET-relative defaults for the intraday screening window
_DEFAULT_INTRADAY_PRE_OPEN_MINUTES = 10
_DEFAULT_INTRADAY_POST_OPEN_MINUTES = 30
# ET-relative defaults for the open window detail
_DEFAULT_OPEN_WINDOW_PRE_OPEN_MINUTES = 1
_DEFAULT_OPEN_WINDOW_POST_OPEN_SECONDS = 5 * 60 + 59  # 5:59 after open
# ET-relative defaults for close-imbalance detail
DEFAULT_CLOSE_IMBALANCE_WINDOW_START_ET = time(15, 50)
DEFAULT_CLOSE_IMBALANCE_AUCTION_TIME_ET = time(16, 0)
DEFAULT_CLOSE_IMBALANCE_WINDOW_END_ET = time(16, 5)
DEFAULT_CLOSE_IMBALANCE_AFTERHOURS_END_ET = time(20, 0)
DEFAULT_CLOSE_IMBALANCE_NEXT_DAY_OUTCOME_TIME_ET = time(10, 0)
_DATABENTO_FLAG_BAD_TS_RECV = 1 << 3
_DATABENTO_FLAG_MAYBE_BAD_BOOK = 1 << 2
_DATABENTO_FLAG_PUBLISHER_SPECIFIC = 1 << 1

MAX_SYMBOLS_PER_REQUEST = 2000
SYMBOL_SUPPORT_CHECK_BATCH_SIZE = 250
SYMBOL_SUPPORT_LOOKBACK_DAYS = 14
SYMBOL_SUPPORT_CACHE_TTL_SECONDS = 7 * 24 * 3600  # 7 days
DATA_CACHE_TTL_SECONDS = 4 * 3600  # 4 hours – guards against stale intra-day caches
RECENT_INTRADAY_CACHE_TTL_SECONDS = DATA_CACHE_TTL_SECONDS
DATABENTO_GET_RANGE_MAX_ATTEMPTS = 3
DATABENTO_SYMBOL_ALIASES = {
    "BRK-A": "BRK.A",
    "BRK-B": "BRK.B",
    "BRK/A": "BRK.A",
    "BRK/B": "BRK.B",
    "BF-B": "BF.B",
    "MKC-V": "MKC.V",
    "MOG-A": "MOG.A",
}
DATABENTO_UNSUPPORTED_SYMBOLS = {
    "CTA-PA",
}
_DATABENTO_INVALID_CHAR_RE = re.compile(r"[^A-Z0-9.]")
_DATABENTO_UNIT_OR_WARRANT_SUFFIXES = (
    ".U",
    ".W",
    ".WS",
    ".R",
    ".RT",
)
UNIVERSE_COLUMNS = ["symbol", "company_name", "exchange", "sector", "industry", "market_cap"]
NASDAQ_TRADER_DIRECTORY_SPECS = (
    (
        "https://www.nasdaqtrader.com/dynamic/SymDir/nasdaqtraded.txt",
        "Symbol",
        "Security Name",
        "Listing Exchange",
    ),
    (
        "https://www.nasdaqtrader.com/dynamic/SymDir/otherlisted.txt",
        "ACT Symbol",
        "Security Name",
        "Exchange",
    ),
)
NASDAQ_TRADER_EXCHANGE_CODE_MAP = {
    "NASDAQ": "Q",
    "NYSE": "N",
    "AMEX": "A",
    "NYSEARCA": "P",
    "NYSE ARCA": "P",
    "BATS": "Z",
    "CBOE": "Z",
}
NASDAQ_TRADER_EXCHANGE_NAME_MAP = {
    "Q": "NASDAQ",
    "N": "NYSE",
    "A": "AMEX",
    "P": "NYSE ARCA",
    "Z": "BATS",
}


@dataclass(frozen=True)
class WindowDefinition:
    trade_date: date
    display_timezone: str
    window_start_local: datetime
    window_end_local: datetime
    fetch_start_utc: datetime
    fetch_end_utc: datetime
    regular_open_utc: datetime
    premarket_anchor_utc: datetime


@dataclass
class SymbolDayState:
    symbol: str
    trade_date: date
    first_window_open: float | None = None
    last_window_close: float | None = None
    last_window_timestamp: pd.Timestamp | None = None
    window_high: float | None = None
    window_low: float | None = None
    window_volume: float = 0.0
    second_count: int = 0
    premarket_price: float | None = None
    market_open_price: float | None = None
    realized_var: float = 0.0
    _last_window_close_for_rv: float | None = None


@dataclass(frozen=True)
class DataStatusResult:
    export_generated_at: str | None
    daily_bars_fetched_at: str | None
    intraday_fetched_at: str | None
    premarket_fetched_at: str | None
    second_detail_fetched_at: str | None
    dataset: str | None
    lookback_days: int | None
    trade_dates_covered: tuple[str, ...]
    is_stale: bool
    staleness_reason: str
    manifest_path: str | None


EXACT_EXPORT_STATUS_FILES = {
    "daily_bars_fetched_at": "daily_symbol_features_full_universe.parquet",
    "premarket_fetched_at": "premarket_features_full_universe.parquet",
    "second_detail_fetched_at": "full_universe_second_detail_open.parquet",
}


def get_cache_root(cache_dir: str | Path | None = None) -> Path:
    root = Path(cache_dir) if cache_dir is not None else CACHE_ROOT
    root.mkdir(parents=True, exist_ok=True)
    return root


def build_cache_path(
    cache_dir: str | Path | None,
    category: str,
    *,
    dataset: str,
    parts: list[str],
    suffix: str = ".parquet",
) -> Path:
    safe_dataset = dataset.replace(".", "_").replace("/", "_")
    normalized = [str(part).replace(":", "-").replace("/", "_").replace(" ", "_") for part in parts]
    cache_version = CACHE_VERSION_BY_CATEGORY.get(category, CACHE_VERSION)
    digest = hashlib.sha1("|".join([cache_version, category, dataset, *normalized]).encode("utf-8")).hexdigest()[:12]
    directory = get_cache_root(cache_dir) / category / safe_dataset
    directory.mkdir(parents=True, exist_ok=True)
    filename = "__".join(normalized + [digest]) + suffix
    return directory / filename


def _read_cached_frame(path: Path, *, max_age_seconds: int | None = None) -> pd.DataFrame | None:
    if not path.exists():
        return None
    if max_age_seconds is not None:
        age = (datetime.now(UTC) - datetime.fromtimestamp(path.stat().st_mtime, tz=UTC)).total_seconds()
        if age > max_age_seconds:
            logger.info("Cache expired (%.1f h old, TTL %.1f h): %s", age / 3600, max_age_seconds / 3600, path.name)
            return None
    try:
        return pd.read_parquet(path)
    except Exception:
        logger.warning("Corrupt cache file removed: %s", path, exc_info=True)
        try:
            path.unlink()
        except OSError:
            pass
        return None


def _trade_day_cache_max_age_seconds(trade_day: date, latest_trade_day: date | None) -> int | None:
    if latest_trade_day is None:
        return DATA_CACHE_TTL_SECONDS
    if trade_day >= latest_trade_day:
        return 0
    if trade_day >= latest_trade_day - timedelta(days=1):
        return RECENT_INTRADAY_CACHE_TTL_SECONDS
    return None


def _write_cached_frame(path: Path, frame: pd.DataFrame) -> None:
    _write_parquet_atomic(path, frame)


def _make_atomic_temp_path(path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    fd, temp_name = tempfile.mkstemp(dir=path.parent, prefix=f".{path.name}.", suffix=".tmp")
    os.close(fd)
    return Path(temp_name)


def _replace_atomic(path: Path, write_temp: Callable[[Path], None]) -> None:
    temp_path = _make_atomic_temp_path(path)
    try:
        write_temp(temp_path)
        os.replace(temp_path, path)
    except Exception:
        try:
            temp_path.unlink(missing_ok=True)
        except OSError:
            pass
        raise


def _write_text_atomic(path: Path, content: str, *, encoding: str = "utf-8") -> None:
    def write_temp(temp_path: Path) -> None:
        temp_path.write_text(content, encoding=encoding)

    _replace_atomic(path, write_temp)


def _write_bytes_atomic(path: Path, content: bytes) -> None:
    def write_temp(temp_path: Path) -> None:
        temp_path.write_bytes(content)

    _replace_atomic(path, write_temp)


def _write_parquet_atomic(path: Path, frame: pd.DataFrame) -> None:
    def write_temp(temp_path: Path) -> None:
        frame.to_parquet(temp_path, index=False)

    _replace_atomic(path, write_temp)


def normalize_symbol_for_databento(symbol: str) -> str:
    normalized = str(symbol).strip().upper()
    if not normalized:
        return ""
    normalized = DATABENTO_SYMBOL_ALIASES.get(normalized, normalized)
    # Databento equity symbology does not resolve many preferred/unit/warrant
    # encodings coming from broad exchange directories; filter these upfront.
    if _DATABENTO_INVALID_CHAR_RE.search(normalized):
        return ""
    if normalized.endswith(_DATABENTO_UNIT_OR_WARRANT_SUFFIXES):
        return ""
    if normalized in DATABENTO_UNSUPPORTED_SYMBOLS:
        return ""
    return normalized


def _normalize_symbols(symbols: set[str] | list[str] | tuple[str, ...]) -> list[str]:
    normalized = {
        normalized_symbol
        for symbol in symbols
        if (normalized_symbol := normalize_symbol_for_databento(str(symbol)))
    }
    return sorted(normalized)


def _symbol_scope_token(symbols: set[str] | list[str] | tuple[str, ...]) -> str:
    normalized = _normalize_symbols(symbols)
    digest = hashlib.sha1("|".join(normalized).encode("utf-8")).hexdigest()[:12] if normalized else "empty"
    return f"{len(normalized)}_{digest}"


def _symbol_day_scope_token(symbol_day_scope: pd.DataFrame | None) -> str:
    if symbol_day_scope is None or symbol_day_scope.empty:
        return "all"
    scope = symbol_day_scope.copy()
    scope["trade_date"] = pd.to_datetime(scope["trade_date"], errors="coerce").dt.strftime("%Y-%m-%d")
    scope["symbol"] = scope["symbol"].astype(str).map(normalize_symbol_for_databento)
    scope = scope.dropna(subset=["trade_date", "symbol"])
    if scope.empty:
        return "empty"
    scope = scope.drop_duplicates(subset=["trade_date", "symbol"]).sort_values(["trade_date", "symbol"])
    payload = "|".join(f"{row.trade_date}:{row.symbol}" for row in scope.itertuples(index=False))
    digest = hashlib.sha1(payload.encode("utf-8")).hexdigest()[:12]
    return f"{len(scope)}_{digest}"


def _normalize_symbol_day_scope(symbol_day_scope: pd.DataFrame | None) -> pd.DataFrame:
    if symbol_day_scope is None or symbol_day_scope.empty:
        return pd.DataFrame(columns=["trade_date", "symbol"])
    scope = symbol_day_scope.copy()
    scope["trade_date"] = pd.to_datetime(scope["trade_date"], errors="coerce").dt.date
    scope["symbol"] = scope["symbol"].astype(str).map(normalize_symbol_for_databento)
    scope = scope.dropna(subset=["trade_date", "symbol"])
    scope = scope[scope["symbol"].astype(str).ne("")]
    return scope[["trade_date", "symbol"]].drop_duplicates(subset=["trade_date", "symbol"]).reset_index(drop=True)


def _iter_symbol_batches(
    symbols: set[str] | list[str] | tuple[str, ...],
    *,
    batch_size: int = MAX_SYMBOLS_PER_REQUEST,
) -> list[list[str]]:
    normalized = _normalize_symbols(symbols)
    return [normalized[index:index + batch_size] for index in range(0, len(normalized), batch_size)]


def _empty_intraday_frame() -> pd.DataFrame:
    return pd.DataFrame(
        columns=[
            "trade_date", "symbol", "previous_close", "premarket_price", "market_open_price", "window_start_price",
            "current_price", "current_price_timestamp", "window_high", "window_low", "window_volume", "seconds_in_window", "window_return_pct",
            "window_range_pct", "realized_vol_pct", "has_premarket_data", "prev_close_to_premarket_abs", "prev_close_to_premarket_pct",
            "premarket_to_open_abs", "premarket_to_open_pct", "open_to_current_abs", "open_to_current_pct",
        ]
    )


def _add_transition_columns(frame: pd.DataFrame) -> pd.DataFrame:
    if frame.empty:
        return frame.copy()

    enriched = frame.copy()
    previous_close = pd.to_numeric(enriched.get("previous_close"), errors="coerce")
    premarket_price = pd.to_numeric(enriched.get("premarket_price"), errors="coerce")
    market_open_price = pd.to_numeric(enriched.get("market_open_price"), errors="coerce")
    current_price = pd.to_numeric(enriched.get("current_price"), errors="coerce")

    enriched["prev_close_to_premarket_abs"] = premarket_price - previous_close
    enriched["prev_close_to_premarket_pct"] = np.where(
        (previous_close > 0) & (premarket_price > 0),
        ((premarket_price / previous_close) - 1.0) * 100.0,
        np.nan,
    )

    enriched["premarket_to_open_abs"] = market_open_price - premarket_price
    enriched["premarket_to_open_pct"] = np.where(
        (premarket_price > 0) & (market_open_price > 0),
        ((market_open_price / premarket_price) - 1.0) * 100.0,
        np.nan,
    )

    enriched["open_to_current_abs"] = current_price - market_open_price
    enriched["open_to_current_pct"] = np.where(
        (market_open_price > 0) & (current_price > 0),
        ((current_price / market_open_price) - 1.0) * 100.0,
        np.nan,
    )

    return enriched


def _symbol_support_cache_path(cache_dir: str | Path | None, dataset: str) -> Path:
    return build_cache_path(cache_dir, "symbol_support", dataset=dataset, parts=["support"])


def _read_symbol_support_cache(path: Path) -> dict[str, bool]:
    if not path.exists():
        return {}
    age_seconds = (datetime.now(UTC) - datetime.fromtimestamp(path.stat().st_mtime, tz=UTC)).total_seconds()
    if age_seconds > SYMBOL_SUPPORT_CACHE_TTL_SECONDS:
        logger.info("Symbol support cache expired (%.0f hours old), refreshing.", age_seconds / 3600)
        return {}
    cached = _read_cached_frame(path)
    if cached is None or cached.empty or "symbol" not in cached.columns or "is_supported" not in cached.columns:
        return {}
    frame = cached.copy()
    frame["symbol"] = frame["symbol"].astype(str).str.upper()
    frame["is_supported"] = frame["is_supported"].astype(bool)
    return dict(zip(frame["symbol"], frame["is_supported"], strict=False))


def _write_symbol_support_cache(path: Path, support_map: dict[str, bool]) -> None:
    if not support_map:
        return
    frame = pd.DataFrame(
        {
            "symbol": sorted(support_map),
            "is_supported": [bool(support_map[symbol]) for symbol in sorted(support_map)],
        }
    )
    _write_cached_frame(path, frame)


def _symbols_requiring_support_check(symbols: set[str] | list[str] | tuple[str, ...]) -> list[str]:
    return _normalize_symbols(symbols)


def _extract_unresolved_symbols_from_warning_messages(messages: list[str]) -> set[str]:
    unresolved: set[str] = set()
    for message in messages:
        match = re.search(r"did not resolve:\s*(.+)$", str(message), flags=re.IGNORECASE)
        if not match:
            continue
        for raw_symbol in match.group(1).split(","):
            cleaned = raw_symbol.replace("...", "").strip().upper()
            cleaned = cleaned.strip(" .;:\t\n\r\"'")
            normalized = normalize_symbol_for_databento(cleaned)
            if normalized:
                unresolved.add(normalized)
    return unresolved


def _probe_symbol_support(
    databento_api_key: str,
    *,
    dataset: str,
    symbols: list[str],
) -> dict[str, bool]:
    if not symbols:
        return {}
    client = _make_databento_client(databento_api_key)
    current_utc_day = datetime.now(UTC).date()
    conditions = client.metadata.get_dataset_condition(
        dataset=dataset,
        start_date=(current_utc_day - timedelta(days=SYMBOL_SUPPORT_LOOKBACK_DAYS)).isoformat(),
        end_date=current_utc_day.isoformat(),
    )
    available_days = [
        date.fromisoformat(str(item.get("date")))
        for item in conditions
        if isinstance(item, dict) and item.get("condition") == "available" and item.get("date")
    ]
    available_days = [day for day in available_days if day < current_utc_day]
    if not available_days:
        return {symbol: True for symbol in symbols}
    probe_day = available_days[-1]
    probe_start = datetime.combine(probe_day, time(9, 30), tzinfo=US_EASTERN_TZ).astimezone(UTC).isoformat()
    probe_end = _exclusive_ohlcv_1s_end(
        datetime.combine(probe_day, time(9, 31), tzinfo=US_EASTERN_TZ).astimezone(UTC)
    ).isoformat()
    support_map: dict[str, bool] = {}

    for batch in _iter_symbol_batches(symbols, batch_size=SYMBOL_SUPPORT_CHECK_BATCH_SIZE):
        try:
            with warnings.catch_warnings(record=True) as caught_warnings:
                warnings.simplefilter("always")
                store = _databento_get_range_with_retry(
                    client,
                    context="_probe_symbol_support",
                    dataset=dataset,
                    symbols=batch,
                    schema="ohlcv-1s",
                    start=probe_start,
                    end=probe_end,
                )
                batch_df = store.to_df(count=100_000)
                if isinstance(batch_df, pd.DataFrame):
                    batch_frame = _coerce_timestamp_frame(batch_df)
                else:
                    chunks = list(batch_df)
                    batch_frame = _coerce_timestamp_frame(pd.concat(chunks, ignore_index=False)) if chunks else pd.DataFrame()
        except Exception:
            logger.warning("Symbol support probe failed for batch of %d symbols, assuming supported.", len(batch), exc_info=True)
            for symbol in batch:
                support_map.setdefault(symbol, True)
            continue
        resolved = set()
        if not batch_frame.empty and "symbol" in batch_frame.columns:
            resolved = {
                normalize_symbol_for_databento(symbol)
                for symbol in batch_frame["symbol"].astype(str).tolist()
                if normalize_symbol_for_databento(symbol)
            }
        unresolved = _extract_unresolved_symbols_from_warning_messages([str(item.message) for item in caught_warnings])
        for symbol in batch:
            if symbol in resolved:
                support_map[symbol] = True
            elif symbol in unresolved:
                support_map[symbol] = False
    return support_map


def filter_supported_universe_for_databento(
    databento_api_key: str,
    *,
    dataset: str,
    universe: pd.DataFrame,
    cache_dir: str | Path | None = None,
    use_file_cache: bool = False,
    force_refresh: bool = False,
) -> tuple[pd.DataFrame, list[str]]:
    if universe.empty or "symbol" not in universe.columns:
        return universe.copy(), []
    candidate_symbols = _symbols_requiring_support_check(universe["symbol"].dropna().astype(str).tolist())
    if not candidate_symbols:
        return universe.copy(), []

    cache_path = _symbol_support_cache_path(cache_dir, dataset)
    support_map = _read_symbol_support_cache(cache_path) if use_file_cache and not force_refresh else {}
    missing = [symbol for symbol in candidate_symbols if symbol not in support_map]
    if missing:
        support_map.update(_probe_symbol_support(databento_api_key, dataset=dataset, symbols=missing))
        if use_file_cache:
            _write_symbol_support_cache(cache_path, support_map)

    unsupported = sorted(symbol for symbol in candidate_symbols if support_map.get(symbol) is False)
    if not unsupported:
        return universe.copy(), []
    filtered = universe[~universe["symbol"].astype(str).str.upper().isin(unsupported)].copy().reset_index(drop=True)
    return filtered, unsupported


def _import_databento() -> Any:
    existing_loop_ids = {
        id(obj)
        for obj in gc.get_objects()
        if isinstance(obj, asyncio.AbstractEventLoop)
    }
    import databento as db

    for obj in gc.get_objects():
        if not isinstance(obj, asyncio.AbstractEventLoop):
            continue
        if id(obj) in existing_loop_ids:
            continue
        if obj.is_closed() or obj.is_running():
            continue
        obj.close()

    return db


def _make_databento_client(api_key: str | None = None) -> Any:
    db = _import_databento()
    return db.Historical(api_key or os.getenv("DATABENTO_API_KEY"))


def _get_schema_available_end(client: Any, dataset: str, schema: str) -> pd.Timestamp | None:
    try:
        dataset_range = client.metadata.get_dataset_range(dataset=dataset)
    except Exception:
        logger.debug("metadata.get_dataset_range failed for %s/%s; clamping disabled", dataset, schema, exc_info=True)
        return None
    if not isinstance(dataset_range, dict):
        return None
    schema_ranges = dataset_range.get("schema")
    if isinstance(schema_ranges, dict):
        schema_info = schema_ranges.get(schema)
        if isinstance(schema_info, dict):
            end_value = schema_info.get("end")
            if end_value:
                return pd.Timestamp(end_value, tz=UTC)
    end_value = dataset_range.get("end")
    if not end_value:
        return None
    return pd.Timestamp(end_value, tz=UTC)


def _clamp_request_end(requested_end: pd.Timestamp, available_end: pd.Timestamp | None) -> pd.Timestamp:
    if available_end is None:
        return requested_end
    return min(requested_end, available_end)


def _exclusive_ohlcv_1s_end(logical_end: datetime | pd.Timestamp) -> pd.Timestamp:
    end_timestamp = pd.Timestamp(logical_end)
    if end_timestamp.tzinfo is None:
        end_timestamp = end_timestamp.tz_localize(UTC)
    return end_timestamp + pd.Timedelta(seconds=1)


def _daily_request_end_exclusive(last_trading_day: date, available_end: pd.Timestamp | None) -> date:
    requested_end = pd.Timestamp(last_trading_day + timedelta(days=1), tz=UTC)
    clamped_end = _clamp_request_end(requested_end, available_end)
    # The clamped timestamp may fall *within* the last trading day (e.g.
    # available_end = 2024-03-15 20:00 UTC).  Converting that directly to
    # a date would give 2024-03-15 as the *exclusive* end, effectively
    # dropping that day from the request.  When the timestamp has a non-zero
    # time component (i.e. falls during a day, not at midnight), add +1 day
    # so that calendar day is still included in the request.
    clamped_dt = pd.Timestamp(clamped_end).to_pydatetime()
    end_date = date(clamped_dt.year, clamped_dt.month, clamped_dt.day)
    if clamped_dt.hour or clamped_dt.minute or clamped_dt.second or clamped_dt.microsecond:
        end_date += timedelta(days=1)
    return end_date


def list_accessible_datasets(databento_api_key: str | None = None) -> list[str]:
    client = _make_databento_client(databento_api_key)
    datasets = client.metadata.list_datasets()
    return sorted({str(dataset) for dataset in datasets if dataset})


def choose_default_dataset(
    available_datasets: list[str],
    requested_dataset: str | None = None,
) -> str:
    normalized = [str(dataset).strip() for dataset in available_datasets if str(dataset).strip()]
    available_lookup = {dataset.upper(): dataset for dataset in normalized}
    requested_normalized = str(requested_dataset).strip() if requested_dataset else None
    if requested_normalized:
        matched_requested = available_lookup.get(requested_normalized.upper())
        if matched_requested:
            return matched_requested
        logger.warning("Requested dataset %r not in available datasets %r, falling back.", requested_dataset, normalized)
    for dataset in PREFERRED_DATABENTO_DATASETS:
        matched_preferred = available_lookup.get(dataset.upper())
        if matched_preferred:
            return matched_preferred
    if normalized:
        return normalized[0]
    return requested_normalized or PREFERRED_DATABENTO_DATASETS[0]


def _safe_float(value: Any, default: float | None = None) -> float | None:
    try:
        if value is None or value == "":
            return default
        result = float(value)
    except (TypeError, ValueError):
        return default
    if math.isnan(result) or math.isinf(result):
        return default
    return result


def _coerce_timestamp_frame(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df.copy()
    frame = df.copy()
    if isinstance(frame.index, pd.DatetimeIndex):
        frame = frame.reset_index()
        idx_name = frame.columns[0]
        frame = frame.rename(columns={idx_name: "ts"})
    elif "ts_event" in frame.columns:
        frame = frame.rename(columns={"ts_event": "ts"})
    elif "ts_recv" in frame.columns:
        frame = frame.rename(columns={"ts_recv": "ts"})
    elif "index" in frame.columns:
        frame = frame.rename(columns={"index": "ts"})
    else:
        raise ValueError("No timestamp column found in Databento frame")
    frame["ts"] = pd.to_datetime(frame["ts"], utc=True)
    return frame


def _redact_sensitive_error_text(text: str) -> str:
    redacted = str(text)
    for pattern in _API_KEY_REDACTION_PATTERNS:
        redacted = pattern.sub(r"\1***", redacted)
    return redacted


def _warn_with_redacted_exception(message: str, exc: BaseException, *, include_traceback: bool = False) -> None:
    logger.warning("%s: %s", message, _redact_sensitive_error_text(str(exc)), exc_info=include_traceback)


def _is_retryable_databento_get_range_error(exc: BaseException) -> bool:
    message = _redact_sensitive_error_text(str(exc)).lower()
    if not message:
        return False
    retryable_fragments = (
        "read timed out",
        "timed out",
        "too many requests",
        "429",
        "503",
        "504",
        "service unavailable",
        "gateway timeout",
        "connection reset",
        "temporarily unavailable",
    )
    return any(fragment in message for fragment in retryable_fragments)


def _databento_get_range_with_retry(client: Any, *, context: str, **kwargs: Any) -> Any:
    last_exc: BaseException | None = None
    for attempt in range(1, DATABENTO_GET_RANGE_MAX_ATTEMPTS + 1):
        try:
            return client.timeseries.get_range(**kwargs)
        except Exception as exc:
            last_exc = exc
            if attempt >= DATABENTO_GET_RANGE_MAX_ATTEMPTS or not _is_retryable_databento_get_range_error(exc):
                raise
            wait_seconds = float(2 ** (attempt - 1))
            logger.warning(
                "%s: transient Databento get_range failure (%s). Retrying in %.0fs (%d/%d).",
                context,
                _redact_sensitive_error_text(str(exc)),
                wait_seconds,
                attempt,
                DATABENTO_GET_RANGE_MAX_ATTEMPTS,
            )
            time_module.sleep(wait_seconds)
    if last_exc is not None:
        raise last_exc
    raise RuntimeError(f"{context}: Databento get_range retry loop exited unexpectedly")


def _validate_frame_columns(frame: pd.DataFrame, *, required: set[str], context: str) -> pd.DataFrame:
    missing = sorted(required - set(frame.columns))
    if missing:
        raise ValueError(f"{context} missing required columns: {', '.join(missing)}")
    return frame


def _store_to_frame(store: Any, *, count: int | None = None, context: str) -> pd.DataFrame:
    payload = store.to_df(count=count) if count is not None else store.to_df()
    if isinstance(payload, pd.DataFrame):
        return _coerce_timestamp_frame(payload)
    chunks = list(payload)
    if not chunks:
        return pd.DataFrame()
    concatenated = pd.concat(chunks, ignore_index=False)
    return _coerce_timestamp_frame(concatenated)


def resolve_display_timezone(display_timezone: str) -> tzinfo:
    tz = SUPPORTED_DISPLAY_TZ.get(display_timezone)
    if tz is None:
        raise ValueError(f"Unsupported display timezone: {display_timezone}")
    return tz


def compute_market_relative_window(
    trade_date: date,
    display_timezone: str,
    *,
    pre_open_minutes: int = _DEFAULT_INTRADAY_PRE_OPEN_MINUTES,
    post_open_minutes: int = _DEFAULT_INTRADAY_POST_OPEN_MINUTES,
    post_open_seconds: int | None = None,
) -> tuple[time, time]:
    if pre_open_minutes < 0:
        raise ValueError(f"pre_open_minutes must be >= 0, got {pre_open_minutes}")
    if post_open_minutes < 0:
        raise ValueError(f"post_open_minutes must be >= 0, got {post_open_minutes}")
    if post_open_seconds is not None and post_open_seconds < 0:
        raise ValueError(f"post_open_seconds must be >= 0, got {post_open_seconds}")
    tz = resolve_display_timezone(display_timezone)
    regular_open_local = datetime.combine(trade_date, time(9, 30), tzinfo=US_EASTERN_TZ).astimezone(tz)
    start_local = regular_open_local - timedelta(minutes=pre_open_minutes)
    if post_open_seconds is not None:
        end_local = regular_open_local + timedelta(seconds=post_open_seconds)
    else:
        end_local = regular_open_local + timedelta(minutes=post_open_minutes)
    return start_local.time(), end_local.time()


def _resolve_window_for_date(
    trade_date: date,
    display_timezone: str,
    window_start: time | None,
    window_end: time | None,
    *,
    default_pre_open_minutes: int = _DEFAULT_INTRADAY_PRE_OPEN_MINUTES,
    default_post_open_minutes: int = _DEFAULT_INTRADAY_POST_OPEN_MINUTES,
    default_post_open_seconds: int | None = None,
) -> tuple[time, time]:
    if window_start is not None and window_end is not None:
        return window_start, window_end
    return compute_market_relative_window(
        trade_date,
        display_timezone,
        pre_open_minutes=default_pre_open_minutes,
        post_open_minutes=default_post_open_minutes,
        post_open_seconds=default_post_open_seconds,
    )


def build_window_definition(
    trade_date: date,
    *,
    display_timezone: str,
    window_start: time,
    window_end: time,
    premarket_anchor_et: time,
) -> WindowDefinition:
    tz = resolve_display_timezone(display_timezone)
    local_start = datetime.combine(trade_date, window_start, tzinfo=tz)
    local_end = datetime.combine(trade_date, window_end, tzinfo=tz)
    if local_end <= local_start:
        raise ValueError("Window end must be after window start")
    regular_open_utc = datetime.combine(trade_date, time(9, 30), tzinfo=US_EASTERN_TZ).astimezone(UTC)
    premarket_anchor_utc = datetime.combine(trade_date, premarket_anchor_et, tzinfo=US_EASTERN_TZ).astimezone(UTC)
    fetch_start_utc = min(local_start.astimezone(UTC), premarket_anchor_utc)
    fetch_end_utc = local_end.astimezone(UTC)
    return WindowDefinition(
        trade_date=trade_date,
        display_timezone=display_timezone,
        window_start_local=local_start,
        window_end_local=local_end,
        fetch_start_utc=fetch_start_utc,
        fetch_end_utc=fetch_end_utc,
        regular_open_utc=regular_open_utc,
        premarket_anchor_utc=premarket_anchor_utc,
    )


def fetch_us_equity_universe(
    fmp_api_key: str = "",
    *,
    min_market_cap: float | None = None,
    exchanges: str = "NASDAQ,NYSE,AMEX",
) -> pd.DataFrame:
    frame, _metadata = fetch_us_equity_universe_with_metadata(
        fmp_api_key,
        min_market_cap=min_market_cap,
        exchanges=exchanges,
    )
    return frame


def fetch_us_equity_universe_with_metadata(
    fmp_api_key: str = "",
    *,
    min_market_cap: float | None = None,
    exchanges: str = "NASDAQ,NYSE,AMEX",
) -> tuple[pd.DataFrame, dict[str, Any]]:
    requested_min_market_cap = float(min_market_cap) if min_market_cap is not None else None
    if min_market_cap is not None and fmp_api_key:
        client = FMPClient(fmp_api_key)
        return (
            _fetch_us_equity_universe_via_screener(
                client,
                min_market_cap=min_market_cap,
                exchanges=exchanges,
            ),
            {
                "source": "fmp_company_screener",
                "fallback_source": "nasdaq_trader_symbol_directory",
                "scope_definition": "US equity universe returned by the FMP company screener for the requested exchanges, filtered by the requested market-cap floor before Databento symbol support is applied.",
                "min_market_cap_requested": requested_min_market_cap,
                "min_market_cap_effective": requested_min_market_cap,
                "min_market_cap_applied": True,
                "selection_reason": "market_cap_filter_requested",
            },
        )
    if min_market_cap is not None and not fmp_api_key:
        logger.warning(
            "min_market_cap=%.0f requested but no FMP API key provided; "
            "returning unfiltered Nasdaq Trader universe instead.",
            min_market_cap,
        )

    official = _fetch_us_equity_universe_via_nasdaq_trader(exchanges=exchanges)
    if not official.empty:
        return (
            official,
            {
                "source": "nasdaq_trader_symbol_directory",
                "fallback_source": "fmp_company_screener" if fmp_api_key else None,
                "scope_definition": "Listed non-ETF, non-test issues from Nasdaq Trader symbol directories for the requested US exchanges; Databento support is applied afterward.",
                "min_market_cap_requested": requested_min_market_cap,
                "min_market_cap_effective": None,
                "min_market_cap_applied": False,
                "selection_reason": "official_directory",
            },
        )

    if fmp_api_key:
        client = FMPClient(fmp_api_key)
        return (
            _fetch_us_equity_universe_via_screener(
                client,
                min_market_cap=min_market_cap,
                exchanges=exchanges,
            ),
            {
                "source": "fmp_company_screener",
                "fallback_source": "nasdaq_trader_symbol_directory",
                "scope_definition": "US equity universe returned by the FMP company screener after the Nasdaq Trader directory fetch failed; Databento support is applied afterward.",
                "min_market_cap_requested": requested_min_market_cap,
                "min_market_cap_effective": requested_min_market_cap,
                "min_market_cap_applied": requested_min_market_cap is not None,
                "selection_reason": "official_directory_failed",
            },
        )

    logger.warning("Nasdaq Trader directory fetch failed and no FMP API key provided; returning empty universe.")
    return (
        _empty_universe_frame(),
        {
            "source": "empty",
            "fallback_source": None,
            "scope_definition": "No universe rows available because the Nasdaq Trader directory fetch failed and no FMP API key was provided for fallback.",
            "min_market_cap_requested": requested_min_market_cap,
            "min_market_cap_effective": None,
            "min_market_cap_applied": False,
            "selection_reason": "no_available_source",
        },
    )


def _empty_universe_frame() -> pd.DataFrame:
    return pd.DataFrame(columns=UNIVERSE_COLUMNS)


def _fetch_us_equity_universe_via_screener(
    client: FMPClient,
    *,
    min_market_cap: float | None = None,
    exchanges: str = "NASDAQ,NYSE,AMEX",
) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    page = 0
    page_size = 1000
    while True:
        try:
            batch = client.get_company_screener(
                country="US",
                market_cap_more_than=min_market_cap,
                exchange=exchanges,
                is_etf=False,
                is_fund=False,
                limit=page_size,
                page=page,
            )
        except Exception:
            logger.warning("FMP company screener request failed on page %d.", page, exc_info=True)
            break
        if not batch:
            break
        rows.extend(batch)
        if len(batch) < page_size:
            break
        page += 1
        if page > 50:
            break
    if not rows:
        return _empty_universe_frame()
    frame = pd.DataFrame(rows)
    if frame.empty:
        return _empty_universe_frame()
    out = pd.DataFrame(
        {
            "symbol": frame.get("symbol", "").astype(str).map(normalize_symbol_for_databento),
            "company_name": frame.get("companyName", frame.get("name", "")),
            "exchange": frame.get("exchangeShortName", frame.get("exchange", "")),
            "sector": frame.get("sector", ""),
            "industry": frame.get("industry", ""),
            "market_cap": pd.to_numeric(frame.get("marketCap"), errors="coerce"),
        }
    )
    out = out[out["symbol"].astype(str).str.len() > 0].copy()
    out = out[out["symbol"].ne("")].drop_duplicates(subset=["symbol"]).reset_index(drop=True)
    return out


def _normalize_requested_exchange_codes(exchanges: str) -> set[str]:
    requested: set[str] = set()
    for item in str(exchanges or "").split(","):
        token = item.strip().upper()
        if not token:
            continue
        requested.add(NASDAQ_TRADER_EXCHANGE_CODE_MAP.get(token, token))
    return requested or {"Q", "N", "A"}


def _download_nasdaq_trader_text(url: str) -> str:
    import ssl
    _ssl_ctx = ssl.create_default_context(cafile=certifi.where())
    request = Request(url, headers={"User-Agent": "Mozilla/5.0"})
    last_error: Exception | None = None
    for attempt in range(3):
        try:
            with urlopen(request, timeout=30, context=_ssl_ctx) as response:
                payload = response.read()
            return bytes(payload).decode("utf-8", errors="replace")
        except Exception as exc:  # pragma: no cover - deterministic in tests via monkeypatch
            last_error = exc
            if attempt < 2:
                time_module.sleep(0.5 * (2 ** attempt))
    assert last_error is not None
    raise last_error


def _parse_nasdaq_trader_directory(
    text: str,
    *,
    symbol_column: str,
    security_name_column: str,
    exchange_column: str,
    allowed_exchange_codes: set[str],
) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for row in pd.read_csv(BytesIO(text.encode("utf-8")), sep="|", dtype=str).fillna("").to_dict(orient="records"):
        first_value = next(iter(row.values()), "")
        if isinstance(first_value, str) and first_value.startswith("File Creation Time:"):
            continue
        symbol = normalize_symbol_for_databento(str(row.get(symbol_column, "")).strip())
        exchange_code = str(row.get(exchange_column, "")).strip().upper()
        if not symbol or exchange_code not in allowed_exchange_codes:
            continue
        if str(row.get("ETF", "")).strip().upper() == "Y":
            continue
        if str(row.get("Test Issue", "")).strip().upper() == "Y":
            continue
        rows.append(
            {
                "symbol": symbol,
                "company_name": str(row.get(security_name_column, "")).strip(),
                "exchange": NASDAQ_TRADER_EXCHANGE_NAME_MAP.get(exchange_code, exchange_code),
                "sector": "",
                "industry": "",
                "market_cap": np.nan,
            }
        )
    if not rows:
        return _empty_universe_frame()
    return pd.DataFrame(rows, columns=UNIVERSE_COLUMNS)


def _fetch_us_equity_universe_via_nasdaq_trader(*, exchanges: str = "NASDAQ,NYSE,AMEX") -> pd.DataFrame:
    allowed_exchange_codes = _normalize_requested_exchange_codes(exchanges)
    frames: list[pd.DataFrame] = []
    for url, symbol_column, security_name_column, exchange_column in NASDAQ_TRADER_DIRECTORY_SPECS:
        try:
            text = _download_nasdaq_trader_text(url)
        except Exception:
            logger.warning("Failed to download Nasdaq Trader directory from %s", url, exc_info=True)
            continue
        frame = _parse_nasdaq_trader_directory(
            text,
            symbol_column=symbol_column,
            security_name_column=security_name_column,
            exchange_column=exchange_column,
            allowed_exchange_codes=allowed_exchange_codes,
        )
        if not frame.empty:
            frames.append(frame)
    if not frames:
        return _empty_universe_frame()
    universe = pd.concat(frames, ignore_index=True)
    universe = universe.drop_duplicates(subset=["symbol"]).reset_index(drop=True)
    return universe


def list_recent_trading_days(
    databento_api_key: str,
    *,
    dataset: str,
    lookback_days: int,
    end_date: date | None = None,
) -> list[date]:
    client = _make_databento_client(databento_api_key)
    current_market_day = datetime.now(US_EASTERN_TZ).date()
    anchor = end_date or current_market_day
    start_date = anchor - timedelta(days=max(lookback_days * 4, 90))
    conditions = client.metadata.get_dataset_condition(
        dataset=dataset,
        start_date=start_date.isoformat(),
        end_date=anchor.isoformat(),
    )
    days = [
        date.fromisoformat(str(item.get("date")))
        for item in conditions
        if isinstance(item, dict) and item.get("condition") == "available"
    ]
    if end_date is None:
        days = [day for day in days if day < current_market_day]
    return days[-lookback_days:]


def _deduplicate_daily_symbol_rows(frame: pd.DataFrame) -> pd.DataFrame:
    if frame.empty or not {"trade_date", "symbol"}.issubset(frame.columns):
        return frame

    duplicate_mask = frame.duplicated(subset=["trade_date", "symbol"], keep=False)
    if not duplicate_mask.any():
        return frame

    duplicate_keys = int(frame.loc[duplicate_mask, ["trade_date", "symbol"]].drop_duplicates().shape[0])
    logger.warning(
        "load_daily_bars: consolidating %d duplicate trade_date-symbol rows by aggregating OHLCV.",
        duplicate_keys,
    )

    ordered = frame.reset_index(drop=True).copy()
    ordered["_row_order"] = np.arange(len(ordered))
    ordered = ordered.sort_values(["trade_date", "symbol", "_row_order"]).reset_index(drop=True)

    aggregations: dict[str, str] = {}
    if "open" in ordered.columns:
        aggregations["open"] = "first"
    if "high" in ordered.columns:
        aggregations["high"] = "max"
    if "low" in ordered.columns:
        aggregations["low"] = "min"
    if "close" in ordered.columns:
        aggregations["close"] = "last"
    if "volume" in ordered.columns:
        aggregations["volume"] = "sum"

    passthrough_columns = [
        column
        for column in ordered.columns
        if column not in {"trade_date", "symbol", "_row_order", *aggregations.keys()}
    ]
    for column in passthrough_columns:
        aggregations[column] = "last"

    return ordered.groupby(["trade_date", "symbol"], as_index=False, sort=True).agg(aggregations).reset_index(drop=True)


def _collapse_duplicate_symbol_seconds(frame: pd.DataFrame, *, context: str) -> pd.DataFrame:
    if frame.empty or not {"symbol", "ts"}.issubset(frame.columns):
        return frame

    duplicate_mask = frame.duplicated(subset=["symbol", "ts"], keep=False)
    if not duplicate_mask.any():
        return frame

    duplicate_keys = int(frame.loc[duplicate_mask, ["symbol", "ts"]].drop_duplicates().shape[0])
    logger.warning(
        "%s: consolidating %d duplicate symbol-second rows by aggregating OHLCV.",
        context,
        duplicate_keys,
    )

    ordered = frame.sort_values(["symbol", "ts"]).reset_index(drop=True)
    aggregations: dict[str, str] = {}
    if "open" in ordered.columns:
        aggregations["open"] = "first"
    if "high" in ordered.columns:
        aggregations["high"] = "max"
    if "low" in ordered.columns:
        aggregations["low"] = "min"
    if "close" in ordered.columns:
        aggregations["close"] = "last"
    if "volume" in ordered.columns:
        aggregations["volume"] = "sum"
    for column in ["trade_count", "count", "n_trades", "num_trades"]:
        if column in ordered.columns:
            aggregations[column] = "sum"

    collapsed = ordered.groupby(["symbol", "ts"], sort=False, as_index=False).agg(aggregations)
    return collapsed.reset_index(drop=True)


def load_daily_bars(
    databento_api_key: str,
    *,
    dataset: str,
    trading_days: list[date],
    universe_symbols: set[str],
    cache_dir: str | Path | None = None,
    use_file_cache: bool = False,
    force_refresh: bool = False,
) -> pd.DataFrame:
    if not trading_days:
        return pd.DataFrame(columns=["trade_date", "symbol", "open", "high", "low", "close", "volume", "previous_close"])
    symbol_scope = _symbol_scope_token(universe_symbols)
    start_date = trading_days[0] - timedelta(days=14)
    end_date = trading_days[-1]
    cache_path = build_cache_path(
        cache_dir,
        "daily_bars",
        dataset=dataset,
        parts=[start_date.strftime("%Y%m%d"), end_date.strftime("%Y%m%d"), symbol_scope],
    )
    frame: pd.DataFrame | None = None
    if use_file_cache and not force_refresh:
        frame = _read_cached_frame(cache_path, max_age_seconds=DATA_CACHE_TTL_SECONDS)
    if frame is None:
        client = _make_databento_client(databento_api_key)
        schema_end = _get_schema_available_end(client, dataset, "ohlcv-1d")
        end_date = _daily_request_end_exclusive(end_date, schema_end)
        if end_date <= start_date:
            return pd.DataFrame(columns=["trade_date", "symbol", "open", "high", "low", "close", "volume", "previous_close"])
        frames: list[pd.DataFrame] = []
        for symbols_batch in _iter_symbol_batches(universe_symbols):
            try:
                store = _databento_get_range_with_retry(
                    client,
                    context="load_daily_bars",
                    dataset=dataset,
                    symbols=symbols_batch,
                    schema="ohlcv-1d",
                    start=start_date.isoformat(),
                    end=end_date.isoformat(),
                )
                batch_frame = _store_to_frame(store, context="load_daily_bars")
                if not batch_frame.empty:
                    _validate_frame_columns(batch_frame, required={"symbol", "open", "high", "low", "close", "volume"}, context="load_daily_bars")
                if not batch_frame.empty:
                    frames.append(batch_frame)
            except Exception as exc:
                _warn_with_redacted_exception(
                    f"Failed to fetch daily bars for batch of {len(symbols_batch)} symbols",
                    exc,
                    include_traceback=True,
                )
        frame = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()
        if use_file_cache and not frame.empty:
            _write_cached_frame(cache_path, frame)
    if frame.empty:
        return pd.DataFrame(columns=["trade_date", "symbol", "open", "high", "low", "close", "volume", "previous_close"])
    normalized_universe_symbols = {
        normalized
        for symbol in universe_symbols
        if (normalized := normalize_symbol_for_databento(symbol)) is not None
    }
    frame["symbol"] = frame.get("symbol", "").map(normalize_symbol_for_databento)
    frame = frame[frame["symbol"].isin(normalized_universe_symbols)].copy()
    frame["trade_date"] = frame["ts"].dt.date
    keep_cols = [col for col in ["trade_date", "symbol", "open", "high", "low", "close", "volume"] if col in frame.columns]
    frame = frame[keep_cols].copy()
    for col in ["open", "high", "low", "close", "volume"]:
        if col in frame.columns:
            frame[col] = pd.to_numeric(frame[col], errors="coerce")
    frame = _deduplicate_daily_symbol_rows(frame)
    frame = frame.sort_values(["symbol", "trade_date"]).reset_index(drop=True)
    frame["previous_close"] = frame.groupby("symbol")["close"].shift(1)
    frame = frame[frame["trade_date"].isin(trading_days)].reset_index(drop=True)
    return frame


def _update_state_from_chunk(
    chunk: pd.DataFrame,
    *,
    window: WindowDefinition,
    universe_symbols: set[str] | None,
    states: dict[str, SymbolDayState],
) -> None:
    if chunk.empty:
        return
    frame = _coerce_timestamp_frame(chunk)
    if "symbol" not in frame.columns:
        return
    frame["symbol"] = frame["symbol"].astype(str).str.upper()
    if universe_symbols is not None:
        frame = frame[frame["symbol"].isin(universe_symbols)].copy()
    if frame.empty:
        return
    frame = frame.sort_values("ts")

    premarket = frame[
        (frame["ts"] >= pd.Timestamp(window.premarket_anchor_utc))
        & (frame["ts"] < pd.Timestamp(window.regular_open_utc))
    ]
    if not premarket.empty:
        latest_premarket = premarket.groupby("symbol").tail(1)
        for row in latest_premarket.itertuples(index=False):
            close_px = _safe_float(getattr(row, "close", None))
            if close_px is None or close_px <= 0:
                continue
            state = states.setdefault(row.symbol, SymbolDayState(symbol=row.symbol, trade_date=window.trade_date))
            state.premarket_price = close_px

    regular = frame[
        (frame["ts"] >= pd.Timestamp(window.regular_open_utc))
        & (frame["ts"] <= pd.Timestamp(window.fetch_end_utc))
    ]
    if not regular.empty:
        opening_rows = regular.groupby("symbol").head(1)
        for row in opening_rows.itertuples(index=False):
            open_px = _safe_float(getattr(row, "open", None)) or _safe_float(getattr(row, "close", None))
            if open_px is None or open_px <= 0:
                continue
            state = states.setdefault(row.symbol, SymbolDayState(symbol=row.symbol, trade_date=window.trade_date))
            if state.market_open_price is None:
                state.market_open_price = open_px

    in_window = frame[
        (frame["ts"] >= pd.Timestamp(window.window_start_local.astimezone(UTC)))
        & (frame["ts"] <= pd.Timestamp(window.window_end_local.astimezone(UTC)))
    ].copy()
    if in_window.empty:
        return

    grouped = in_window.groupby("symbol", sort=False)
    summary = grouped.agg(
        first_window_open=("open", "first"),
        last_window_close=("close", "last"),
        last_window_timestamp=("ts", "last"),
        window_high=("high", "max"),
        window_low=("low", "min"),
        window_volume=("volume", "sum"),
        second_count=("close", "size"),
    )
    returns = grouped["close"].apply(lambda s: float(np.square(np.log(s / s.shift(1))).sum(skipna=True)))
    first_close = grouped["close"].first()
    last_close = grouped["close"].last()

    for symbol, row in summary.iterrows():
        state = states.setdefault(symbol, SymbolDayState(symbol=symbol, trade_date=window.trade_date))
        first_open = _safe_float(row.get("first_window_open"))
        if state.first_window_open is None and first_open is not None and first_open > 0:
            state.first_window_open = first_open

        last_window_close = _safe_float(row.get("last_window_close"))
        if last_window_close is not None and last_window_close > 0:
            state.last_window_close = last_window_close
        last_window_timestamp = pd.to_datetime(row.get("last_window_timestamp"), errors="coerce", utc=True)
        if pd.notna(last_window_timestamp):
            state.last_window_timestamp = last_window_timestamp

        high_val = _safe_float(row.get("window_high"))
        low_val = _safe_float(row.get("window_low"))
        if high_val is not None and high_val > 0:
            state.window_high = high_val if state.window_high is None else max(state.window_high, high_val)
        if low_val is not None and low_val > 0:
            state.window_low = low_val if state.window_low is None else min(state.window_low, low_val)

        state.window_volume += float(row.get("window_volume") or 0.0)
        state.second_count += int(row.get("second_count") or 0)

        chunk_var = float(returns.get(symbol, 0.0) or 0.0)
        initial_close = _safe_float(first_close.get(symbol))
        if state._last_window_close_for_rv is not None and initial_close is not None and initial_close > 0:
            state.realized_var += math.log(initial_close / state._last_window_close_for_rv) ** 2
        state.realized_var += chunk_var
        trailing_close = _safe_float(last_close.get(symbol))
        if trailing_close is not None and trailing_close > 0:
            state._last_window_close_for_rv = trailing_close


def summarize_symbol_day(
    state: SymbolDayState,
    *,
    previous_close: float | None,
) -> dict[str, Any]:
    first_open = _safe_float(state.first_window_open)
    last_close = _safe_float(state.last_window_close)
    high_val = _safe_float(state.window_high)
    low_val = _safe_float(state.window_low)
    prev_close = _safe_float(previous_close)
    premarket = _safe_float(state.premarket_price)
    market_open = _safe_float(state.market_open_price)

    window_return_pct = None
    window_range_pct = None
    if first_open and first_open > 0 and last_close and last_close > 0:
        window_return_pct = ((last_close / first_open) - 1.0) * 100.0
    if first_open and first_open > 0 and high_val and low_val and high_val > 0 and low_val > 0:
        window_range_pct = ((high_val - low_val) / first_open) * 100.0

    prev_to_premarket_pct = None
    prev_to_premarket_abs = None
    if prev_close and prev_close > 0 and premarket and premarket > 0:
        prev_to_premarket_abs = premarket - prev_close
        prev_to_premarket_pct = ((premarket / prev_close) - 1.0) * 100.0

    premarket_to_open_pct = None
    premarket_to_open_abs = None
    if premarket and premarket > 0 and market_open and market_open > 0:
        premarket_to_open_abs = market_open - premarket
        premarket_to_open_pct = ((market_open / premarket) - 1.0) * 100.0

    open_to_current_pct = None
    open_to_current_abs = None
    if market_open and market_open > 0 and last_close and last_close > 0:
        open_to_current_abs = last_close - market_open
        open_to_current_pct = ((last_close / market_open) - 1.0) * 100.0

    return {
        "trade_date": state.trade_date,
        "symbol": state.symbol,
        "previous_close": prev_close,
        "premarket_price": premarket,
        "has_premarket_data": bool(premarket and premarket > 0),
        "market_open_price": market_open,
        "window_start_price": first_open,
        "current_price": last_close,
        "current_price_timestamp": state.last_window_timestamp,
        "window_high": high_val,
        "window_low": low_val,
        "window_volume": state.window_volume,
        "seconds_in_window": state.second_count,
        "window_return_pct": window_return_pct,
        "window_range_pct": window_range_pct,
        "realized_vol_pct": math.sqrt(state.realized_var) * 100.0 if state.realized_var > 0 else None,
        "prev_close_to_premarket_abs": prev_to_premarket_abs,
        "prev_close_to_premarket_pct": prev_to_premarket_pct,
        "premarket_to_open_abs": premarket_to_open_abs,
        "premarket_to_open_pct": premarket_to_open_pct,
        "open_to_current_abs": open_to_current_abs,
        "open_to_current_pct": open_to_current_pct,
    }


def run_intraday_screen(
    databento_api_key: str,
    *,
    dataset: str,
    trading_days: list[date],
    universe_symbols: set[str],
    daily_bars: pd.DataFrame,
    display_timezone: str = DEFAULT_DISPLAY_TZ,
    window_start: time | None = None,
    window_end: time | None = None,
    premarket_anchor_et: time = time(8, 0),
    cache_dir: str | Path | None = None,
    use_file_cache: bool = False,
    force_refresh: bool = False,
) -> pd.DataFrame:
    client = _make_databento_client(databento_api_key)
    available_end_1s = _get_schema_available_end(client, dataset, "ohlcv-1s")
    symbol_scope = _symbol_scope_token(universe_symbols)
    runtime_unsupported_symbols: set[str] = set()
    prev_close_lookup = {
        (row.trade_date, row.symbol): _safe_float(row.previous_close)
        for row in daily_bars.itertuples(index=False)
    }
    results: list[dict[str, Any]] = []
    latest_trade_day = max(trading_days) if trading_days else None

    for trade_day in trading_days:
        day_ws, day_we = _resolve_window_for_date(trade_day, display_timezone, window_start, window_end)
        window = build_window_definition(
            trade_day,
            display_timezone=display_timezone,
            window_start=day_ws,
            window_end=day_we,
            premarket_anchor_et=premarket_anchor_et,
        )
        cache_path = build_cache_path(
            cache_dir,
            "intraday_summary",
            dataset=dataset,
            parts=[
                trade_day.isoformat(),
                display_timezone,
                day_ws.strftime("%H%M%S"),
                day_we.strftime("%H%M%S"),
                premarket_anchor_et.strftime("%H%M%S"),
                symbol_scope,
            ],
        )
        day_frame: pd.DataFrame | None = None
        if use_file_cache and not force_refresh:
            day_frame = _read_cached_frame(
                cache_path,
                max_age_seconds=_trade_day_cache_max_age_seconds(trade_day, latest_trade_day),
            )
        if day_frame is None:
            states: dict[str, SymbolDayState] = {}
            active_symbols = set(universe_symbols) - runtime_unsupported_symbols
            for symbols_batch in _iter_symbol_batches(active_symbols):
                try:
                    with warnings.catch_warnings(record=True) as caught_warnings:
                        warnings.simplefilter("always")
                        clamped_end_1s = _clamp_request_end(
                            _exclusive_ohlcv_1s_end(window.fetch_end_utc), available_end_1s,
                        )
                        if clamped_end_1s <= pd.Timestamp(window.fetch_start_utc):
                            continue
                        store = _databento_get_range_with_retry(
                            client,
                            context="run_intraday_screen",
                            dataset=dataset,
                            symbols=symbols_batch,
                            schema="ohlcv-1s",
                            start=window.fetch_start_utc.isoformat(),
                            end=clamped_end_1s.isoformat(),
                        )
                        iterator = store.to_df(count=250_000)
                        if isinstance(iterator, pd.DataFrame):
                            _update_state_from_chunk(iterator, window=window, universe_symbols=None, states=states)
                        else:
                            for chunk in iterator:
                                _update_state_from_chunk(chunk, window=window, universe_symbols=None, states=states)
                    runtime_unsupported_symbols.update(
                        _extract_unresolved_symbols_from_warning_messages([str(item.message) for item in caught_warnings])
                    )
                except Exception as exc:
                    _warn_with_redacted_exception(
                        f"Intraday fetch failed for batch on {trade_day}, skipping batch",
                        exc,
                        include_traceback=True,
                    )
            day_rows = [summarize_symbol_day(state, previous_close=None) for state in states.values()]
            day_frame = pd.DataFrame(day_rows) if day_rows else _empty_intraday_frame()
            if use_file_cache and not day_frame.empty:
                _write_cached_frame(cache_path, day_frame)
        if day_frame.empty:
            continue
        filtered = day_frame[day_frame["symbol"].isin(universe_symbols)].copy()
        if filtered.empty:
            continue
        filtered["previous_close"] = filtered.apply(
            lambda row: prev_close_lookup.get((pd.Timestamp(row["trade_date"]).date(), str(row["symbol"]).upper())),
            axis=1,
        )
        filtered = _add_transition_columns(filtered)
        results.extend(filtered.to_dict(orient="records"))

    if not results:
        return _empty_intraday_frame()
    return pd.DataFrame(results)


def rank_top_fraction_per_day(
    frame: pd.DataFrame,
    *,
    ranking_metric: str,
    top_fraction: float,
) -> pd.DataFrame:
    if not (0 < top_fraction <= 1):
        raise ValueError(f"top_fraction must be in (0, 1], got {top_fraction}")
    if frame.empty:
        return frame.copy()
    if ranking_metric not in frame.columns:
        raise ValueError(f"Ranking metric not found: {ranking_metric}")
    ranked_groups: list[pd.DataFrame] = []
    for trade_date, group in frame.groupby("trade_date", sort=False):
        eligible = group.dropna(subset=[ranking_metric]).sort_values(ranking_metric, ascending=False).copy()
        if eligible.empty:
            continue
        take_n = max(1, math.ceil(len(eligible) * top_fraction))
        top = eligible.head(take_n).copy()
        top.insert(0, "rank", range(1, len(top) + 1))
        top["day_universe_count"] = len(eligible)
        top["top_cutoff_count"] = take_n
        top["ranked_metric"] = ranking_metric
        ranked_groups.append(top)
    if not ranked_groups:
        return frame.iloc[0:0].copy()
    return pd.concat(ranked_groups, ignore_index=True)


def build_summary_table(
    ranked: pd.DataFrame,
    universe: pd.DataFrame,
) -> pd.DataFrame:
    if ranked.empty:
        return ranked.copy()
    # Only bring in columns from universe that are not already in ranked
    # to avoid silent _x/_y suffix pollution.
    extra_cols = ["symbol"] + [c for c in universe.columns if c not in ranked.columns]
    merged = ranked.merge(universe[extra_cols], on="symbol", how="left") if len(extra_cols) > 1 else ranked.copy()
    merged = merged.sort_values(["trade_date", "rank", "symbol"], ascending=[False, True, True]).reset_index(drop=True)
    return _add_transition_columns(merged)


def fetch_symbol_day_detail(
    databento_api_key: str,
    *,
    dataset: str,
    symbol: str,
    trade_date: date,
    display_timezone: str = DEFAULT_DISPLAY_TZ,
    window_start: time | None = None,
    window_end: time | None = None,
    premarket_anchor_et: time = time(8, 0),
    previous_close: float | None = None,
    cache_dir: str | Path | None = None,
    use_file_cache: bool = False,
    force_refresh: bool = False,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    normalized_symbol = normalize_symbol_for_databento(symbol)
    if not normalized_symbol:
        logger.info("Skipping unsupported or empty symbol: %r", symbol)
        return pd.DataFrame(), pd.DataFrame()
    ws, we = _resolve_window_for_date(trade_date, display_timezone, window_start, window_end)
    window = build_window_definition(
        trade_date,
        display_timezone=display_timezone,
        window_start=ws,
        window_end=we,
        premarket_anchor_et=premarket_anchor_et,
    )
    second_cache_path = build_cache_path(
        cache_dir,
        "symbol_detail_second",
        dataset=dataset,
        parts=[trade_date.isoformat(), normalized_symbol, display_timezone, ws.strftime("%H%M%S"), we.strftime("%H%M%S"), premarket_anchor_et.strftime("%H%M%S")],
    )
    minute_cache_path = build_cache_path(
        cache_dir,
        "symbol_detail_minute",
        dataset=dataset,
        parts=[trade_date.isoformat(), normalized_symbol, display_timezone, ws.strftime("%H%M%S"), we.strftime("%H%M%S"), premarket_anchor_et.strftime("%H%M%S")],
    )
    if use_file_cache and not force_refresh:
        cached_second = _read_cached_frame(second_cache_path, max_age_seconds=DATA_CACHE_TTL_SECONDS)
        cached_minute = _read_cached_frame(minute_cache_path, max_age_seconds=DATA_CACHE_TTL_SECONDS)
        if cached_second is not None and cached_minute is not None:
            return cached_second, cached_minute

    client = _make_databento_client(databento_api_key)
    available_end_1s = _get_schema_available_end(client, dataset, "ohlcv-1s")
    clamped_end_1s = _clamp_request_end(
        _exclusive_ohlcv_1s_end(window.fetch_end_utc), available_end_1s,
    )
    if clamped_end_1s <= pd.Timestamp(window.fetch_start_utc):
        return pd.DataFrame(), pd.DataFrame()
    try:
        store = _databento_get_range_with_retry(
            client,
            context="fetch_symbol_day_detail",
            dataset=dataset,
            symbols=[normalized_symbol],
            schema="ohlcv-1s",
            start=window.fetch_start_utc.isoformat(),
            end=clamped_end_1s.isoformat(),
        )
        frame = _store_to_frame(store, context="fetch_symbol_day_detail")
        if not frame.empty:
            _validate_frame_columns(frame, required={"symbol", "open", "high", "low", "close", "volume"}, context="fetch_symbol_day_detail")
    except Exception as exc:
        _warn_with_redacted_exception(
            f"Detail fetch failed for {normalized_symbol} on {trade_date}",
            exc,
            include_traceback=True,
        )
        return pd.DataFrame(), pd.DataFrame()
    if frame.empty:
        return pd.DataFrame(), pd.DataFrame()
    frame = frame.copy()
    frame["symbol"] = frame.get("symbol", normalized_symbol).astype(str).map(normalize_symbol_for_databento)
    frame = frame[frame["symbol"] == normalized_symbol].copy()
    if frame.empty:
        return pd.DataFrame(), pd.DataFrame()
    frame = _collapse_duplicate_symbol_seconds(frame, context="fetch_symbol_day_detail")

    tz = resolve_display_timezone(display_timezone)
    frame["display_ts"] = frame["ts"].dt.tz_convert(tz)
    frame["session"] = np.where(frame["ts"] < pd.Timestamp(window.regular_open_utc), "premarket", "regular")
    frame["second_delta_pct"] = np.where(
        pd.to_numeric(frame.get("close"), errors="coerce").shift(1) > 0,
        ((pd.to_numeric(frame.get("close"), errors="coerce") / pd.to_numeric(frame.get("close"), errors="coerce").shift(1)) - 1.0) * 100.0,
        np.nan,
    )
    if previous_close and previous_close > 0:
        frame["from_previous_close_pct"] = ((pd.to_numeric(frame.get("close"), errors="coerce") / previous_close) - 1.0) * 100.0
    else:
        frame["from_previous_close_pct"] = np.nan

    detail_window = frame[
        (frame["ts"] >= pd.Timestamp(window.window_start_local.astimezone(UTC)))
        & (frame["ts"] <= pd.Timestamp(window.window_end_local.astimezone(UTC)))
    ].copy()
    if detail_window.empty:
        minute_detail = pd.DataFrame()
    else:
        detail_window["minute"] = detail_window["display_ts"].dt.floor("min")
        minute_detail = (
            detail_window.groupby("minute", as_index=False)
            .agg(
                open=("open", "first"),
                high=("high", "max"),
                low=("low", "min"),
                close=("close", "last"),
                volume=("volume", "sum"),
                seconds=("close", "size"),
            )
            .sort_values("minute")
        )
        minute_detail["minute_delta_pct"] = np.where(
            pd.to_numeric(minute_detail["open"], errors="coerce") > 0,
            ((pd.to_numeric(minute_detail["close"], errors="coerce") / pd.to_numeric(minute_detail["open"], errors="coerce")) - 1.0) * 100.0,
            np.nan,
        )
        first_open = _safe_float(minute_detail["open"].iloc[0])
        if first_open and first_open > 0:
            minute_detail["cumulative_pct"] = ((pd.to_numeric(minute_detail["close"], errors="coerce") / first_open) - 1.0) * 100.0
        else:
            minute_detail["cumulative_pct"] = np.nan

    second_detail = frame[
        [
            "display_ts", "session", "open", "high", "low", "close", "volume",
            "second_delta_pct", "from_previous_close_pct",
        ]
    ].rename(columns={"display_ts": "timestamp"})
    second_detail = second_detail.reset_index(drop=True)
    minute_detail = minute_detail.reset_index(drop=True)
    if use_file_cache:
        _write_cached_frame(second_cache_path, second_detail)
        _write_cached_frame(minute_cache_path, minute_detail)
    return second_detail, minute_detail


def collect_detail_tables_for_summary(
    databento_api_key: str,
    *,
    dataset: str,
    summary: pd.DataFrame,
    display_timezone: str = DEFAULT_DISPLAY_TZ,
    window_start: time | None = None,
    window_end: time | None = None,
    premarket_anchor_et: time = time(8, 0),
    cache_dir: str | Path | None = None,
    use_file_cache: bool = False,
    force_refresh: bool = False,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    required_columns = {"trade_date", "symbol"}
    if summary.empty or not required_columns.issubset(summary.columns):
        return pd.DataFrame(), pd.DataFrame()

    detail_rows = (
        summary[[column for column in ["trade_date", "symbol", "previous_close"] if column in summary.columns]]
        .dropna(subset=["trade_date", "symbol"])
        .drop_duplicates(subset=["trade_date", "symbol"])
        .reset_index(drop=True)
    )
    if detail_rows.empty:
        return pd.DataFrame(), pd.DataFrame()

    combined_second: list[pd.DataFrame] = []
    combined_minute: list[pd.DataFrame] = []

    for row in detail_rows.itertuples(index=False):
        trade_date_value = pd.Timestamp(row.trade_date).date()
        symbol_value = str(row.symbol)
        previous_close = _safe_float(getattr(row, "previous_close", None))
        day_ws, day_we = _resolve_window_for_date(
            trade_date_value, display_timezone, window_start, window_end,
            default_pre_open_minutes=_DEFAULT_OPEN_WINDOW_PRE_OPEN_MINUTES,
            default_post_open_seconds=_DEFAULT_OPEN_WINDOW_POST_OPEN_SECONDS,
        )
        second_detail, minute_detail = fetch_symbol_day_detail(
            databento_api_key,
            dataset=dataset,
            symbol=symbol_value,
            trade_date=trade_date_value,
            display_timezone=display_timezone,
            window_start=day_ws,
            window_end=day_we,
            premarket_anchor_et=premarket_anchor_et,
            previous_close=previous_close,
            cache_dir=cache_dir,
            use_file_cache=use_file_cache,
            force_refresh=force_refresh,
        )
        if not second_detail.empty:
            tagged_second = second_detail.copy()
            tagged_second.insert(0, "symbol", symbol_value)
            tagged_second.insert(0, "trade_date", trade_date_value)
            combined_second.append(tagged_second)
        if not minute_detail.empty:
            tagged_minute = minute_detail.copy()
            tagged_minute.insert(0, "symbol", symbol_value)
            tagged_minute.insert(0, "trade_date", trade_date_value)
            combined_minute.append(tagged_minute)

    second_detail_all = pd.concat(combined_second, ignore_index=True) if combined_second else pd.DataFrame()
    minute_detail_all = pd.concat(combined_minute, ignore_index=True) if combined_minute else pd.DataFrame()
    return second_detail_all, minute_detail_all


def collect_full_universe_open_window_second_detail(
    databento_api_key: str,
    *,
    dataset: str,
    trading_days: list[date],
    universe_symbols: set[str],
    daily_bars: pd.DataFrame,
    symbol_day_scope: pd.DataFrame | None = None,
    display_timezone: str = DEFAULT_DISPLAY_TZ,
    window_start: time | None = None,
    window_end: time | None = None,
    premarket_anchor_et: time = time(8, 0),
    cache_dir: str | Path | None = None,
    use_file_cache: bool = False,
    force_refresh: bool = False,
) -> pd.DataFrame:
    if not trading_days or not universe_symbols:
        return pd.DataFrame(
            columns=[
                "trade_date", "symbol", "timestamp", "session", "open", "high", "low", "close", "volume",
                "trade_count", "second_delta_pct", "from_previous_close_pct",
            ]
        )

    client = _make_databento_client(databento_api_key)
    available_end_1s = _get_schema_available_end(client, dataset, "ohlcv-1s")
    normalized_scope = _normalize_symbol_day_scope(symbol_day_scope)
    scope_by_day = {
        trade_day: set(group["symbol"].astype(str).tolist())
        for trade_day, group in normalized_scope.groupby("trade_date", sort=False)
    } if not normalized_scope.empty else {}
    symbol_scope = _symbol_day_scope_token(normalized_scope) if scope_by_day else _symbol_scope_token(universe_symbols)
    previous_close_lookup = {
        (row.trade_date, row.symbol): _safe_float(row.previous_close)
        for row in daily_bars.itertuples(index=False)
    }
    all_rows: list[pd.DataFrame] = []
    runtime_unsupported_symbols: set[str] = set()
    latest_trade_day = max(trading_days) if trading_days else None

    for trade_day in trading_days:
        day_universe_symbols = scope_by_day.get(trade_day, set(universe_symbols)) if scope_by_day else set(universe_symbols)
        if not day_universe_symbols:
            continue
        day_ws, day_we = _resolve_window_for_date(
            trade_day, display_timezone, window_start, window_end,
            default_pre_open_minutes=_DEFAULT_OPEN_WINDOW_PRE_OPEN_MINUTES,
            default_post_open_seconds=_DEFAULT_OPEN_WINDOW_POST_OPEN_SECONDS,
        )
        cache_path = build_cache_path(
            cache_dir,
            "full_universe_open_second_detail",
            dataset=dataset,
            parts=[
                trade_day.isoformat(),
                display_timezone,
                day_ws.strftime("%H%M%S"),
                day_we.strftime("%H%M%S"),
                premarket_anchor_et.strftime("%H%M%S"),
                symbol_scope,
            ],
        )
        day_frame: pd.DataFrame | None = None
        if use_file_cache and not force_refresh:
            day_frame = _read_cached_frame(
                cache_path,
                max_age_seconds=_trade_day_cache_max_age_seconds(trade_day, latest_trade_day),
            )

        if day_frame is None:
            window = build_window_definition(
                trade_day,
                display_timezone=display_timezone,
                window_start=day_ws,
                window_end=day_we,
                premarket_anchor_et=premarket_anchor_et,
            )
            day_parts: list[pd.DataFrame] = []
            previous_close_by_symbol = {
                symbol: previous_close_lookup.get((trade_day, symbol))
                for symbol in day_universe_symbols
            }
            active_symbols = set(day_universe_symbols) - runtime_unsupported_symbols
            clamped_end_1s = _clamp_request_end(
                _exclusive_ohlcv_1s_end(window.window_end_local.astimezone(UTC)), available_end_1s,
            )
            if clamped_end_1s <= pd.Timestamp(window.fetch_start_utc):
                continue
            for symbols_batch in _iter_symbol_batches(active_symbols):
                try:
                    with warnings.catch_warnings(record=True) as caught_warnings:
                        warnings.simplefilter("always")
                        store = _databento_get_range_with_retry(
                            client,
                            context="collect_full_universe_open_window_second_detail",
                            dataset=dataset,
                            symbols=symbols_batch,
                            schema="ohlcv-1s",
                            start=window.fetch_start_utc.isoformat(),
                            end=clamped_end_1s.isoformat(),
                        )
                    frame = _store_to_frame(store, count=250_000, context="collect_full_universe_open_window_second_detail")
                    runtime_unsupported_symbols.update(
                        _extract_unresolved_symbols_from_warning_messages([str(item.message) for item in caught_warnings])
                    )
                    if not frame.empty:
                        _validate_frame_columns(
                            frame,
                            required={"symbol", "open", "high", "low", "close", "volume"},
                            context="collect_full_universe_open_window_second_detail",
                        )
                except Exception as exc:
                    _warn_with_redacted_exception(
                        f"Open window detail fetch failed for batch on {trade_day}, skipping",
                        exc,
                        include_traceback=True,
                    )
                    continue
                if frame.empty or "symbol" not in frame.columns:
                    continue
                frame = frame.copy()
                frame["symbol"] = frame["symbol"].astype(str).str.upper()
                frame = frame[frame["symbol"].isin(day_universe_symbols)].copy()
                if frame.empty:
                    continue
                frame = _collapse_duplicate_symbol_seconds(frame, context="collect_full_universe_open_window_second_detail")

                frame = frame.sort_values(["symbol", "ts"]).reset_index(drop=True)
                display_tz = resolve_display_timezone(display_timezone)
                frame["timestamp"] = frame["ts"].dt.tz_convert(display_tz)
                frame["session"] = np.where(frame["ts"] < pd.Timestamp(window.regular_open_utc), "premarket", "regular")
                trade_count_source: str | None = None
                for candidate in ("trade_count", "count", "n_trades", "num_trades"):
                    if candidate in frame.columns:
                        trade_count_source = candidate
                        break
                if trade_count_source is not None:
                    frame["trade_count"] = pd.to_numeric(frame[trade_count_source], errors="coerce")
                else:
                    frame["trade_count"] = np.nan
                frame["second_delta_pct"] = frame.groupby("symbol")["close"].pct_change() * 100.0
                frame["previous_close"] = frame["symbol"].map(previous_close_by_symbol)
                frame["from_previous_close_pct"] = np.where(
                    pd.to_numeric(frame["previous_close"], errors="coerce") > 0,
                    ((pd.to_numeric(frame["close"], errors="coerce") / pd.to_numeric(frame["previous_close"], errors="coerce")) - 1.0) * 100.0,
                    np.nan,
                )
                frame.insert(0, "trade_date", trade_day)
                day_parts.append(
                    frame[
                        [
                            "trade_date", "symbol", "timestamp", "session", "open", "high", "low", "close", "volume",
                            "trade_count", "second_delta_pct", "from_previous_close_pct",
                        ]
                    ].reset_index(drop=True)
                )

            day_frame = pd.concat(day_parts, ignore_index=True) if day_parts else pd.DataFrame()
            if use_file_cache and not day_frame.empty:
                _write_cached_frame(cache_path, day_frame)

        if day_frame is not None and not day_frame.empty:
            all_rows.append(day_frame)

    return pd.concat(all_rows, ignore_index=True) if all_rows else pd.DataFrame(
        columns=[
            "trade_date", "symbol", "timestamp", "session", "open", "high", "low", "close", "volume",
            "trade_count", "second_delta_pct", "from_previous_close_pct",
        ]
    )


def _build_expected_symbol_day_frame(trading_days: list[date], universe: pd.DataFrame) -> pd.DataFrame:
    if not trading_days or universe.empty or "symbol" not in universe.columns:
        return pd.DataFrame(columns=["trade_date", "symbol"])
    symbols = sorted({str(symbol).upper() for symbol in universe["symbol"].dropna().astype(str) if str(symbol).strip()})
    if not symbols:
        return pd.DataFrame(columns=["trade_date", "symbol"])
    expected = pd.MultiIndex.from_product([trading_days, symbols], names=["trade_date", "symbol"]).to_frame(index=False)
    return expected.reset_index(drop=True)


def _build_open_window_aggregates(
    second_detail_all: pd.DataFrame,
    *,
    trading_days: list[date],
    display_timezone: str,
    open_window_start: time | None = None,
    open_window_end: time | None = None,
    premarket_anchor_et: time = time(8, 0),
    reference_open_et: time = time(9, 30),
    metric_prefix: str = "",
) -> pd.DataFrame:
    def _m(name: str) -> str:
        return f"{metric_prefix}{name}" if metric_prefix else name

    metric_columns = [
        "trade_date", "symbol", _m("open_window_second_rows"), _m("open_1m_volume"), _m("open_5m_volume"), _m("open_30s_volume"),
        _m("regular_open_second_rows"), _m("regular_open_5m_second_rows"), _m("regular_open_30s_second_rows"),
        _m("regular_open_reference_price"), _m("early_dip_low_10s"), _m("early_dip_pct_10s"), _m("early_dip_second"),
        _m("reclaimed_start_price_within_30s"), _m("reclaim_second_30s"),
    ]
    if second_detail_all.empty:
        return pd.DataFrame(columns=metric_columns)

    detail = second_detail_all.copy()
    detail["trade_date"] = pd.to_datetime(detail["trade_date"], errors="coerce").dt.date
    detail["symbol"] = detail["symbol"].astype(str).str.upper()
    detail["timestamp"] = pd.to_datetime(detail["timestamp"], errors="coerce", utc=True)
    detail["timestamp"] = detail["timestamp"].dt.tz_convert(resolve_display_timezone(display_timezone))
    detail["volume"] = pd.to_numeric(detail.get("volume"), errors="coerce").fillna(0.0)
    detail = detail.dropna(subset=["trade_date", "symbol", "timestamp"])
    if detail.empty:
        return pd.DataFrame(columns=metric_columns)

    windows = {}
    for trade_day in trading_days:
        day_ws, day_we = _resolve_window_for_date(
            trade_day, display_timezone, open_window_start, open_window_end,
            default_pre_open_minutes=_DEFAULT_OPEN_WINDOW_PRE_OPEN_MINUTES,
            default_post_open_seconds=_DEFAULT_OPEN_WINDOW_POST_OPEN_SECONDS,
        )
        windows[trade_day] = build_window_definition(
            trade_day,
            display_timezone=display_timezone,
            window_start=day_ws,
            window_end=day_we,
            premarket_anchor_et=premarket_anchor_et,
        )
    tz = resolve_display_timezone(display_timezone)
    regular_open_by_day = {
        trade_day: pd.Timestamp(datetime.combine(trade_day, reference_open_et, tzinfo=US_EASTERN_TZ).astimezone(tz))
        for trade_day in windows.keys()
    }

    metrics: list[dict[str, Any]] = []
    for (trade_day, symbol), group in detail.groupby(["trade_date", "symbol"], sort=False):
        regular_open = regular_open_by_day.get(trade_day)
        if regular_open is None:
            continue
        one_minute_end = regular_open + pd.Timedelta(minutes=1)
        five_minute_end = regular_open + pd.Timedelta(minutes=5)
        thirty_second_end = regular_open + pd.Timedelta(seconds=LONG_DIP_ENTRY_RECLAIM_MAX_SECONDS)
        regular_rows = group[group["timestamp"] >= regular_open]
        first_minute = regular_rows[regular_rows["timestamp"] < one_minute_end]
        first_five = regular_rows[regular_rows["timestamp"] < five_minute_end]
        first_thirty = regular_rows[regular_rows["timestamp"] < thirty_second_end]

        reference_price = np.nan
        early_dip_low = np.nan
        early_dip_pct = np.nan
        early_dip_second = np.nan
        reclaimed_start_price_within_30s = False
        reclaim_second = np.nan
        if not regular_rows.empty:
            ordered = regular_rows.sort_values("timestamp").copy()
            ordered["offset_seconds"] = (ordered["timestamp"] - regular_open).dt.total_seconds().round().astype(int)
            reference_price = float(pd.to_numeric(ordered.iloc[0].get("open"), errors="coerce"))
            if math.isfinite(reference_price) and reference_price > 0:
                first_ten = ordered[ordered["offset_seconds"] <= LONG_DIP_ENTRY_EARLY_DIP_MAX_SECONDS].copy()
                if not first_ten.empty:
                    early_dip_low = float(pd.to_numeric(first_ten.get("low"), errors="coerce").min())
                    if math.isfinite(early_dip_low):
                        low_rows = first_ten[pd.to_numeric(first_ten.get("low"), errors="coerce") == early_dip_low].sort_values("timestamp")
                        if not low_rows.empty:
                            early_dip_second = float(low_rows.iloc[0]["offset_seconds"])
                            early_dip_pct = ((early_dip_low / reference_price) - 1.0) * 100.0
                            reclaim_rows = ordered[
                                (ordered["offset_seconds"] > int(early_dip_second))
                                & (ordered["offset_seconds"] < LONG_DIP_ENTRY_RECLAIM_MAX_SECONDS)
                                & (pd.to_numeric(ordered.get("high"), errors="coerce") >= reference_price)
                            ].sort_values("timestamp")
                            if not reclaim_rows.empty:
                                reclaimed_start_price_within_30s = True
                                reclaim_second = float(reclaim_rows.iloc[0]["offset_seconds"])
        metrics.append(
            {
                "trade_date": trade_day,
                "symbol": symbol,
                _m("open_window_second_rows"): int(len(group)),
                _m("open_1m_volume"): float(first_minute["volume"].sum()),
                _m("open_5m_volume"): float(first_five["volume"].sum()),
                _m("open_30s_volume"): float(first_thirty["volume"].sum()),
                _m("regular_open_second_rows"): int(len(first_minute)),
                _m("regular_open_5m_second_rows"): int(len(first_five)),
                _m("regular_open_30s_second_rows"): int(len(first_thirty)),
                _m("regular_open_reference_price"): reference_price,
                _m("early_dip_low_10s"): early_dip_low,
                _m("early_dip_pct_10s"): early_dip_pct,
                _m("early_dip_second"): early_dip_second,
                _m("reclaimed_start_price_within_30s"): reclaimed_start_price_within_30s,
                _m("reclaim_second_30s"): reclaim_second,
            }
        )
    return pd.DataFrame(metrics)


def _build_close_imbalance_aggregates(
    second_detail_all: pd.DataFrame,
    *,
    trading_days: list[date],
    display_timezone: str,
    close_window_start: time = DEFAULT_CLOSE_IMBALANCE_WINDOW_START_ET,
    close_auction_time: time = DEFAULT_CLOSE_IMBALANCE_AUCTION_TIME_ET,
    close_window_end: time = DEFAULT_CLOSE_IMBALANCE_WINDOW_END_ET,
) -> pd.DataFrame:
    metric_columns = [
        "trade_date",
        "symbol",
        "close_window_second_rows",
        "close_preclose_second_rows",
        "close_last_minute_second_rows",
        "close_postclose_second_rows",
        "close_10m_volume",
        "close_last_1m_volume",
        "close_postclose_5m_volume",
        "close_last_1m_volume_share",
        "close_postclose_volume_share",
        "close_reference_price",
        "close_auction_reference_price",
        "close_preclose_return_pct",
        "close_postclose_return_pct",
        "close_postclose_high_pct",
        "close_postclose_low_pct",
    ]
    if second_detail_all.empty:
        return pd.DataFrame(columns=metric_columns)

    detail = second_detail_all.copy()
    detail["trade_date"] = pd.to_datetime(detail["trade_date"], errors="coerce").dt.date
    detail["symbol"] = detail["symbol"].astype(str).str.upper()
    detail["timestamp"] = pd.to_datetime(detail["timestamp"], errors="coerce", utc=True)
    detail["timestamp"] = detail["timestamp"].dt.tz_convert(resolve_display_timezone(display_timezone))
    detail["volume"] = pd.to_numeric(detail.get("volume"), errors="coerce").fillna(0.0)
    detail = detail.dropna(subset=["trade_date", "symbol", "timestamp"])
    if detail.empty:
        return pd.DataFrame(columns=metric_columns)

    tz = resolve_display_timezone(display_timezone)
    windows = {
        trade_day: {
            "start": pd.Timestamp(datetime.combine(trade_day, close_window_start, tzinfo=US_EASTERN_TZ).astimezone(tz)),
            "auction": pd.Timestamp(datetime.combine(trade_day, close_auction_time, tzinfo=US_EASTERN_TZ).astimezone(tz)),
            "end": pd.Timestamp(datetime.combine(trade_day, close_window_end, tzinfo=US_EASTERN_TZ).astimezone(tz)),
        }
        for trade_day in trading_days
    }

    metrics: list[dict[str, Any]] = []
    for (trade_day, symbol), group in detail.groupby(["trade_date", "symbol"], sort=False):
        window = windows.get(trade_day)
        if window is None:
            continue
        ordered = group.sort_values("timestamp").copy()
        window_rows = ordered[(ordered["timestamp"] >= window["start"]) & (ordered["timestamp"] < window["end"])]
        if window_rows.empty:
            continue

        preclose_rows = window_rows[window_rows["timestamp"] < window["auction"]]
        last_minute_start = window["auction"] - pd.Timedelta(minutes=1)
        last_minute_rows = preclose_rows[preclose_rows["timestamp"] >= last_minute_start]
        postclose_rows = window_rows[window_rows["timestamp"] >= window["auction"]]

        close_reference_price = np.nan
        close_auction_reference_price = np.nan
        close_preclose_return_pct = np.nan
        close_postclose_return_pct = np.nan
        close_postclose_high_pct = np.nan
        close_postclose_low_pct = np.nan

        if not preclose_rows.empty:
            close_reference_price = float(pd.to_numeric(preclose_rows.iloc[0].get("open"), errors="coerce"))
            close_auction_reference_price = float(pd.to_numeric(preclose_rows.iloc[-1].get("close"), errors="coerce"))
            if math.isfinite(close_reference_price) and close_reference_price > 0 and math.isfinite(close_auction_reference_price):
                close_preclose_return_pct = ((close_auction_reference_price / close_reference_price) - 1.0) * 100.0

        if not postclose_rows.empty and math.isfinite(close_auction_reference_price) and close_auction_reference_price > 0:
            postclose_last_close = float(pd.to_numeric(postclose_rows.iloc[-1].get("close"), errors="coerce"))
            postclose_high = float(pd.to_numeric(postclose_rows.get("high"), errors="coerce").max())
            postclose_low = float(pd.to_numeric(postclose_rows.get("low"), errors="coerce").min())
            if math.isfinite(postclose_last_close):
                close_postclose_return_pct = ((postclose_last_close / close_auction_reference_price) - 1.0) * 100.0
            if math.isfinite(postclose_high):
                close_postclose_high_pct = ((postclose_high / close_auction_reference_price) - 1.0) * 100.0
            if math.isfinite(postclose_low):
                close_postclose_low_pct = ((postclose_low / close_auction_reference_price) - 1.0) * 100.0

        close_10m_volume = float(preclose_rows["volume"].sum())
        close_last_1m_volume = float(last_minute_rows["volume"].sum())
        close_postclose_5m_volume = float(postclose_rows["volume"].sum())
        close_window_total_volume = float(window_rows["volume"].sum())
        close_last_1m_volume_share = (close_last_1m_volume / close_10m_volume) if close_10m_volume > 0 else np.nan
        close_postclose_volume_share = (close_postclose_5m_volume / close_window_total_volume) if close_window_total_volume > 0 else np.nan

        metrics.append(
            {
                "trade_date": trade_day,
                "symbol": symbol,
                "close_window_second_rows": int(len(window_rows)),
                "close_preclose_second_rows": int(len(preclose_rows)),
                "close_last_minute_second_rows": int(len(last_minute_rows)),
                "close_postclose_second_rows": int(len(postclose_rows)),
                "close_10m_volume": close_10m_volume,
                "close_last_1m_volume": close_last_1m_volume,
                "close_postclose_5m_volume": close_postclose_5m_volume,
                "close_last_1m_volume_share": close_last_1m_volume_share,
                "close_postclose_volume_share": close_postclose_volume_share,
                "close_reference_price": close_reference_price,
                "close_auction_reference_price": close_auction_reference_price,
                "close_preclose_return_pct": close_preclose_return_pct,
                "close_postclose_return_pct": close_postclose_return_pct,
                "close_postclose_high_pct": close_postclose_high_pct,
                "close_postclose_low_pct": close_postclose_low_pct,
            }
        )
    return pd.DataFrame(metrics)


def _normalize_trade_price_series(series: pd.Series | Any) -> pd.Series:
    numeric = pd.to_numeric(series, errors="coerce")
    if numeric.empty:
        return numeric
    numeric = numeric.where(numeric.abs() < 9e18, np.nan)
    finite = numeric[np.isfinite(numeric)]
    if finite.empty:
        return numeric
    median_abs = float(finite.abs().median())
    if median_abs >= 1_000_000.0:
        return numeric / 1_000_000_000.0
    return numeric


def _load_databento_publisher_lookup(client: Any) -> dict[int, dict[str, str]]:
    try:
        publishers = client.metadata.list_publishers()
    except Exception:
        logger.debug("metadata.list_publishers failed; close venue metadata unavailable", exc_info=True)
        return {}
    if not isinstance(publishers, list):
        return {}

    lookup: dict[int, dict[str, str]] = {}
    for item in publishers:
        if not isinstance(item, dict):
            continue
        publisher_id = item.get("publisher_id")
        if publisher_id in (None, ""):
            continue
        try:
            publisher_id_int = int(str(publisher_id))
        except (TypeError, ValueError):
            continue
        description = str(item.get("description") or item.get("publisher") or item.get("venue") or "").strip()
        description_lower = description.lower()
        venue_label = description.partition("-")[2].strip() if "-" in description else description
        if not venue_label:
            venue_label = description or f"publisher_{publisher_id_int}"
        is_trf = "trf" in description_lower
        lookup[publisher_id_int] = {
            "description": description or venue_label,
            "venue_label": venue_label,
            "venue_class": "off_exchange_trf" if is_trf else "lit_exchange",
        }
    return lookup


def collect_full_universe_close_trade_detail(
    databento_api_key: str,
    *,
    dataset: str,
    trading_days: list[date],
    universe_symbols: set[str],
    symbol_day_scope: pd.DataFrame | None = None,
    display_timezone: str = DEFAULT_DISPLAY_TZ,
    cache_dir: str | Path | None = None,
    use_file_cache: bool = False,
    force_refresh: bool = False,
) -> pd.DataFrame:
    output_columns = [
        "trade_date", "symbol", "timestamp", "ts_event", "ts_recv", "publisher_id", "publisher", "venue_class",
        "side", "price", "size", "flags", "sequence", "ts_in_delta",
    ]
    if not trading_days or not universe_symbols:
        return pd.DataFrame(columns=output_columns)

    client = _make_databento_client(databento_api_key)
    available_end = _get_schema_available_end(client, dataset, "trades")
    publisher_lookup = _load_databento_publisher_lookup(client)
    normalized_scope = _normalize_symbol_day_scope(symbol_day_scope)
    scope_by_day = {
        trade_day: set(group["symbol"].astype(str).tolist())
        for trade_day, group in normalized_scope.groupby("trade_date", sort=False)
    } if not normalized_scope.empty else {}
    symbol_scope = _symbol_day_scope_token(normalized_scope) if scope_by_day else _symbol_scope_token(universe_symbols)

    display_tz = resolve_display_timezone(display_timezone)
    all_rows: list[pd.DataFrame] = []
    runtime_unsupported_symbols: set[str] = set()
    latest_trade_day = max(trading_days) if trading_days else None
    for trade_day in trading_days:
        day_universe_symbols = scope_by_day.get(trade_day, set(universe_symbols)) if scope_by_day else set(universe_symbols)
        if not day_universe_symbols:
            continue
        local_start = datetime.combine(trade_day, DEFAULT_CLOSE_IMBALANCE_WINDOW_START_ET, tzinfo=US_EASTERN_TZ).astimezone(display_tz)
        local_end = datetime.combine(trade_day, DEFAULT_CLOSE_IMBALANCE_WINDOW_END_ET, tzinfo=US_EASTERN_TZ).astimezone(display_tz)
        fetch_start_utc = pd.Timestamp(local_start.astimezone(UTC))
        fetch_end_utc = _clamp_request_end(pd.Timestamp(local_end.astimezone(UTC)), available_end)
        if fetch_end_utc <= fetch_start_utc:
            continue
        cache_path = build_cache_path(
            cache_dir,
            "full_universe_close_trade_detail",
            dataset=dataset,
            parts=[trade_day.isoformat(), display_timezone, symbol_scope],
        )
        day_frame: pd.DataFrame | None = None
        if use_file_cache and not force_refresh:
            day_frame = _read_cached_frame(
                cache_path,
                max_age_seconds=_trade_day_cache_max_age_seconds(trade_day, latest_trade_day),
            )

        if day_frame is None:
            day_parts: list[pd.DataFrame] = []
            active_symbols = set(day_universe_symbols) - runtime_unsupported_symbols
            for symbols_batch in _iter_symbol_batches(active_symbols):
                try:
                    with warnings.catch_warnings(record=True) as caught_warnings:
                        warnings.simplefilter("always")
                        store = _databento_get_range_with_retry(
                            client,
                            context="collect_full_universe_close_trade_detail",
                            dataset=dataset,
                            symbols=symbols_batch,
                            schema="trades",
                            start=fetch_start_utc.isoformat(),
                            end=fetch_end_utc.isoformat(),
                        )
                    frame = _store_to_frame(store, count=250_000, context="collect_full_universe_close_trade_detail")
                    runtime_unsupported_symbols.update(
                        _extract_unresolved_symbols_from_warning_messages([str(item.message) for item in caught_warnings])
                    )
                    if not frame.empty:
                        _validate_frame_columns(
                            frame,
                            required={"symbol", "publisher_id", "size", "price", "flags", "sequence"},
                            context="collect_full_universe_close_trade_detail",
                        )
                except Exception as exc:
                    _warn_with_redacted_exception(
                        f"Close trade detail fetch failed for batch on {trade_day}, skipping",
                        exc,
                        include_traceback=True,
                    )
                    continue
                if frame.empty or "symbol" not in frame.columns:
                    continue
                frame = frame.copy()
                frame["symbol"] = frame["symbol"].astype(str).str.upper()
                frame = frame[frame["symbol"].isin(day_universe_symbols)].copy()
                if frame.empty:
                    continue
                frame["timestamp"] = frame["ts"].dt.tz_convert(display_tz)
                frame["ts_recv"] = pd.to_datetime(frame["ts"], errors="coerce", utc=True)
                frame["ts_event"] = pd.to_datetime(frame.get("ts_event"), errors="coerce", utc=True)
                frame["publisher_id"] = pd.to_numeric(frame.get("publisher_id"), errors="coerce").astype("Int64")
                frame["publisher"] = frame["publisher_id"].map(lambda value: publisher_lookup.get(int(value), {}).get("venue_label") if pd.notna(value) else None)
                frame["venue_class"] = frame["publisher_id"].map(lambda value: publisher_lookup.get(int(value), {}).get("venue_class") if pd.notna(value) else None)
                frame["side"] = frame.get("side", "N").astype(str).str.upper().replace("", "N")
                frame["price"] = _normalize_trade_price_series(frame.get("price"))
                frame["size"] = pd.to_numeric(frame.get("size"), errors="coerce")
                frame["flags"] = pd.to_numeric(frame.get("flags"), errors="coerce").fillna(0).astype(int)
                frame["sequence"] = pd.to_numeric(frame.get("sequence"), errors="coerce")
                frame["ts_in_delta"] = pd.to_numeric(frame.get("ts_in_delta"), errors="coerce")
                frame.insert(0, "trade_date", trade_day)
                day_parts.append(frame[output_columns].reset_index(drop=True))

            day_frame = pd.concat(day_parts, ignore_index=True) if day_parts else pd.DataFrame(columns=output_columns)
            if use_file_cache and not day_frame.empty:
                _write_cached_frame(cache_path, day_frame)

        if day_frame is not None and not day_frame.empty:
            all_rows.append(day_frame)

    return pd.concat(all_rows, ignore_index=True) if all_rows else pd.DataFrame(columns=output_columns)


def collect_full_universe_close_outcome_minute_detail(
    databento_api_key: str,
    *,
    dataset: str,
    trading_days: list[date],
    universe_symbols: set[str],
    symbol_day_scope: pd.DataFrame | None = None,
    display_timezone: str = DEFAULT_DISPLAY_TZ,
    cache_dir: str | Path | None = None,
    use_file_cache: bool = False,
    force_refresh: bool = False,
) -> pd.DataFrame:
    output_columns = ["trade_date", "symbol", "timestamp", "open", "high", "low", "close", "volume"]
    if not trading_days or not universe_symbols:
        return pd.DataFrame(columns=output_columns)

    client = _make_databento_client(databento_api_key)
    available_end = _get_schema_available_end(client, dataset, "ohlcv-1m")
    normalized_scope = _normalize_symbol_day_scope(symbol_day_scope)
    scope_by_day = {
        trade_day: set(group["symbol"].astype(str).tolist())
        for trade_day, group in normalized_scope.groupby("trade_date", sort=False)
    } if not normalized_scope.empty else {}
    symbol_scope = _symbol_day_scope_token(normalized_scope) if scope_by_day else _symbol_scope_token(universe_symbols)
    display_tz = resolve_display_timezone(display_timezone)
    all_rows: list[pd.DataFrame] = []
    runtime_unsupported_symbols: set[str] = set()
    latest_trade_day = max(trading_days) if trading_days else None

    for trade_day in trading_days:
        day_universe_symbols = scope_by_day.get(trade_day, set(universe_symbols)) if scope_by_day else set(universe_symbols)
        if not day_universe_symbols:
            continue
        local_start = datetime.combine(trade_day, DEFAULT_CLOSE_IMBALANCE_AUCTION_TIME_ET, tzinfo=US_EASTERN_TZ).astimezone(display_tz)
        local_end = datetime.combine(trade_day, DEFAULT_CLOSE_IMBALANCE_AFTERHOURS_END_ET, tzinfo=US_EASTERN_TZ).astimezone(display_tz)
        fetch_start_utc = pd.Timestamp(local_start.astimezone(UTC))
        fetch_end_utc = _clamp_request_end(pd.Timestamp(local_end.astimezone(UTC)), available_end)
        if fetch_end_utc <= fetch_start_utc:
            continue
        cache_path = build_cache_path(
            cache_dir,
            "full_universe_close_outcome_minute_detail",
            dataset=dataset,
            parts=[trade_day.isoformat(), display_timezone, symbol_scope],
        )
        day_frame: pd.DataFrame | None = None
        if use_file_cache and not force_refresh:
            day_frame = _read_cached_frame(
                cache_path,
                max_age_seconds=_trade_day_cache_max_age_seconds(trade_day, latest_trade_day),
            )

        if day_frame is None:
            day_parts: list[pd.DataFrame] = []
            active_symbols = set(day_universe_symbols) - runtime_unsupported_symbols
            for symbols_batch in _iter_symbol_batches(active_symbols):
                try:
                    with warnings.catch_warnings(record=True) as caught_warnings:
                        warnings.simplefilter("always")
                        store = _databento_get_range_with_retry(
                            client,
                            context="collect_full_universe_close_outcome_minute_detail",
                            dataset=dataset,
                            symbols=symbols_batch,
                            schema="ohlcv-1m",
                            start=fetch_start_utc.isoformat(),
                            end=fetch_end_utc.isoformat(),
                        )
                    frame = _store_to_frame(store, count=250_000, context="collect_full_universe_close_outcome_minute_detail")
                    runtime_unsupported_symbols.update(
                        _extract_unresolved_symbols_from_warning_messages([str(item.message) for item in caught_warnings])
                    )
                    if not frame.empty:
                        _validate_frame_columns(
                            frame,
                            required={"symbol", "open", "high", "low", "close", "volume"},
                            context="collect_full_universe_close_outcome_minute_detail",
                        )
                except Exception as exc:
                    _warn_with_redacted_exception(
                        f"Close outcome minute detail fetch failed for batch on {trade_day}, skipping",
                        exc,
                        include_traceback=True,
                    )
                    continue
                if frame.empty or "symbol" not in frame.columns:
                    continue
                frame = frame.copy()
                frame["symbol"] = frame["symbol"].astype(str).str.upper()
                frame = frame[frame["symbol"].isin(day_universe_symbols)].copy()
                if frame.empty:
                    continue
                frame["timestamp"] = frame["ts"].dt.tz_convert(display_tz)
                frame.insert(0, "trade_date", trade_day)
                day_parts.append(frame[output_columns].reset_index(drop=True))
            day_frame = pd.concat(day_parts, ignore_index=True) if day_parts else pd.DataFrame(columns=output_columns)
            if use_file_cache and not day_frame.empty:
                _write_cached_frame(cache_path, day_frame)

        if day_frame is not None and not day_frame.empty:
            all_rows.append(day_frame)

    return pd.concat(all_rows, ignore_index=True) if all_rows else pd.DataFrame(columns=output_columns)


def _build_close_trade_aggregates(
    close_trade_detail_all: pd.DataFrame,
    *,
    trading_days: list[date],
    display_timezone: str,
    close_auction_time: time = DEFAULT_CLOSE_IMBALANCE_AUCTION_TIME_ET,
) -> pd.DataFrame:
    metric_columns = [
        "trade_date", "symbol", "close_trade_print_count", "close_trade_share_volume", "close_trade_clean_print_count",
        "close_trade_clean_share_volume", "close_trade_bad_ts_recv_count", "close_trade_maybe_bad_book_count",
        "close_trade_publisher_specific_flag_count", "close_trade_sequence_break_count", "close_trade_event_time_regression_count",
        "close_trade_unknown_side_count", "close_trade_unknown_side_share", "close_trade_hygiene_score",
        "close_trade_unique_publishers", "close_trade_trf_print_count", "close_trade_trf_share_volume",
        "close_trade_lit_print_count", "close_trade_lit_share_volume", "close_trade_trf_volume_share",
        "close_trade_lit_volume_share", "close_trade_has_trf_activity", "close_trade_has_lit_activity",
        "close_trade_has_lit_followthrough",
    ]
    if close_trade_detail_all.empty:
        return pd.DataFrame(columns=metric_columns)

    detail = close_trade_detail_all.copy()
    detail["trade_date"] = pd.to_datetime(detail["trade_date"], errors="coerce").dt.date
    detail["symbol"] = detail["symbol"].astype(str).str.upper()
    detail["timestamp"] = pd.to_datetime(detail["timestamp"], errors="coerce", utc=True).dt.tz_convert(resolve_display_timezone(display_timezone))
    detail["ts_recv"] = pd.to_datetime(detail.get("ts_recv"), errors="coerce", utc=True)
    detail["ts_event"] = pd.to_datetime(detail.get("ts_event"), errors="coerce", utc=True)
    detail["size"] = pd.to_numeric(detail.get("size"), errors="coerce").fillna(0.0)
    detail["price"] = _normalize_trade_price_series(detail.get("price"))
    detail["flags"] = pd.to_numeric(detail.get("flags"), errors="coerce").fillna(0).astype(int)
    detail["sequence"] = pd.to_numeric(detail.get("sequence"), errors="coerce")
    detail["publisher_id"] = pd.to_numeric(detail.get("publisher_id"), errors="coerce").astype("Int64")
    detail["side"] = detail.get("side", "N").astype(str).str.upper().replace("", "N")
    detail["venue_class"] = detail.get("venue_class", "").astype(str)
    detail = detail.dropna(subset=["trade_date", "symbol", "timestamp"])
    if detail.empty:
        return pd.DataFrame(columns=metric_columns)

    tz = resolve_display_timezone(display_timezone)
    auction_by_day = {
        trade_day: pd.Timestamp(datetime.combine(trade_day, close_auction_time, tzinfo=US_EASTERN_TZ).astimezone(tz))
        for trade_day in trading_days
    }

    metrics: list[dict[str, Any]] = []
    for (trade_day, symbol), group in detail.groupby(["trade_date", "symbol"], sort=False):
        auction_ts = auction_by_day.get(trade_day)
        if auction_ts is None:
            continue
        ordered = group.sort_values(["timestamp", "sequence"], na_position="last").copy()
        trade_count = int(len(ordered))
        if trade_count == 0:
            continue
        volume_total = float(ordered["size"].sum())
        bad_ts_mask = (ordered["flags"] & _DATABENTO_FLAG_BAD_TS_RECV) != 0
        maybe_bad_book_mask = (ordered["flags"] & _DATABENTO_FLAG_MAYBE_BAD_BOOK) != 0
        publisher_specific_mask = (ordered["flags"] & _DATABENTO_FLAG_PUBLISHER_SPECIFIC) != 0
        clean_mask = (~bad_ts_mask) & (~maybe_bad_book_mask) & (ordered["size"] > 0) & (ordered["price"] > 0)
        seq_break_count = 0
        for _, publisher_group in ordered.groupby("publisher_id", dropna=False, sort=False):
            diffs = pd.to_numeric(publisher_group["sequence"], errors="coerce").diff()
            seq_break_count += int((diffs < 0).fillna(False).sum())
        event_regression_count = int((ordered["ts_event"].diff() < pd.Timedelta(0)).fillna(False).sum()) if ordered["ts_event"].notna().any() else 0
        unknown_side_count = int(ordered["side"].eq("N").sum())
        trf_mask = ordered["venue_class"].eq("off_exchange_trf")
        lit_mask = ordered["venue_class"].eq("lit_exchange")
        last_minute = ordered[(ordered["timestamp"] >= (auction_ts - pd.Timedelta(minutes=1))) & (ordered["timestamp"] < auction_ts)]
        trf_last_minute = last_minute[last_minute["venue_class"].eq("off_exchange_trf")]
        lit_followthrough = False
        if not trf_last_minute.empty:
            first_trf_ts = trf_last_minute["timestamp"].min()
            lit_followthrough = bool(last_minute[last_minute["venue_class"].eq("lit_exchange")]["timestamp"].ge(first_trf_ts).any())
        hygiene_penalty = (
            int(bad_ts_mask.sum())
            + int(maybe_bad_book_mask.sum())
            + seq_break_count
            + event_regression_count
        ) / max(trade_count, 1)
        metrics.append(
            {
                "trade_date": trade_day,
                "symbol": symbol,
                "close_trade_print_count": trade_count,
                "close_trade_share_volume": volume_total,
                "close_trade_clean_print_count": int(clean_mask.sum()),
                "close_trade_clean_share_volume": float(ordered.loc[clean_mask, "size"].sum()),
                "close_trade_bad_ts_recv_count": int(bad_ts_mask.sum()),
                "close_trade_maybe_bad_book_count": int(maybe_bad_book_mask.sum()),
                "close_trade_publisher_specific_flag_count": int(publisher_specific_mask.sum()),
                "close_trade_sequence_break_count": seq_break_count,
                "close_trade_event_time_regression_count": event_regression_count,
                "close_trade_unknown_side_count": unknown_side_count,
                "close_trade_unknown_side_share": (unknown_side_count / trade_count) if trade_count > 0 else np.nan,
                "close_trade_hygiene_score": max(0.0, 1.0 - hygiene_penalty),
                "close_trade_unique_publishers": int(ordered["publisher_id"].dropna().nunique()),
                "close_trade_trf_print_count": int(trf_mask.sum()),
                "close_trade_trf_share_volume": float(ordered.loc[trf_mask, "size"].sum()),
                "close_trade_lit_print_count": int(lit_mask.sum()),
                "close_trade_lit_share_volume": float(ordered.loc[lit_mask, "size"].sum()),
                "close_trade_trf_volume_share": (float(ordered.loc[trf_mask, "size"].sum()) / volume_total) if volume_total > 0 else np.nan,
                "close_trade_lit_volume_share": (float(ordered.loc[lit_mask, "size"].sum()) / volume_total) if volume_total > 0 else np.nan,
                "close_trade_has_trf_activity": bool(trf_mask.any()),
                "close_trade_has_lit_activity": bool(lit_mask.any()),
                "close_trade_has_lit_followthrough": lit_followthrough,
            }
        )
    return pd.DataFrame(metrics)


def _build_close_outcome_aggregates(close_outcome_minute_detail_all: pd.DataFrame) -> pd.DataFrame:
    metric_columns = [
        "trade_date", "symbol", "close_afterhours_minute_rows", "close_afterhours_volume",
        "close_last_price_2000", "close_high_price_1600_2000", "close_low_price_1600_2000",
    ]
    if close_outcome_minute_detail_all.empty:
        return pd.DataFrame(columns=metric_columns)

    detail = close_outcome_minute_detail_all.copy()
    detail["trade_date"] = pd.to_datetime(detail["trade_date"], errors="coerce").dt.date
    detail["symbol"] = detail["symbol"].astype(str).str.upper()
    detail["timestamp"] = pd.to_datetime(detail["timestamp"], errors="coerce", utc=True)
    detail["volume"] = pd.to_numeric(detail.get("volume"), errors="coerce").fillna(0.0)
    detail = detail.dropna(subset=["trade_date", "symbol", "timestamp"])
    if detail.empty:
        return pd.DataFrame(columns=metric_columns)

    metrics: list[dict[str, Any]] = []
    for (trade_day, symbol), group in detail.groupby(["trade_date", "symbol"], sort=False):
        ordered = group.sort_values("timestamp")
        metrics.append(
            {
                "trade_date": trade_day,
                "symbol": symbol,
                "close_afterhours_minute_rows": int(len(ordered)),
                "close_afterhours_volume": float(ordered["volume"].sum()),
                "close_last_price_2000": float(pd.to_numeric(ordered.iloc[-1].get("close"), errors="coerce")),
                "close_high_price_1600_2000": float(pd.to_numeric(ordered.get("high"), errors="coerce").max()),
                "close_low_price_1600_2000": float(pd.to_numeric(ordered.get("low"), errors="coerce").min()),
            }
        )
    return pd.DataFrame(metrics)


def build_daily_features_full_universe(
    *,
    trading_days: list[date],
    universe: pd.DataFrame,
    daily_bars: pd.DataFrame,
    intraday: pd.DataFrame,
    second_detail_all: pd.DataFrame,
    close_detail_all: pd.DataFrame | None = None,
    close_trade_detail_all: pd.DataFrame | None = None,
    close_outcome_minute_detail_all: pd.DataFrame | None = None,
    display_timezone: str = DEFAULT_DISPLAY_TZ,
    premarket_anchor_et: time = time(8, 0),
    open_window_start: time | None = None,
    open_window_end: time | None = None,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    expected = _build_expected_symbol_day_frame(trading_days, universe)
    if expected.empty:
        empty_features = pd.DataFrame(columns=["trade_date", "symbol"])
        empty_coverage = pd.DataFrame(columns=["trade_date", "symbol", "has_daily_bar", "has_intraday_summary", "has_open_window_detail", "has_close_window_detail", "exclusion_reason"])
        return empty_features, empty_coverage

    if close_detail_all is None:
        close_detail_all = pd.DataFrame()
    if close_trade_detail_all is None:
        close_trade_detail_all = pd.DataFrame()
    if close_outcome_minute_detail_all is None:
        close_outcome_minute_detail_all = pd.DataFrame()

    optional_universe_columns = [column for column in FULL_UNIVERSE_OPTIONAL_FEATURE_COLUMNS if column in universe.columns]
    universe_columns = [column for column in ["symbol", "company_name", "exchange", "sector", "industry", "market_cap", *optional_universe_columns] if column in universe.columns]
    universe_frame = universe[universe_columns].copy() if universe_columns else pd.DataFrame(columns=["symbol"])
    if not universe_frame.empty:
        universe_frame["symbol"] = universe_frame["symbol"].astype(str).str.upper()
        universe_frame = universe_frame.drop_duplicates(subset=["symbol"]).reset_index(drop=True)

    daily = daily_bars.copy()
    if not daily.empty:
        daily["trade_date"] = pd.to_datetime(daily["trade_date"], errors="coerce").dt.date
        daily["symbol"] = daily["symbol"].astype(str).str.upper()
        daily = daily.rename(
            columns={
                "open": "day_open",
                "high": "day_high",
                "low": "day_low",
                "close": "day_close",
                "volume": "day_volume",
            }
        )
        daily = daily[[column for column in ["trade_date", "symbol", "day_open", "day_high", "day_low", "day_close", "day_volume", "previous_close"] if column in daily.columns]]

    intraday_columns = [
        "trade_date", "symbol", "previous_close", "premarket_price", "has_premarket_data", "market_open_price",
        "window_start_price", "current_price", "exact_1000_price", "window_high", "window_low", "window_volume", "seconds_in_window",
        "window_return_pct", "window_range_pct", "realized_vol_pct", "prev_close_to_premarket_abs",
        "prev_close_to_premarket_pct", "premarket_to_open_abs", "premarket_to_open_pct", "open_to_current_abs",
        "open_to_current_pct",
    ]
    intraday_frame = intraday[[column for column in intraday_columns if column in intraday.columns]].copy() if not intraday.empty else pd.DataFrame(columns=["trade_date", "symbol"])
    if not intraday_frame.empty:
        intraday_frame["trade_date"] = pd.to_datetime(intraday_frame["trade_date"], errors="coerce").dt.date
        intraday_frame["symbol"] = intraday_frame["symbol"].astype(str).str.upper()
        intraday_frame = intraday_frame.drop_duplicates(subset=["trade_date", "symbol"]).reset_index(drop=True)

    open_window = _build_open_window_aggregates(
        second_detail_all,
        trading_days=trading_days,
        display_timezone=display_timezone,
        open_window_start=open_window_start,
        open_window_end=open_window_end,
        premarket_anchor_et=premarket_anchor_et,
        reference_open_et=time(9, 30),
        metric_prefix="",
    )
    open_window_0800 = _build_open_window_aggregates(
        second_detail_all,
        trading_days=trading_days,
        display_timezone=display_timezone,
        open_window_start=open_window_start,
        open_window_end=open_window_end,
        premarket_anchor_et=premarket_anchor_et,
        reference_open_et=time(8, 0),
        metric_prefix="focus_0800_",
    )
    open_window_0400 = _build_open_window_aggregates(
        second_detail_all,
        trading_days=trading_days,
        display_timezone=display_timezone,
        open_window_start=open_window_start,
        open_window_end=open_window_end,
        premarket_anchor_et=premarket_anchor_et,
        reference_open_et=time(4, 0),
        metric_prefix="focus_0400_",
    )
    if not open_window.empty:
        open_window["trade_date"] = pd.to_datetime(open_window["trade_date"], errors="coerce").dt.date
        open_window["symbol"] = open_window["symbol"].astype(str).str.upper()
    if not open_window_0800.empty:
        open_window_0800["trade_date"] = pd.to_datetime(open_window_0800["trade_date"], errors="coerce").dt.date
        open_window_0800["symbol"] = open_window_0800["symbol"].astype(str).str.upper()
    if not open_window_0400.empty:
        open_window_0400["trade_date"] = pd.to_datetime(open_window_0400["trade_date"], errors="coerce").dt.date
        open_window_0400["symbol"] = open_window_0400["symbol"].astype(str).str.upper()
    close_window = _build_close_imbalance_aggregates(
        close_detail_all,
        trading_days=trading_days,
        display_timezone=display_timezone,
    )
    close_trade_window = _build_close_trade_aggregates(
        close_trade_detail_all,
        trading_days=trading_days,
        display_timezone=display_timezone,
    )
    close_outcome_window = _build_close_outcome_aggregates(close_outcome_minute_detail_all)
    if not close_window.empty:
        close_window["trade_date"] = pd.to_datetime(close_window["trade_date"], errors="coerce").dt.date
        close_window["symbol"] = close_window["symbol"].astype(str).str.upper()
    if not close_trade_window.empty:
        close_trade_window["trade_date"] = pd.to_datetime(close_trade_window["trade_date"], errors="coerce").dt.date
        close_trade_window["symbol"] = close_trade_window["symbol"].astype(str).str.upper()
    if not close_outcome_window.empty:
        close_outcome_window["trade_date"] = pd.to_datetime(close_outcome_window["trade_date"], errors="coerce").dt.date
        close_outcome_window["symbol"] = close_outcome_window["symbol"].astype(str).str.upper()

    features = expected.merge(universe_frame, on="symbol", how="left")
    if not daily.empty:
        features = features.merge(daily, on=["trade_date", "symbol"], how="left")
    if not intraday_frame.empty:
        features = features.merge(intraday_frame, on=["trade_date", "symbol"], how="left", suffixes=("", "_intraday"))
    if not open_window.empty:
        features = features.merge(open_window, on=["trade_date", "symbol"], how="left")
    if not open_window_0800.empty:
        features = features.merge(open_window_0800, on=["trade_date", "symbol"], how="left")
    if not open_window_0400.empty:
        features = features.merge(open_window_0400, on=["trade_date", "symbol"], how="left")
    if not close_window.empty:
        features = features.merge(close_window, on=["trade_date", "symbol"], how="left")
    if not close_trade_window.empty:
        features = features.merge(close_trade_window, on=["trade_date", "symbol"], how="left")
    if not close_outcome_window.empty:
        features = features.merge(close_outcome_window, on=["trade_date", "symbol"], how="left")

    if "previous_close_intraday" in features.columns:
        features["previous_close"] = features["previous_close"].combine_first(features["previous_close_intraday"])
        features = features.drop(columns=["previous_close_intraday"])

    if "has_premarket_data" in features.columns:
        features["has_premarket_data"] = features["has_premarket_data"].where(features["has_premarket_data"].notna(), False).astype(bool)
    else:
        features["has_premarket_data"] = False
    if "open_window_second_rows" not in features.columns:
        features["open_window_second_rows"] = 0
    features["open_window_second_rows"] = pd.to_numeric(features["open_window_second_rows"], errors="coerce").fillna(0).astype(int)
    if "focus_0800_open_window_second_rows" not in features.columns:
        features["focus_0800_open_window_second_rows"] = 0
    features["focus_0800_open_window_second_rows"] = pd.to_numeric(features["focus_0800_open_window_second_rows"], errors="coerce").fillna(0).astype(int)
    if "focus_0400_open_window_second_rows" not in features.columns:
        features["focus_0400_open_window_second_rows"] = 0
    features["focus_0400_open_window_second_rows"] = pd.to_numeric(features["focus_0400_open_window_second_rows"], errors="coerce").fillna(0).astype(int)
    features["has_open_window_detail"] = features["open_window_second_rows"] > 0
    if "regular_open_second_rows" in features.columns:
        features["has_open_window_detail"] = features["has_open_window_detail"] | (features["regular_open_second_rows"] > 0)
    if "focus_0800_regular_open_second_rows" in features.columns:
        features["has_open_window_detail"] = features["has_open_window_detail"] | (pd.to_numeric(features["focus_0800_regular_open_second_rows"], errors="coerce").fillna(0).astype(int) > 0)
    if "focus_0400_regular_open_second_rows" in features.columns:
        features["has_open_window_detail"] = features["has_open_window_detail"] | (pd.to_numeric(features["focus_0400_regular_open_second_rows"], errors="coerce").fillna(0).astype(int) > 0)
    if "close_window_second_rows" not in features.columns:
        features["close_window_second_rows"] = 0
    features["close_window_second_rows"] = pd.to_numeric(features["close_window_second_rows"], errors="coerce").fillna(0).astype(int)
    features["has_close_window_detail"] = features["close_window_second_rows"] > 0
    features["premarket_anchor_et"] = premarket_anchor_et.strftime("%H:%M:%S")
    features["premarket_price_source"] = "last_ohlcv_1s_close_between_anchor_and_regular_open"
    features["internal_display_timezone"] = display_timezone
    features = features.sort_values(["symbol", "trade_date"]).reset_index(drop=True)

    if "close_trade_has_lit_followthrough" in features.columns:
        features["close_trade_has_lit_followthrough"] = pd.Series(features["close_trade_has_lit_followthrough"], dtype="boolean").fillna(False).astype(bool)
    if "close_trade_has_trf_activity" in features.columns:
        features["close_trade_has_trf_activity"] = pd.Series(features["close_trade_has_trf_activity"], dtype="boolean").fillna(False).astype(bool)
    if "close_trade_has_lit_activity" in features.columns:
        features["close_trade_has_lit_activity"] = pd.Series(features["close_trade_has_lit_activity"], dtype="boolean").fillna(False).astype(bool)
    close_ref = pd.to_numeric(features.get("close_auction_reference_price"), errors="coerce")
    close_last_2000 = pd.to_numeric(features.get("close_last_price_2000"), errors="coerce")
    close_high_2000 = pd.to_numeric(features.get("close_high_price_1600_2000"), errors="coerce")
    close_low_2000 = pd.to_numeric(features.get("close_low_price_1600_2000"), errors="coerce")
    features["close_to_2000_return_pct"] = np.where(
        close_ref > 0,
        ((close_last_2000 / close_ref) - 1.0) * 100.0,
        np.nan,
    )
    features["close_to_2000_high_pct"] = np.where(
        close_ref > 0,
        ((close_high_2000 / close_ref) - 1.0) * 100.0,
        np.nan,
    )
    features["close_to_2000_low_pct"] = np.where(
        close_ref > 0,
        ((close_low_2000 / close_ref) - 1.0) * 100.0,
        np.nan,
    )
    features["next_trade_date"] = features.groupby("symbol")["trade_date"].shift(-1)
    if "market_open_price" in features.columns:
        next_day_open_from_intraday = pd.to_numeric(features.groupby("symbol")["market_open_price"].shift(-1), errors="coerce")
    else:
        next_day_open_from_intraday = pd.Series(np.nan, index=features.index, dtype=float)
    if "day_open" in features.columns:
        next_day_open_from_daily = pd.to_numeric(features.groupby("symbol")["day_open"].shift(-1), errors="coerce")
    else:
        next_day_open_from_daily = pd.Series(np.nan, index=features.index, dtype=float)
    features["next_day_open_price"] = pd.Series(next_day_open_from_intraday).combine_first(next_day_open_from_daily)
    if "exact_1000_price" in features.columns:
        next_day_window_end_from_exact = pd.to_numeric(features.groupby("symbol")["exact_1000_price"].shift(-1), errors="coerce")
    else:
        next_day_window_end_from_exact = pd.Series(np.nan, index=features.index, dtype=float)
    if "current_price" in features.columns:
        next_day_window_end_from_current = pd.to_numeric(features.groupby("symbol")["current_price"].shift(-1), errors="coerce")
    else:
        next_day_window_end_from_current = pd.Series(np.nan, index=features.index, dtype=float)
    features["next_day_window_end_price"] = pd.Series(next_day_window_end_from_exact).combine_first(next_day_window_end_from_current)
    next_open = pd.to_numeric(features.get("next_day_open_price"), errors="coerce")
    next_window_end = pd.to_numeric(features.get("next_day_window_end_price"), errors="coerce")
    features["close_to_next_open_return_pct"] = np.where(
        close_ref > 0,
        ((next_open / close_ref) - 1.0) * 100.0,
        np.nan,
    )
    features["next_open_to_window_end_return_pct"] = np.where(
        next_open > 0,
        ((next_window_end / next_open) - 1.0) * 100.0,
        np.nan,
    )
    features["close_to_next_window_end_return_pct"] = np.where(
        close_ref > 0,
        ((next_window_end / close_ref) - 1.0) * 100.0,
        np.nan,
    )
    features["has_next_day_outcome"] = features["next_trade_date"].notna() & next_open.gt(0) & next_window_end.gt(0)

    if "open_1m_volume" not in features.columns:
        features["open_1m_volume"] = np.nan
    if "open_5m_volume" not in features.columns:
        features["open_5m_volume"] = np.nan
    if "open_30s_volume" not in features.columns:
        features["open_30s_volume"] = np.nan
    if "day_volume" not in features.columns:
        features["day_volume"] = np.nan
    if "reclaimed_start_price_within_30s" not in features.columns:
        features["reclaimed_start_price_within_30s"] = False

    focus_dual_cols = [
        "open_window_second_rows",
        "open_30s_volume",
        "early_dip_pct_10s",
        "early_dip_second",
        "reclaimed_start_price_within_30s",
        "reclaim_second_30s",
    ]
    for col in focus_dual_cols:
        if col in features.columns:
            features[f"focus_0930_{col}"] = features[col]

    coalesce_pairs = [
        ("open_window_second_rows", "focus_0800_open_window_second_rows"),
        ("open_window_second_rows", "focus_0400_open_window_second_rows"),
        ("open_1m_volume", "focus_0800_open_1m_volume"),
        ("open_1m_volume", "focus_0400_open_1m_volume"),
        ("open_5m_volume", "focus_0800_open_5m_volume"),
        ("open_5m_volume", "focus_0400_open_5m_volume"),
        ("open_30s_volume", "focus_0800_open_30s_volume"),
        ("open_30s_volume", "focus_0400_open_30s_volume"),
        ("regular_open_second_rows", "focus_0800_regular_open_second_rows"),
        ("regular_open_second_rows", "focus_0400_regular_open_second_rows"),
        ("regular_open_5m_second_rows", "focus_0800_regular_open_5m_second_rows"),
        ("regular_open_5m_second_rows", "focus_0400_regular_open_5m_second_rows"),
        ("regular_open_30s_second_rows", "focus_0800_regular_open_30s_second_rows"),
        ("regular_open_30s_second_rows", "focus_0400_regular_open_30s_second_rows"),
        ("regular_open_reference_price", "focus_0800_regular_open_reference_price"),
        ("regular_open_reference_price", "focus_0400_regular_open_reference_price"),
        ("early_dip_low_10s", "focus_0800_early_dip_low_10s"),
        ("early_dip_low_10s", "focus_0400_early_dip_low_10s"),
        ("early_dip_pct_10s", "focus_0800_early_dip_pct_10s"),
        ("early_dip_pct_10s", "focus_0400_early_dip_pct_10s"),
        ("early_dip_second", "focus_0800_early_dip_second"),
        ("early_dip_second", "focus_0400_early_dip_second"),
        ("reclaim_second_30s", "focus_0800_reclaim_second_30s"),
        ("reclaim_second_30s", "focus_0400_reclaim_second_30s"),
    ]
    for primary, secondary in coalesce_pairs:
        if primary in features.columns and secondary in features.columns:
            features[primary] = features[primary].combine_first(features[secondary])
    if "reclaimed_start_price_within_30s" in features.columns and "focus_0800_reclaimed_start_price_within_30s" in features.columns:
        primary_bool = pd.Series(features["reclaimed_start_price_within_30s"], dtype="boolean").fillna(False)
        secondary_bool = pd.Series(features["focus_0800_reclaimed_start_price_within_30s"], dtype="boolean").fillna(False)
        features["reclaimed_start_price_within_30s"] = (primary_bool | secondary_bool).astype(bool)
    if "reclaimed_start_price_within_30s" in features.columns and "focus_0400_reclaimed_start_price_within_30s" in features.columns:
        primary_bool = pd.Series(features["reclaimed_start_price_within_30s"], dtype="boolean").fillna(False)
        secondary_bool = pd.Series(features["focus_0400_reclaimed_start_price_within_30s"], dtype="boolean").fillna(False)
        features["reclaimed_start_price_within_30s"] = (primary_bool | secondary_bool).astype(bool)

    reclaimed_flag = pd.Series(features["reclaimed_start_price_within_30s"], dtype="boolean")
    features["reclaimed_start_price_within_30s"] = reclaimed_flag.fillna(False).astype(bool)

    features["avg_open_1m_volume_20d"] = features.groupby("symbol")["open_1m_volume"].transform(lambda s: s.shift(1).rolling(20, min_periods=1).mean())
    features["avg_open_5m_volume_20d"] = features.groupby("symbol")["open_5m_volume"].transform(lambda s: s.shift(1).rolling(20, min_periods=1).mean())
    features["avg_day_volume_20d"] = features.groupby("symbol")["day_volume"].transform(lambda s: s.shift(1).rolling(20, min_periods=1).mean())
    features["open_1m_rvol_20d"] = np.where(
        pd.to_numeric(features["avg_open_1m_volume_20d"], errors="coerce") > 0,
        pd.to_numeric(features["open_1m_volume"], errors="coerce") / pd.to_numeric(features["avg_open_1m_volume_20d"], errors="coerce"),
        np.nan,
    )
    features["open_5m_rvol_20d"] = np.where(
        pd.to_numeric(features["avg_open_5m_volume_20d"], errors="coerce") > 0,
        pd.to_numeric(features["open_5m_volume"], errors="coerce") / pd.to_numeric(features["avg_open_5m_volume_20d"], errors="coerce"),
        np.nan,
    )
    features["day_volume_rvol_20d"] = np.where(
        pd.to_numeric(features["avg_day_volume_20d"], errors="coerce") > 0,
        pd.to_numeric(features["day_volume"], errors="coerce") / pd.to_numeric(features["avg_day_volume_20d"], errors="coerce"),
        np.nan,
    )

    coverage = expected.copy()
    coverage = coverage.merge(
        daily[["trade_date", "symbol"]].assign(has_daily_bar=True) if not daily.empty else pd.DataFrame(columns=["trade_date", "symbol", "has_daily_bar"]),
        on=["trade_date", "symbol"],
        how="left",
    )
    coverage = coverage.merge(
        intraday_frame[["trade_date", "symbol"]].assign(has_intraday_summary=True) if not intraday_frame.empty else pd.DataFrame(columns=["trade_date", "symbol", "has_intraday_summary"]),
        on=["trade_date", "symbol"],
        how="left",
    )
    coverage = coverage.merge(
        open_window[["trade_date", "symbol", "open_window_second_rows"]] if not open_window.empty else pd.DataFrame(columns=["trade_date", "symbol", "open_window_second_rows"]),
        on=["trade_date", "symbol"],
        how="left",
    )
    coverage = coverage.merge(
        open_window_0800[["trade_date", "symbol", "focus_0800_open_window_second_rows"]] if not open_window_0800.empty else pd.DataFrame(columns=["trade_date", "symbol", "focus_0800_open_window_second_rows"]),
        on=["trade_date", "symbol"],
        how="left",
    )
    coverage = coverage.merge(
        close_window[["trade_date", "symbol", "close_window_second_rows"]] if not close_window.empty else pd.DataFrame(columns=["trade_date", "symbol", "close_window_second_rows"]),
        on=["trade_date", "symbol"],
        how="left",
    )
    coverage["has_daily_bar"] = coverage["has_daily_bar"].where(coverage["has_daily_bar"].notna(), False).astype(bool)
    coverage["has_intraday_summary"] = coverage["has_intraday_summary"].where(coverage["has_intraday_summary"].notna(), False).astype(bool)
    if "open_window_second_rows" not in coverage.columns:
        coverage["open_window_second_rows"] = 0
    if "focus_0800_open_window_second_rows" not in coverage.columns:
        coverage["focus_0800_open_window_second_rows"] = 0
    if "close_window_second_rows" not in coverage.columns:
        coverage["close_window_second_rows"] = 0
    coverage["open_window_second_rows"] = pd.to_numeric(coverage["open_window_second_rows"], errors="coerce").fillna(0).astype(int)
    coverage["focus_0800_open_window_second_rows"] = pd.to_numeric(coverage["focus_0800_open_window_second_rows"], errors="coerce").fillna(0).astype(int)
    coverage["close_window_second_rows"] = pd.to_numeric(coverage["close_window_second_rows"], errors="coerce").fillna(0).astype(int)
    coverage["has_open_window_detail"] = coverage["open_window_second_rows"] > 0
    if "regular_open_second_rows" in coverage.columns:
        coverage["has_open_window_detail"] = coverage["has_open_window_detail"] | (coverage["regular_open_second_rows"] > 0)
    coverage["has_open_window_detail"] = coverage["has_open_window_detail"] | (coverage["focus_0800_open_window_second_rows"] > 0)
    coverage["has_close_window_detail"] = coverage["close_window_second_rows"] > 0
    coverage["exclusion_reason"] = np.select(
        [
            ~coverage["has_daily_bar"],
            coverage["has_daily_bar"] & ~coverage["has_intraday_summary"],
            coverage["has_intraday_summary"] & ~coverage["has_open_window_detail"],
        ],
        [
            "missing_daily_bar",
            "missing_intraday_summary",
            "no_open_window_seconds",
        ],
        default="",
    )
    coverage = coverage.sort_values(["symbol", "trade_date"]).reset_index(drop=True)
    return features, coverage


def _filter_materialized_detail_rows(
    frame: pd.DataFrame,
    *,
    selected_date: str | date | datetime | pd.Timestamp,
    selected_symbol: str,
) -> pd.DataFrame:
    if frame.empty or "trade_date" not in frame.columns or "symbol" not in frame.columns:
        return pd.DataFrame()

    selected_date_value = pd.Timestamp(selected_date).strftime("%Y-%m-%d")
    trade_dates = pd.to_datetime(frame["trade_date"], errors="coerce").dt.strftime("%Y-%m-%d")
    selected_symbol_norm = normalize_symbol_for_databento(selected_symbol)
    if not selected_symbol_norm:
        return pd.DataFrame()
    frame_symbol_norm = frame["symbol"].astype(str).map(normalize_symbol_for_databento)
    return frame[(trade_dates == selected_date_value) & (frame_symbol_norm == selected_symbol_norm)].reset_index(drop=True)


def resolve_selected_detail_tables(
    second_detail_all: pd.DataFrame,
    minute_detail_all: pd.DataFrame,
    *,
    selected_date: str | date | datetime | pd.Timestamp,
    selected_symbol: str,
    fallback_loader: Callable[[], tuple[pd.DataFrame, pd.DataFrame]] | None = None,
    allow_explicit_refetch: bool = False,
) -> tuple[pd.DataFrame, pd.DataFrame, bool]:
    second_detail = _filter_materialized_detail_rows(
        second_detail_all,
        selected_date=selected_date,
        selected_symbol=selected_symbol,
    )
    minute_detail = _filter_materialized_detail_rows(
        minute_detail_all,
        selected_date=selected_date,
        selected_symbol=selected_symbol,
    )

    if not second_detail.empty and not minute_detail.empty:
        return second_detail, minute_detail, False

    if not allow_explicit_refetch or fallback_loader is None:
        return second_detail, minute_detail, False

    fallback_second, fallback_minute = fallback_loader()
    if second_detail.empty and isinstance(fallback_second, pd.DataFrame):
        second_detail = fallback_second.reset_index(drop=True)
    if minute_detail.empty and isinstance(fallback_minute, pd.DataFrame):
        minute_detail = fallback_minute.reset_index(drop=True)
    return second_detail, minute_detail, True


def estimate_databento_costs(
    databento_api_key: str,
    *,
    dataset: str,
    trading_days: list[date],
    display_timezone: str = DEFAULT_DISPLAY_TZ,
    window_start: time | None = None,
    window_end: time | None = None,
    premarket_anchor_et: time = time(8, 0),
) -> pd.DataFrame:
    if not trading_days:
        return pd.DataFrame(columns=["scope", "cost_usd", "billable_size_bytes"])
    client = _make_databento_client(databento_api_key)
    rows: list[dict[str, Any]] = []
    daily_start = trading_days[0] - timedelta(days=14)
    schema_end = _get_schema_available_end(client, dataset, "ohlcv-1d")
    daily_end_exclusive = _daily_request_end_exclusive(trading_days[-1], schema_end)
    try:
        if daily_end_exclusive <= daily_start:
            rows.append({"scope": "daily_ohlcv_1d", "cost_usd": None, "billable_size_bytes": None})
        else:
            rows.append(
                {
                    "scope": "daily_ohlcv_1d",
                    "cost_usd": client.metadata.get_cost(
                        dataset=dataset,
                        start=daily_start.isoformat(),
                        end=daily_end_exclusive.isoformat(),
                        symbols="ALL_SYMBOLS",
                        schema="ohlcv-1d",
                    ),
                    "billable_size_bytes": client.metadata.get_billable_size(
                        dataset=dataset,
                        start=daily_start.isoformat(),
                        end=daily_end_exclusive.isoformat(),
                        symbols="ALL_SYMBOLS",
                        schema="ohlcv-1d",
                    ),
                }
            )
    except Exception:
        rows.append({"scope": "daily_ohlcv_1d", "cost_usd": None, "billable_size_bytes": None})

    intraday_cost = 0.0
    intraday_size = 0
    try:
        for trade_day in trading_days:
            day_ws, day_we = _resolve_window_for_date(trade_day, display_timezone, window_start, window_end)
            window = build_window_definition(
                trade_day,
                display_timezone=display_timezone,
                window_start=day_ws,
                window_end=day_we,
                premarket_anchor_et=premarket_anchor_et,
            )
            intraday_cost += float(
                client.metadata.get_cost(
                    dataset=dataset,
                    start=window.fetch_start_utc.isoformat(),
                    end=_exclusive_ohlcv_1s_end(window.fetch_end_utc).isoformat(),
                    symbols="ALL_SYMBOLS",
                    schema="ohlcv-1s",
                )
            )
            intraday_size += int(
                client.metadata.get_billable_size(
                    dataset=dataset,
                    start=window.fetch_start_utc.isoformat(),
                    end=_exclusive_ohlcv_1s_end(window.fetch_end_utc).isoformat(),
                    symbols="ALL_SYMBOLS",
                    schema="ohlcv-1s",
                )
            )
    except Exception:
        logger.warning("Intraday cost estimate failed.", exc_info=True)
        intraday_cost = float("nan")
        intraday_size = -1
    rows.append(
        {
            "scope": "intraday_ohlcv_1s_total",
            "cost_usd": None if math.isnan(intraday_cost) else intraday_cost,
            "billable_size_bytes": None if intraday_size < 0 else intraday_size,
        }
    )
    return pd.DataFrame(rows)


def default_export_directory() -> Path:
    return Path.home() / "Downloads"


def _safe_iso_from_file(path: Path) -> str | None:
    if not path.exists():
        return None
    return datetime.fromtimestamp(path.stat().st_mtime, tz=UTC).isoformat(timespec="seconds")


def _manifest_candidates(export_dir: Path) -> list[Path]:
    manifests = sorted(export_dir.glob("databento_*_manifest.json"), key=lambda candidate: candidate.stat().st_mtime, reverse=True)
    if not manifests:
        manifests = sorted(export_dir.glob("*_manifest.json"), key=lambda candidate: candidate.stat().st_mtime, reverse=True)
    return manifests


def _load_latest_parseable_manifest(export_dir: Path) -> tuple[Path | None, dict[str, Any]]:
    fallback_path: Path | None = None
    for candidate in _manifest_candidates(export_dir):
        if fallback_path is None:
            fallback_path = candidate
        try:
            return candidate, json.loads(candidate.read_text(encoding="utf-8"))
        except Exception:
            logger.warning("Failed to parse manifest JSON: %s", candidate, exc_info=True)
    return fallback_path, {}


def _latest_manifest_path(export_dir: str | Path | None = None) -> Path | None:
    target_dir = Path(export_dir) if export_dir is not None else default_export_directory()
    if not target_dir.exists():
        return None
    manifests = _manifest_candidates(target_dir)
    return manifests[0] if manifests else None


def build_data_status_result(export_dir: str | Path | None = None, *, stale_after_minutes: int = 60) -> DataStatusResult:
    target_dir = Path(export_dir) if export_dir is not None else default_export_directory()
    manifest_path, manifest = _load_latest_parseable_manifest(target_dir)

    export_generated_at = manifest.get("export_generated_at") or manifest.get("exported_at")
    is_fast_manifest = str(manifest.get("mode") or "") == "preopen_fast_reduced_scope"
    daily_bars_fetched_at = manifest.get("daily_bars_fetched_at") or _safe_iso_from_file(target_dir / EXACT_EXPORT_STATUS_FILES["daily_bars_fetched_at"])
    intraday_fetched_at = manifest.get("intraday_fetched_at")
    premarket_fetched_at = manifest.get("premarket_fetched_at") or _safe_iso_from_file(target_dir / EXACT_EXPORT_STATUS_FILES["premarket_fetched_at"])
    second_detail_fetched_at = manifest.get("second_detail_fetched_at")
    if second_detail_fetched_at is None and not is_fast_manifest:
        second_detail_fetched_at = _safe_iso_from_file(target_dir / EXACT_EXPORT_STATUS_FILES["second_detail_fetched_at"])

    if not intraday_fetched_at:
        intraday_fetched_at = second_detail_fetched_at or premarket_fetched_at
    if not export_generated_at:
        export_generated_at = premarket_fetched_at or intraday_fetched_at or daily_bars_fetched_at

    if not export_generated_at:
        return DataStatusResult(
            export_generated_at=None,
            daily_bars_fetched_at=daily_bars_fetched_at,
            intraday_fetched_at=intraday_fetched_at,
            premarket_fetched_at=premarket_fetched_at,
            second_detail_fetched_at=second_detail_fetched_at,
            dataset=manifest.get("dataset"),
            lookback_days=manifest.get("lookback_days"),
            trade_dates_covered=tuple(str(item) for item in manifest.get("trade_dates_covered", [])),
            is_stale=True,
            staleness_reason="No production export found yet.",
            manifest_path=str(manifest_path) if manifest_path is not None else None,
        )

    try:
        reference_timestamp = pd.Timestamp(export_generated_at)
        if reference_timestamp.tzinfo is None:
            reference_timestamp = reference_timestamp.tz_localize(UTC)
        else:
            reference_timestamp = reference_timestamp.tz_convert(UTC)
    except Exception:
        return DataStatusResult(
            export_generated_at=str(export_generated_at) if export_generated_at else None,
            daily_bars_fetched_at=str(daily_bars_fetched_at) if daily_bars_fetched_at else None,
            intraday_fetched_at=str(intraday_fetched_at) if intraday_fetched_at else None,
            premarket_fetched_at=str(premarket_fetched_at) if premarket_fetched_at else None,
            second_detail_fetched_at=str(second_detail_fetched_at) if second_detail_fetched_at else None,
            dataset=str(manifest.get("dataset")) if manifest.get("dataset") else None,
            lookback_days=int(manifest["lookback_days"]) if manifest.get("lookback_days") is not None and str(manifest.get("lookback_days")).isdigit() else None,
            trade_dates_covered=tuple(str(item) for item in manifest.get("trade_dates_covered", [])),
            is_stale=True,
            staleness_reason="Invalid export timestamp in manifest.",
            manifest_path=str(manifest_path) if manifest_path is not None else None,
        )
    age_minutes = max(0.0, (pd.Timestamp(datetime.now(UTC)) - reference_timestamp).total_seconds() / 60.0)
    is_stale = age_minutes > stale_after_minutes

    return DataStatusResult(
        export_generated_at=str(export_generated_at) if export_generated_at else None,
        daily_bars_fetched_at=str(daily_bars_fetched_at) if daily_bars_fetched_at else None,
        intraday_fetched_at=str(intraday_fetched_at) if intraday_fetched_at else None,
        premarket_fetched_at=str(premarket_fetched_at) if premarket_fetched_at else None,
        second_detail_fetched_at=str(second_detail_fetched_at) if second_detail_fetched_at else None,
        dataset=str(manifest.get("dataset")) if manifest.get("dataset") else None,
        lookback_days=int(manifest["lookback_days"]) if manifest.get("lookback_days") is not None and str(manifest.get("lookback_days")).isdigit() else None,
        trade_dates_covered=tuple(str(item) for item in manifest.get("trade_dates_covered", [])),
        is_stale=is_stale,
        staleness_reason="Fresh" if not is_stale else f"Last successful export is {age_minutes:.0f} minutes old.",
        manifest_path=str(manifest_path) if manifest_path is not None else None,
    )


def build_status_table(status: DataStatusResult) -> pd.DataFrame:
    rows = [
        {"field": "Export generated", "value": status.export_generated_at or "n/a"},
        {"field": "Daily bars fetched", "value": status.daily_bars_fetched_at or "n/a"},
        {"field": "Intraday fetched", "value": status.intraday_fetched_at or "n/a"},
        {"field": "Premarket fetched", "value": status.premarket_fetched_at or "n/a"},
        {"field": "1s open window fetched", "value": status.second_detail_fetched_at or "n/a"},
        {"field": "Dataset", "value": status.dataset or "n/a"},
        {"field": "Lookback days", "value": status.lookback_days if status.lookback_days is not None else "n/a"},
        {"field": "Trade dates covered", "value": ", ".join(status.trade_dates_covered) if status.trade_dates_covered else "n/a"},
        {"field": "Freshness", "value": "stale" if status.is_stale else "fresh"},
        {"field": "Reason", "value": status.staleness_reason},
    ]
    table = pd.DataFrame(rows)
    table["value"] = table["value"].map(str)
    return table


def build_config_table(config_snapshot: dict[str, Any]) -> pd.DataFrame:
    def _format_config_value(value: Any) -> str:
        if isinstance(value, (list, tuple, dict)):
            try:
                return json.dumps(value, ensure_ascii=True, default=str)
            except Exception:
                return str(value)
        return str(value)

    return pd.DataFrame(
        {
            "setting": list(config_snapshot.keys()),
            "value": [_format_config_value(value) for value in config_snapshot.values()],
        }
    )


def _safe_timestamp(value: Any) -> pd.Timestamp | None:
    try:
        ts = pd.Timestamp(value)
    except Exception:
        return None
    if pd.isna(ts):
        return None
    if ts.tzinfo is None:
        return ts.tz_localize(UTC)
    return ts.tz_convert(UTC)


def _safe_number(value: Any) -> float | None:
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    if not math.isfinite(number):
        return None
    return number


def build_entry_checklist_table(
    *,
    status: DataStatusResult,
    selected_row: pd.Series,
    watchlist_table: pd.DataFrame,
    watchlist_config: dict[str, Any] | None = None,
) -> tuple[pd.DataFrame, str, int]:
    def _check_level(is_ok: bool, is_borderline: bool) -> str:
        if is_ok:
            return "erfuellt"
        if is_borderline:
            return "grenzwertig"
        return "nicht_erfuellt"

    export_ts = _safe_timestamp(status.export_generated_at)
    export_today_et = bool(
        export_ts is not None
        and export_ts.tz_convert(US_EASTERN_TZ).date() == datetime.now(US_EASTERN_TZ).date()
    )
    status_ok = (not status.is_stale) and export_today_et
    status_borderline = (not status.is_stale) and (not export_today_et)

    config = watchlist_config or {}
    gap_threshold = float(config.get("min_gap_pct", LONG_DIP_MIN_GAP_PCT))
    pmdv_threshold = float(config.get("min_premarket_dollar_volume", LONG_DIP_MIN_PREMARKET_DOLLAR_VOLUME))

    gap_value = _safe_number(selected_row.get("prev_close_to_premarket_pct"))
    gap_ok = bool(gap_value is not None and gap_value >= gap_threshold)
    gap_borderline = False

    pmdv_value = _safe_number(selected_row.get("premarket_dollar_volume"))
    pmdv_ok = bool(pmdv_value is not None and pmdv_value >= pmdv_threshold)
    pmdv_borderline = False

    early_dip_pct = _safe_number(selected_row.get("early_dip_pct_10s"))
    early_dip_second = _safe_number(selected_row.get("early_dip_second"))
    early_dip_ok = bool(
        early_dip_pct is not None
        and early_dip_second is not None
        and early_dip_pct <= LONG_DIP_ENTRY_EARLY_DIP_MIN_PCT
        and early_dip_second <= LONG_DIP_ENTRY_EARLY_DIP_MAX_SECONDS
    )
    early_dip_borderline = False

    open30_volume = _safe_number(selected_row.get("open_30s_volume"))
    open30_ok = bool(open30_volume is not None and open30_volume >= LONG_DIP_ENTRY_OPEN30_VOLUME_MIN)
    open30_borderline = False

    reclaim_flag_raw = selected_row.get("reclaimed_start_price_within_30s")
    reclaim_flag = bool(reclaim_flag_raw) if reclaim_flag_raw is not None and not pd.isna(reclaim_flag_raw) else False
    reclaim_second = _safe_number(selected_row.get("reclaim_second_30s"))
    reclaim_ok = bool(
        reclaim_flag
        and reclaim_second is not None
        and reclaim_second < LONG_DIP_ENTRY_RECLAIM_MAX_SECONDS
    )
    reclaim_borderline = False

    rows = [
        {
            "check": "1. Status check",
            "erfuellt": status_ok,
            "status": _check_level(status_ok, status_borderline),
            "details": f"fresh={not status.is_stale}, export_heute_et={export_today_et}",
        },
        {
            "check": "2. Gap filter",
            "erfuellt": gap_ok,
            "status": _check_level(gap_ok, gap_borderline),
            "details": f"prev_close_to_premarket_pct={gap_value if gap_value is not None else 'n/a'} (>= {gap_threshold:.1f})",
        },
        {
            "check": "3. Premarket $ volume",
            "erfuellt": pmdv_ok,
            "status": _check_level(pmdv_ok, pmdv_borderline),
            "details": f"premarket_dollar_volume={pmdv_value if pmdv_value is not None else 'n/a'} (>= {pmdv_threshold:,.0f})",
        },
        {
            "check": "4. Early dip <= -0.5%",
            "erfuellt": early_dip_ok,
            "status": _check_level(early_dip_ok, early_dip_borderline),
            "details": (
                f"early_dip_pct_10s={early_dip_pct if early_dip_pct is not None else 'n/a'} | "
                f"early_dip_second={int(early_dip_second) if early_dip_second is not None else 'n/a'} "
                f"(<= {LONG_DIP_ENTRY_EARLY_DIP_MAX_SECONDS}s)"
            ),
        },
        {
            "check": "5. Open30 vol + reclaim",
            "erfuellt": open30_ok and reclaim_ok,
            "status": _check_level(open30_ok and reclaim_ok, open30_borderline or reclaim_borderline),
            "details": (
                f"open_30s_volume={open30_volume if open30_volume is not None else 'n/a'} "
                f"(>= {LONG_DIP_ENTRY_OPEN30_VOLUME_MIN:,.0f}) | "
                f"reclaim={reclaim_flag} @ {int(reclaim_second) if reclaim_second is not None else 'n/a'}s "
                f"(< {LONG_DIP_ENTRY_RECLAIM_MAX_SECONDS}s)"
            ),
        },
    ]

    checklist = pd.DataFrame(rows)
    score = int(checklist["erfuellt"].sum())
    rule_note = (
        f"Regelwerk: Gap >= {gap_threshold:.1f}%, "
        f"Premarket-Dollar-Volumen >= {pmdv_threshold:,.0f}, "
        f"frueher Dip <= {LONG_DIP_ENTRY_EARLY_DIP_MIN_PCT:.1f}% in den ersten {LONG_DIP_ENTRY_EARLY_DIP_MAX_SECONDS}s, "
        f"Open30-Volumen >= {LONG_DIP_ENTRY_OPEN30_VOLUME_MIN:,.0f} und "
        f"Reclaim des Startpreises innerhalb von {LONG_DIP_ENTRY_RECLAIM_MAX_SECONDS}s."
    )
    return checklist, rule_note, score


def resolve_watchlist_display_table(
    *,
    watchlist_result: dict[str, Any],
    view_mode: str,
) -> tuple[pd.DataFrame, str]:
    historical_watchlist_table = watchlist_result["watchlist_table"]
    active_watchlist_table = watchlist_result.get("active_watchlist_table")
    if not isinstance(active_watchlist_table, pd.DataFrame):
        active_watchlist_table = historical_watchlist_table

    use_full_history = view_mode == "Full history"
    display_watchlist_table = historical_watchlist_table if use_full_history else active_watchlist_table
    active_trade_date = watchlist_result.get("trade_date") or "n/a"
    base_caption = (
        f"Watchlist generated at {watchlist_result['generated_at']}. "
        f"Source data fetched at {watchlist_result.get('source_data_fetched_at') or 'n/a'}."
    )

    if use_full_history and len(historical_watchlist_table) != len(active_watchlist_table):
        trade_dates = pd.to_datetime(historical_watchlist_table.get("trade_date"), errors="coerce").dt.date.dropna()
        trade_date_count = int(trade_dates.nunique()) if len(trade_dates) else 0
        caption = (
            f"{base_caption} Showing full history ({len(historical_watchlist_table)} rows across {trade_date_count} trade dates). "
            f"Latest trade date is {active_trade_date}."
        )
    elif len(historical_watchlist_table) != len(active_watchlist_table):
        caption = (
            f"{base_caption} Showing latest trade date {active_trade_date} "
            f"({len(active_watchlist_table)} rows, {len(historical_watchlist_table)} historical rows total)."
        )
    else:
        caption = f"{base_caption} Active trade date: {active_trade_date}."

    return display_watchlist_table, caption


def build_export_basename(
    *,
    prefix: str = "databento_volatility_screener",
    run_timestamp: datetime | None = None,
) -> str:
    stamp = (run_timestamp or datetime.now(UTC)).strftime("%Y%m%d_%H%M%S")
    return f"{prefix}_{stamp}"


def build_run_manifest_frame(manifest: dict[str, Any]) -> pd.DataFrame:
    rows: list[dict[str, str]] = []
    for key, value in manifest.items():
        if isinstance(value, (list, tuple, set)):
            rendered = json.dumps(list(value), ensure_ascii=True)
        elif isinstance(value, dict):
            rendered = json.dumps(value, ensure_ascii=True, sort_keys=True)
        else:
            rendered = "" if value is None else str(value)
        rows.append({"field": str(key), "value": rendered})
    return pd.DataFrame(rows)


def _prepare_frame_for_excel(frame: pd.DataFrame) -> pd.DataFrame:
    sanitized = frame.copy()
    for column in sanitized.columns:
        series = sanitized[column]
        if isinstance(series.dtype, pd.DatetimeTZDtype):
            sanitized[column] = series.dt.tz_localize(None)
        elif is_datetime64_any_dtype(series):
            continue
        elif series.dtype == object:
            sanitized[column] = series.map(
                lambda value: (
                    value.tz_localize(None) if isinstance(value, pd.Timestamp) and value.tzinfo is not None
                    else value.replace(tzinfo=None) if isinstance(value, datetime) and value.tzinfo is not None
                    else value
                )
            )
    return sanitized


def create_excel_workbook(
    summary: pd.DataFrame,
    *,
    minute_detail: pd.DataFrame | None = None,
    second_detail: pd.DataFrame | None = None,
    additional_sheets: dict[str, pd.DataFrame] | None = None,
) -> bytes:
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        _prepare_frame_for_excel(summary).to_excel(writer, sheet_name="summary", index=False)
        if additional_sheets:
            for sheet_name, frame in additional_sheets.items():
                if frame is None or frame.empty:
                    continue
                safe_sheet_name = str(sheet_name)[:31]
                _prepare_frame_for_excel(frame).to_excel(writer, sheet_name=safe_sheet_name, index=False)
        if minute_detail is not None and not minute_detail.empty:
            _prepare_frame_for_excel(minute_detail).to_excel(writer, sheet_name="minute_detail", index=False)
        if second_detail is not None and not second_detail.empty:
            _prepare_frame_for_excel(second_detail).to_excel(writer, sheet_name="second_detail", index=False)

        workbook = writer.book
        for worksheet in workbook.worksheets:
            worksheet.freeze_panes = "A2"
            worksheet.auto_filter.ref = worksheet.dimensions
            header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
            header_font = Font(color="FFFFFF", bold=True)
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            for column_cells in worksheet.columns:
                max_length = max(len(str(cell.value or "")) for cell in column_cells)
                worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_length + 2, 12), 28)

        summary_sheet = workbook["summary"]
        headers = {cell.value: idx + 1 for idx, cell in enumerate(summary_sheet[1])}
        heat_columns = [
            "window_range_pct",
            "realized_vol_pct",
            "window_return_pct",
            "prev_close_to_premarket_pct",
            "premarket_to_open_pct",
            "open_to_current_pct",
        ]
        for col_name in heat_columns:
            col_idx = headers.get(col_name)
            if col_idx is None or summary_sheet.max_row < 2:
                continue
            letter = get_column_letter(col_idx)
            summary_sheet.conditional_formatting.add(
                f"{letter}2:{letter}{summary_sheet.max_row}",
                ColorScaleRule(
                    start_type="num",
                    start_value=-10,
                    start_color="C00000",
                    mid_type="num",
                    mid_value=0,
                    mid_color="FFF2CC",
                    end_type="num",
                    end_value=10,
                    end_color="63BE7B",
                ),
            )
    return output.getvalue()


def _normalize_tradingview_exchange_prefix(value: object) -> str:
    normalized = str(value or "").strip().upper()
    if normalized in {"NYSE ARCA", "NYSEARCA", "ARCA", "NYSE AMERICAN", "NYSEAMERICAN"}:
        return "AMEX"
    return normalized.replace(" ", "")


def _build_tradingview_watchlist_text(frame: pd.DataFrame) -> str:
    if frame.empty or "symbol" not in frame.columns:
        return ""

    exchange_series = frame.get("exchange")
    if exchange_series is None:
        exchange_series = pd.Series("", index=frame.index, dtype=str)

    entries: list[str] = []
    seen: set[str] = set()
    for exchange, symbol in zip(exchange_series, frame["symbol"], strict=False):
        symbol_text = str(symbol).strip().upper() if pd.notna(symbol) else ""
        exchange_text = _normalize_tradingview_exchange_prefix(exchange)
        if not symbol_text or not exchange_text:
            continue
        entry = f"{exchange_text}:{symbol_text}"
        if entry in seen:
            continue
        seen.add(entry)
        entries.append(entry)
    if not entries:
        return ""
    return ",".join(entries) + ","


def _write_tradingview_watchlist_exports(
    export_dir: Path,
    basename: str,
    named_frames: dict[str, pd.DataFrame],
) -> dict[str, Path]:
    created: dict[str, Path] = {}
    export_dir.mkdir(parents=True, exist_ok=True)
    for name, frame in named_frames.items():
        if frame is None or frame.empty:
            continue
        if "symbol" not in frame.columns or "exchange" not in frame.columns:
            continue
        text = _build_tradingview_watchlist_text(frame)
        if not text:
            continue
        path = export_dir / f"{basename}__{name}.txt"
        _write_text_atomic(path, text, encoding="utf-8")
        created[name] = path
    return created


def _write_streamlit_watchlist_txt_exports(export_dir: Path, watchlist_result: dict[str, Any]) -> dict[str, Path]:
    created: dict[str, Path] = {}
    export_dir.mkdir(parents=True, exist_ok=True)

    latest_table = watchlist_result.get("active_watchlist_table")
    if not isinstance(latest_table, pd.DataFrame) or latest_table.empty:
        latest_table = watchlist_result.get("watchlist_table")
    if isinstance(latest_table, pd.DataFrame) and not latest_table.empty:
        latest_text = _build_tradingview_watchlist_text(latest_table)
        if latest_text:
            latest_path = export_dir / "tradingview_watchlist_topn_latest.txt"
            _write_text_atomic(latest_path, latest_text, encoding="utf-8")
            created["txt_topn_latest"] = latest_path

    history_table = watchlist_result.get("watchlist_table")
    if isinstance(history_table, pd.DataFrame) and not history_table.empty:
        history_text = _build_tradingview_watchlist_text(history_table)
        if history_text:
            history_path = export_dir / "tradingview_watchlist_topn_full_history.txt"
            _write_text_atomic(history_path, history_text, encoding="utf-8")
            created["txt_topn_full_history"] = history_path

    return created


def _write_exact_named_export_state(
    export_dir: Path,
    *,
    manifest: dict[str, Any],
    artifact_paths: dict[str, Path],
    source_manifest_path: Path | None = None,
) -> Path:
    export_dir.mkdir(parents=True, exist_ok=True)
    payload = {
        "manifest": manifest,
        "source_manifest_path": str(source_manifest_path) if source_manifest_path is not None else None,
        "artifact_paths": {name: str(path) for name, path in artifact_paths.items()},
    }
    state_path = export_dir / EXACT_NAMED_EXPORT_STATE_FILE
    _write_text_atomic(state_path, json.dumps(payload, indent=2, sort_keys=True, default=str), encoding="utf-8")
    return state_path


def _read_exact_named_export_state(export_dir: Path) -> dict[str, Any]:
    state_path = export_dir / EXACT_NAMED_EXPORT_STATE_FILE
    if not state_path.exists():
        return {}
    try:
        payload = json.loads(state_path.read_text(encoding="utf-8"))
    except Exception:
        logger.warning("Failed to parse exact-named export state: %s", state_path, exc_info=True)
        return {}
    return payload if isinstance(payload, dict) else {}


def _load_ui_runtime_state(export_dir: Path) -> dict[str, Any]:
    payload = _read_exact_named_export_state(export_dir)
    runtime_state = payload.get(UI_RUNTIME_STATE_KEY)
    return runtime_state if isinstance(runtime_state, dict) else {}


def _persist_ui_runtime_state(
    export_dir: Path,
    *,
    refresh_seconds: float | None = None,
    watchlist_seconds: float | None = None,
    action_message: str | None = None,
) -> Path:
    export_dir.mkdir(parents=True, exist_ok=True)
    payload = _read_exact_named_export_state(export_dir)
    runtime_state = payload.get(UI_RUNTIME_STATE_KEY)
    if not isinstance(runtime_state, dict):
        runtime_state = {}
    if refresh_seconds is not None:
        runtime_state["last_refresh_seconds"] = float(refresh_seconds)
        runtime_state["last_refresh_recorded_at"] = datetime.now(UTC).isoformat(timespec="seconds")
    if watchlist_seconds is not None:
        runtime_state["last_watchlist_seconds"] = float(watchlist_seconds)
        runtime_state["last_watchlist_recorded_at"] = datetime.now(UTC).isoformat(timespec="seconds")
    if action_message is not None:
        runtime_state["last_action_message"] = str(action_message)
    payload[UI_RUNTIME_STATE_KEY] = runtime_state
    state_path = export_dir / EXACT_NAMED_EXPORT_STATE_FILE
    _write_text_atomic(state_path, json.dumps(payload, indent=2, sort_keys=True, default=str), encoding="utf-8")
    return state_path


def _watchlist_snapshot_path(export_dir: Path) -> Path:
    return export_dir / WATCHLIST_SNAPSHOT_FILE


def _empty_watchlist_snapshot_history() -> pd.DataFrame:
    return pd.DataFrame(
        columns=[
            "snapshot_at",
            "trade_date",
            "symbol",
            "watchlist_rank",
            "source_data_fetched_at",
            "watchlist_generated_at",
            "trigger",
        ]
    )


def _coerce_watchlist_snapshot_history(frame: pd.DataFrame | None) -> pd.DataFrame:
    if frame is None or frame.empty:
        return _empty_watchlist_snapshot_history()
    history = frame.copy()
    for column in _empty_watchlist_snapshot_history().columns:
        if column not in history.columns:
            history[column] = pd.NA
    history["snapshot_at"] = pd.to_datetime(history["snapshot_at"], errors="coerce", utc=True)
    history["trade_date"] = pd.to_datetime(history["trade_date"], errors="coerce").dt.date
    history["symbol"] = history["symbol"].astype(str).str.strip().str.upper()
    history["watchlist_rank"] = pd.to_numeric(history["watchlist_rank"], errors="coerce")
    history["source_data_fetched_at"] = history["source_data_fetched_at"].astype(str)
    history["watchlist_generated_at"] = history["watchlist_generated_at"].astype(str)
    history["trigger"] = history["trigger"].astype(str)
    history = history.dropna(subset=["snapshot_at", "trade_date", "symbol", "watchlist_rank"])
    if history.empty:
        return _empty_watchlist_snapshot_history()
    return history.sort_values(["snapshot_at", "trade_date", "watchlist_rank", "symbol"]).reset_index(drop=True)


def _format_rank_change_label(previous_rank: Any, current_rank: Any, *, missing_label: str = "new") -> str:
    previous = pd.to_numeric(pd.Series([previous_rank]), errors="coerce").iloc[0]
    current = pd.to_numeric(pd.Series([current_rank]), errors="coerce").iloc[0]
    if pd.isna(current):
        return "n/a"
    if pd.isna(previous):
        return missing_label
    delta = int(previous) - int(current)
    if delta == 0:
        return "flat"
    if delta > 0:
        return f"up {delta}"
    return f"down {abs(delta)}"


def _highlight_rank_change_label(label: Any, delta: Any = None) -> str:
    normalized = str(label or "").strip().lower()
    numeric_delta = pd.to_numeric(pd.Series([delta]), errors="coerce").iloc[0]
    if normalized in {"new", "first"}:
        return normalized.upper()
    if normalized == "flat":
        return "FLAT 0"
    if normalized.startswith("up "):
        amount = normalized.removeprefix("up ").strip()
        return f"UP +{amount}"
    if normalized.startswith("down "):
        amount = normalized.removeprefix("down ").strip()
        return f"DOWN -{amount}"
    if pd.notna(numeric_delta):
        numeric_delta_int = int(numeric_delta)
        if numeric_delta_int > 0:
            return f"UP +{numeric_delta_int}"
        if numeric_delta_int < 0:
            return f"DOWN {numeric_delta_int}"
        return "FLAT 0"
    return str(label or "n/a").upper() or "N/A"


def _format_intraday_reference_time(value: Any, *, display_timezone: str = DEFAULT_DISPLAY_TZ) -> str:
    reference_ts = _safe_timestamp(value)
    if reference_ts is None:
        return "n/a"
    return str(reference_ts.tz_convert(resolve_display_timezone(display_timezone)).strftime("%H:%M:%S %Z"))


def _rank_change_cell_style(label: Any, delta: Any = None) -> str:
    emphasized = _highlight_rank_change_label(label, delta)
    if emphasized.startswith("UP "):
        return "background-color: #dcfce7; color: #166534; font-weight: 700"
    if emphasized.startswith("DOWN "):
        return "background-color: #fee2e2; color: #991b1b; font-weight: 700"
    if emphasized.startswith("FLAT "):
        return "background-color: #e5e7eb; color: #374151; font-weight: 700"
    if emphasized == "NEW":
        return "background-color: #dbeafe; color: #1d4ed8; font-weight: 700"
    if emphasized == "FIRST":
        return "background-color: #fef3c7; color: #92400e; font-weight: 700"
    if emphasized == "N/A":
        return "color: #6b7280"
    return ""


def _rank_delta_cell_style(value: Any) -> str:
    numeric_value = pd.to_numeric(pd.Series([value]), errors="coerce").iloc[0]
    if pd.isna(numeric_value):
        return "color: #6b7280"
    numeric_value = int(numeric_value)
    if numeric_value > 0:
        return "background-color: #dcfce7; color: #166534; font-weight: 700"
    if numeric_value < 0:
        return "background-color: #fee2e2; color: #991b1b; font-weight: 700"
    return "background-color: #e5e7eb; color: #374151; font-weight: 700"


def _build_watchlist_table_style_frame(frame: pd.DataFrame) -> pd.DataFrame:
    styles = pd.DataFrame("", index=frame.index, columns=frame.columns, dtype=object)
    if frame.empty:
        return styles

    if "watchlist_rank_change" in frame.columns:
        watchlist_deltas = frame.get("watchlist_rank_delta", pd.Series(index=frame.index, dtype=float))
        styles["watchlist_rank_change"] = [
            _rank_change_cell_style(label, delta)
            for label, delta in zip(frame["watchlist_rank_change"], watchlist_deltas, strict=False)
        ]
    if "intraday_watchlist_rank_change" in frame.columns:
        intraday_deltas = frame.get("intraday_watchlist_rank_delta", pd.Series(index=frame.index, dtype=float))
        styles["intraday_watchlist_rank_change"] = [
            _rank_change_cell_style(label, delta)
            for label, delta in zip(frame["intraday_watchlist_rank_change"], intraday_deltas, strict=False)
        ]
    if "watchlist_rank_delta" in frame.columns:
        styles["watchlist_rank_delta"] = frame["watchlist_rank_delta"].map(_rank_delta_cell_style)
    if "intraday_watchlist_rank_delta" in frame.columns:
        styles["intraday_watchlist_rank_delta"] = frame["intraday_watchlist_rank_delta"].map(_rank_delta_cell_style)
    return styles


def _build_watchlist_snapshot_panel_frames(
    snapshot_history: pd.DataFrame | None,
    *,
    trade_date: Any,
    active_symbols: list[str] | tuple[str, ...] | None = None,
    display_timezone: str = DEFAULT_DISPLAY_TZ,
    max_snapshots: int = 8,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    summary_columns = ["snapshot_time", "trigger", "symbols", "leader", "top3"]
    history = _coerce_watchlist_snapshot_history(snapshot_history)
    trade_day = pd.to_datetime(trade_date, errors="coerce")
    if pd.isna(trade_day):
        return pd.DataFrame(columns=summary_columns), pd.DataFrame(columns=["symbol"])

    day_history = history.loc[history["trade_date"] == trade_day.date()].copy()
    if day_history.empty:
        return pd.DataFrame(columns=summary_columns), pd.DataFrame(columns=["symbol"])

    unique_snapshots = (
        day_history.loc[:, ["snapshot_at", "trigger"]]
        .drop_duplicates()
        .sort_values("snapshot_at", ascending=False)
        .head(max_snapshots)
        .reset_index(drop=True)
    )
    day_history = day_history.loc[day_history["snapshot_at"].isin(unique_snapshots["snapshot_at"])]

    summary_rows: list[dict[str, Any]] = []
    for snapshot_row in unique_snapshots.itertuples(index=False):
        snapshot_rows = day_history.loc[day_history["snapshot_at"] == snapshot_row.snapshot_at].sort_values(
            ["watchlist_rank", "symbol"]
        )
        symbols = snapshot_rows["symbol"].astype(str).tolist()
        summary_rows.append(
            {
                "snapshot_time": _format_intraday_reference_time(
                    snapshot_row.snapshot_at,
                    display_timezone=display_timezone,
                ),
                "trigger": str(snapshot_row.trigger or "n/a"),
                "symbols": int(snapshot_rows["symbol"].nunique()),
                "leader": symbols[0] if symbols else "n/a",
                "top3": ", ".join(symbols[:3]) if symbols else "n/a",
            }
        )
    summary_frame = pd.DataFrame(summary_rows, columns=summary_columns)

    normalized_symbols = [str(symbol).strip().upper() for symbol in (active_symbols or []) if str(symbol).strip()]
    if not normalized_symbols:
        latest_snapshot_at = unique_snapshots["snapshot_at"].iloc[0]
        normalized_symbols = (
            day_history.loc[day_history["snapshot_at"] == latest_snapshot_at]
            .sort_values(["watchlist_rank", "symbol"])["symbol"]
            .astype(str)
            .head(5)
            .tolist()
        )
    if not normalized_symbols:
        return summary_frame, pd.DataFrame(columns=["symbol"])

    trail_source = day_history.loc[day_history["symbol"].isin(normalized_symbols)].copy()
    if trail_source.empty:
        return summary_frame, pd.DataFrame(columns=["symbol"])

    snapshot_labels = [
        _format_intraday_reference_time(snapshot_at, display_timezone=display_timezone)
        for snapshot_at in unique_snapshots.sort_values("snapshot_at")["snapshot_at"].tolist()
    ]
    trail_source["snapshot_time"] = trail_source["snapshot_at"].map(
        lambda value: _format_intraday_reference_time(value, display_timezone=display_timezone)
    )
    trail_frame = (
        trail_source.pivot_table(index="symbol", columns="snapshot_time", values="watchlist_rank", aggfunc="last")
        .reindex(index=normalized_symbols)
        .reindex(columns=snapshot_labels)
        .reset_index()
    )
    return summary_frame, trail_frame


def _build_watchlist_snapshot_frame(watchlist_result: dict[str, Any]) -> pd.DataFrame:
    active_table = watchlist_result.get("active_watchlist_table")
    if not isinstance(active_table, pd.DataFrame) or active_table.empty:
        active_table = watchlist_result.get("watchlist_table")
    if not isinstance(active_table, pd.DataFrame) or active_table.empty:
        return _empty_watchlist_snapshot_history()
    if not {"trade_date", "symbol", "watchlist_rank"}.issubset(active_table.columns):
        return _empty_watchlist_snapshot_history()

    snapshot_at = pd.to_datetime(watchlist_result.get("generated_at"), errors="coerce", utc=True)
    if pd.isna(snapshot_at):
        snapshot_at = pd.Timestamp(datetime.now(UTC))

    snapshot = active_table.loc[:, ["trade_date", "symbol", "watchlist_rank"]].copy()
    snapshot["trade_date"] = pd.to_datetime(snapshot["trade_date"], errors="coerce").dt.date
    snapshot["symbol"] = snapshot["symbol"].astype(str).str.strip().str.upper()
    snapshot["watchlist_rank"] = pd.to_numeric(snapshot["watchlist_rank"], errors="coerce")
    snapshot = snapshot.dropna(subset=["trade_date", "symbol", "watchlist_rank"])
    if snapshot.empty:
        return _empty_watchlist_snapshot_history()

    snapshot["snapshot_at"] = snapshot_at
    snapshot["source_data_fetched_at"] = str(watchlist_result.get("source_data_fetched_at") or "")
    snapshot["watchlist_generated_at"] = str(watchlist_result.get("generated_at") or snapshot_at.isoformat())
    snapshot["trigger"] = ""
    return snapshot[
        [
            "snapshot_at",
            "trade_date",
            "symbol",
            "watchlist_rank",
            "source_data_fetched_at",
            "watchlist_generated_at",
            "trigger",
        ]
    ].sort_values(["trade_date", "watchlist_rank", "symbol"]).reset_index(drop=True)


def _load_watchlist_snapshot_history(export_dir: Path) -> pd.DataFrame:
    history = _read_cached_frame(_watchlist_snapshot_path(export_dir))
    return _coerce_watchlist_snapshot_history(history)


def _persist_watchlist_snapshot(
    export_dir: Path,
    watchlist_result: dict[str, Any],
    *,
    trigger: str,
) -> pd.DataFrame:
    snapshot = _build_watchlist_snapshot_frame(watchlist_result)
    history_path = _watchlist_snapshot_path(export_dir)
    history = _load_watchlist_snapshot_history(export_dir)
    if snapshot.empty:
        return history
    snapshot["trigger"] = str(trigger)
    if history.empty:
        history = snapshot.copy()
    else:
        history = pd.concat([history, snapshot], ignore_index=True)
    history = history.drop_duplicates(subset=["snapshot_at", "trade_date", "symbol"], keep="last")
    history = history.sort_values(["snapshot_at", "trade_date", "watchlist_rank", "symbol"]).reset_index(drop=True)
    _write_cached_frame(history_path, history)
    return history


def _augment_watchlist_result_with_intraday_context(
    watchlist_result: dict[str, Any],
    snapshot_history: pd.DataFrame | None,
) -> dict[str, Any]:
    current_snapshot = _build_watchlist_snapshot_frame(watchlist_result)
    if current_snapshot.empty:
        return watchlist_result

    history = _coerce_watchlist_snapshot_history(snapshot_history)
    current_snapshot_at = current_snapshot["snapshot_at"].iloc[0]
    current_trade_date = current_snapshot["trade_date"].iloc[0]
    previous_rows = history.loc[
        (history["trade_date"] == current_trade_date) & (history["snapshot_at"] < current_snapshot_at)
    ].copy()
    if previous_rows.empty:
        previous_by_symbol = pd.DataFrame(columns=["symbol", "previous_intraday_watchlist_rank", "intraday_rank_reference_at"])
    else:
        previous_by_symbol = previous_rows.sort_values(["symbol", "snapshot_at"]).drop_duplicates(subset=["symbol"], keep="last")
        previous_by_symbol = previous_by_symbol.rename(
            columns={
                "watchlist_rank": "previous_intraday_watchlist_rank",
                "snapshot_at": "intraday_rank_reference_at",
            }
        )[["symbol", "previous_intraday_watchlist_rank", "intraday_rank_reference_at"]]

    current_context = current_snapshot.merge(previous_by_symbol, on="symbol", how="left")
    current_context["intraday_watchlist_rank_delta"] = (
        pd.to_numeric(current_context["previous_intraday_watchlist_rank"], errors="coerce")
        - pd.to_numeric(current_context["watchlist_rank"], errors="coerce")
    )
    current_context["intraday_watchlist_rank_change"] = [
        _format_rank_change_label(previous_rank, current_rank, missing_label="first")
        for previous_rank, current_rank in zip(
            current_context["previous_intraday_watchlist_rank"],
            current_context["watchlist_rank"],
            strict=False,
        )
    ]
    current_context["intraday_rank_reference_at"] = current_context["intraday_rank_reference_at"].map(
        lambda value: value.isoformat() if isinstance(value, pd.Timestamp) and pd.notna(value) else None
    )
    current_context = current_context[
        [
            "trade_date",
            "symbol",
            "previous_intraday_watchlist_rank",
            "intraday_watchlist_rank_delta",
            "intraday_watchlist_rank_change",
            "intraday_rank_reference_at",
        ]
    ]

    augmented = dict(watchlist_result)
    context_columns = [
        "previous_intraday_watchlist_rank",
        "intraday_watchlist_rank_delta",
        "intraday_watchlist_rank_change",
        "intraday_rank_reference_at",
    ]
    for table_key in ("active_watchlist_table", "watchlist_table"):
        table = augmented.get(table_key)
        if not isinstance(table, pd.DataFrame) or table.empty:
            continue
        merge_table = table.copy()
        if "trade_date" in merge_table.columns:
            merge_table["trade_date"] = pd.to_datetime(merge_table["trade_date"], errors="coerce").dt.date
        for column in context_columns:
            if column in merge_table.columns:
                merge_table = merge_table.drop(columns=column)
        augmented[table_key] = merge_table.merge(current_context, on=["trade_date", "symbol"], how="left")
    return augmented


def _numeric_series_or_nan(frame: pd.DataFrame, column: str) -> pd.Series:
    if column in frame.columns:
        return pd.to_numeric(frame[column], errors="coerce")
    return pd.Series(np.nan, index=frame.index, dtype=float)


def _open_pattern_metric_columns_for_view(focus_view: str) -> tuple[str, str, str]:
    if focus_view == "09:30 only":
        return (
            "focus_0930_open_30s_volume",
            "focus_0930_early_dip_pct_10s",
            "focus_0930_reclaim_second_30s",
        )
    if focus_view == "08:00 only":
        return (
            "focus_0800_open_30s_volume",
            "focus_0800_early_dip_pct_10s",
            "focus_0800_reclaim_second_30s",
        )
    if focus_view == "04:00 only":
        return (
            "focus_0400_open_30s_volume",
            "focus_0400_early_dip_pct_10s",
            "focus_0400_reclaim_second_30s",
        )
    return ("open_30s_volume", "early_dip_pct_10s", "reclaim_second_30s")


def _build_open_pattern_status_series(frame: pd.DataFrame, focus_view: str) -> pd.Series:
    open_30s_col, early_dip_col, reclaim_col = _open_pattern_metric_columns_for_view(focus_view)
    open_pattern_missing = (
        _numeric_series_or_nan(frame, open_30s_col).isna()
        & _numeric_series_or_nan(frame, early_dip_col).isna()
        & _numeric_series_or_nan(frame, reclaim_col).isna()
    )
    if focus_view == "All (04:00 + 08:00 + 09:30)":
        available_label = "available via >=1 focus window"
        missing_label = "missing across all focus windows"
    else:
        focus_label = focus_view.replace(" only", "")
        available_label = f"available at {focus_label}"
        missing_label = f"missing at {focus_label}"
    return pd.Series(
        np.where(open_pattern_missing, missing_label, available_label),
        index=frame.index,
        dtype=object,
    )


def _build_focus_window_coverage_series(frame: pd.DataFrame) -> pd.Series:
    has_any_window_rows = any(
        column in frame.columns
        for column in (
            "focus_0400_open_window_second_rows",
            "focus_0800_open_window_second_rows",
            "focus_0930_open_window_second_rows",
            "open_window_second_rows",
        )
    )
    if not has_any_window_rows:
        return pd.Series("unavailable", index=frame.index, dtype=object)
    has_0400 = _numeric_series_or_nan(frame, "focus_0400_open_window_second_rows").fillna(0) > 0
    has_0800 = _numeric_series_or_nan(frame, "focus_0800_open_window_second_rows").fillna(0) > 0
    if "focus_0930_open_window_second_rows" in frame.columns:
        has_0930 = _numeric_series_or_nan(frame, "focus_0930_open_window_second_rows").fillna(0) > 0
    else:
        has_0930 = _numeric_series_or_nan(frame, "open_window_second_rows").fillna(0) > 0
    return pd.Series(
        np.select(
            [
                has_0400 & has_0800 & has_0930,
                has_0400 & has_0800,
                has_0400 & has_0930,
                has_0800 & has_0930,
                has_0400,
                has_0800,
                has_0930,
            ],
            [
                "04:00 + 08:00 + 09:30",
                "04:00 + 08:00",
                "04:00 + 09:30",
                "08:00 + 09:30",
                "04:00",
                "08:00",
                "09:30",
            ],
            default="none",
        ),
        index=frame.index,
        dtype=object,
    )


def _streamlit_button_compat(button_callable, label: str, **kwargs: Any):
    try:
        return button_callable(label, width="stretch", **kwargs)
    except TypeError:
        return button_callable(label, use_container_width=True, **kwargs)


def _streamlit_dataframe_compat(dataframe_callable, data: Any, **kwargs: Any):
    try:
        return dataframe_callable(data, width="stretch", **kwargs)
    except TypeError:
        return dataframe_callable(data, use_container_width=True, **kwargs)


def _resolve_watchlist_snapshot_trigger(*, generate_watchlist: bool, fast_pipeline: bool, fast_refresh: bool) -> str:
    if generate_watchlist:
        return "generate_watchlist"
    if fast_pipeline:
        return "fast_pipeline"
    if fast_refresh:
        return "fast_refresh_auto_generate"
    return "auto_load"


def _run_full_history_refresh_with_status(*, status_container: Any, run_pipeline: Callable[[], None]) -> None:
    try:
        run_pipeline()
    except Exception:
        status_container.update(label="Full history refresh: failed.", state="error", expanded=True)
        raise
    else:
        status_container.update(label="Full history refresh: complete.", state="complete", expanded=False)


def _format_reclaim_status_series(frame: pd.DataFrame, column: str = "reclaimed_start_price_within_30s") -> pd.Series:
    if column not in frame.columns:
        return pd.Series("n/a", index=frame.index, dtype=object)
    values = frame[column]
    bool_values = pd.Series(values, dtype="boolean").fillna(False).astype(bool)
    return pd.Series(
        np.where(values.isna(), "n/a", np.where(bool_values, "yes", "no")),
        index=frame.index,
        dtype=object,
    )


def export_run_artifacts(
    *,
    export_dir: str | Path | None,
    basename: str,
    summary: pd.DataFrame,
    universe: pd.DataFrame,
    daily_bars: pd.DataFrame,
    intraday: pd.DataFrame,
    ranked: pd.DataFrame,
    minute_detail: pd.DataFrame | None = None,
    second_detail: pd.DataFrame | None = None,
    minute_detail_all: pd.DataFrame | None = None,
    second_detail_all: pd.DataFrame | None = None,
    additional_sheets: dict[str, pd.DataFrame] | None = None,
    additional_parquet_targets: dict[str, pd.DataFrame] | None = None,
    cost_estimate: pd.DataFrame | None = None,
    unsupported_symbols: list[str] | None = None,
    manifest: dict[str, Any] | None = None,
) -> dict[str, Path]:
    target_dir = Path(export_dir) if export_dir is not None else default_export_directory()
    target_dir.mkdir(parents=True, exist_ok=True)

    manifest_payload = dict(manifest or {})
    manifest_payload.setdefault("basename", basename)
    manifest_payload.setdefault("exported_at", datetime.now(UTC).isoformat(timespec="seconds"))

    workbook_sheets: dict[str, pd.DataFrame] = {
        "manifest": build_run_manifest_frame(manifest_payload),
        "universe": universe,
        "daily_bars": daily_bars,
        "intraday_all": intraday,
        "ranked": ranked,
    }
    if additional_sheets:
        workbook_sheets.update(additional_sheets)
    if cost_estimate is not None and not cost_estimate.empty:
        workbook_sheets["cost_estimate"] = cost_estimate
    if unsupported_symbols:
        workbook_sheets["unsupported_symbols"] = pd.DataFrame({"symbol": unsupported_symbols})

    excel_path = target_dir / f"{basename}.xlsx"
    _write_bytes_atomic(
        excel_path,
        create_excel_workbook(
            summary,
            minute_detail=minute_detail,
            second_detail=second_detail,
            additional_sheets=workbook_sheets,
        ),
    )

    parquet_targets: dict[str, pd.DataFrame] = {
        "summary": summary,
        "universe": universe,
        "daily_bars": daily_bars,
        "intraday": intraday,
        "ranked": ranked,
    }
    if minute_detail is not None and not minute_detail.empty:
        parquet_targets["minute_detail"] = minute_detail
    if second_detail is not None and not second_detail.empty:
        parquet_targets["second_detail"] = second_detail
    if minute_detail_all is not None and not minute_detail_all.empty:
        parquet_targets["minute_detail_all"] = minute_detail_all
    if second_detail_all is not None and not second_detail_all.empty:
        parquet_targets["second_detail_all"] = second_detail_all
    if cost_estimate is not None and not cost_estimate.empty:
        parquet_targets["cost_estimate"] = cost_estimate
    if unsupported_symbols:
        parquet_targets["unsupported_symbols"] = pd.DataFrame({"symbol": unsupported_symbols})
    if additional_parquet_targets:
        parquet_targets.update(additional_parquet_targets)

    created_paths: dict[str, Path] = {
        "excel": excel_path,
    }
    for name, frame in parquet_targets.items():
        parquet_path = target_dir / f"{basename}__{name}.parquet"
        _write_parquet_atomic(parquet_path, frame)
        created_paths[f"parquet_{name}"] = parquet_path
    txt_paths = _write_tradingview_watchlist_exports(target_dir, basename, parquet_targets)
    for name, path in txt_paths.items():
        created_paths[f"txt_{name}"] = path

    manifest_path = target_dir / f"{basename}_manifest.json"
    _write_text_atomic(manifest_path, json.dumps(manifest_payload, indent=2, sort_keys=True, default=str), encoding="utf-8")
    created_paths["manifest"] = manifest_path
    return created_paths


def run_streamlit_app() -> None:
    from dataclasses import replace
    import os
    import streamlit as st
    from datetime import UTC, datetime
    from dotenv import load_dotenv
    from pathlib import Path

    from scripts.bullish_quality_config import (
        BULLISH_QUALITY_SCORE_PROFILES,
        DEFAULT_BULLISH_QUALITY_SCORE_PROFILE,
        build_default_bullish_quality_config,
        normalize_bullish_quality_score_profile,
    )
    from scripts.databento_preopen_fast import run_preopen_fast_refresh
    from scripts.databento_production_export import run_production_export_pipeline
    from scripts.generate_bullish_quality_scanner import generate_bullish_quality_scanner_result
    from scripts.generate_databento_watchlist import LongDipConfig, generate_watchlist_result
    from strategy_config import LONG_DIP_DEFAULTS

    repo_env_path = Path(__file__).resolve().parent / ".env"
    load_dotenv(repo_env_path, override=True)
    st.set_page_config(page_title="Databento Volatility Screener", layout="wide")
    raw_default_top_n = LONG_DIP_DEFAULTS.get("top_n", 5)
    default_top_n: int = raw_default_top_n if isinstance(raw_default_top_n, int) else 5

    st.title("Databento Volatility Screener")
    st.caption("Single point of view for data freshness, pipeline actions, top-N watchlist and per-entry strategy details.")

    if "dvs_databento_api_key" not in st.session_state:
        st.session_state["dvs_databento_api_key"] = os.getenv("DATABENTO_API_KEY", "")
    if "dvs_fmp_api_key" not in st.session_state:
        st.session_state["dvs_fmp_api_key"] = os.getenv("FMP_API_KEY", "")
    if "dvs_bullish_score_profile" not in st.session_state:
        st.session_state["dvs_bullish_score_profile"] = DEFAULT_BULLISH_QUALITY_SCORE_PROFILE

    with st.sidebar:
        scanner_mode = st.radio("Scanner mode", options=["Long-Dip Watchlist", "Bullish-Quality Scanner"], index=0)
        st.text_input(
            "Databento API Key",
            type="password",
            key="dvs_databento_api_key",
        )
        st.text_input(
            "FMP API Key (optional, free tier fallback)",
            type="password",
            key="dvs_fmp_api_key",
            help="Optional. Used as fallback universe source when Nasdaq Trader is unavailable. Free tier is sufficient.",
        )
        st.caption("API keys entered here stay in the current Streamlit session and are not written back to .env.")
        databento_api_key = str(st.session_state.get("dvs_databento_api_key", "")).strip()
        fmp_api_key = str(st.session_state.get("dvs_fmp_api_key", "")).strip()
        export_dir = st.text_input("Export directory", value=str(default_export_directory()))
        dataset = st.text_input("Databento dataset", value=os.getenv("DATABENTO_DATASET", "DBEQ.BASIC"))
        lookback_days = st.number_input("Trading days", min_value=1, max_value=90, value=30)
        top_n = st.number_input("Top N watchlist", min_value=1, max_value=25, value=default_top_n)
        bullish_score_profile = st.selectbox(
            "Bullish score profile",
            options=list(BULLISH_QUALITY_SCORE_PROFILES),
            key="dvs_bullish_score_profile",
            help="Controls how strongly structure influences Bullish-Quality window scoring and exported premarket window artifacts.",
        )
        fast_scope_days = st.number_input("Fast scope days override (0 = auto)", min_value=0, max_value=30, value=0)
        force_refresh = st.checkbox("Force refresh", value=False)

    if "dvs_run_logs" not in st.session_state:
        st.session_state["dvs_run_logs"] = []
    persisted_runtime_state = _load_ui_runtime_state(Path(export_dir).expanduser())
    if "dvs_last_refresh_seconds" not in st.session_state:
        persisted_refresh_seconds = persisted_runtime_state.get("last_refresh_seconds")
        st.session_state["dvs_last_refresh_seconds"] = float(persisted_refresh_seconds) if persisted_refresh_seconds is not None else None
    if "dvs_last_watchlist_seconds" not in st.session_state:
        persisted_watchlist_seconds = persisted_runtime_state.get("last_watchlist_seconds")
        st.session_state["dvs_last_watchlist_seconds"] = float(persisted_watchlist_seconds) if persisted_watchlist_seconds is not None else None
    if "dvs_last_action_message" not in st.session_state:
        st.session_state["dvs_last_action_message"] = str(persisted_runtime_state.get("last_action_message") or "")

    def add_log(message: str) -> None:
        timestamp = datetime.now(UTC).isoformat(timespec="seconds")
        st.session_state["dvs_run_logs"] = [*st.session_state["dvs_run_logs"], f"{timestamp} {message}"][-50:]

    def _truncate_table_for_ui(
        frame: pd.DataFrame,
        *,
        label: str,
        max_rows: int = 1200,
        max_cols: int = 80,
    ) -> pd.DataFrame:
        if not isinstance(frame, pd.DataFrame):
            return frame
        out = frame
        clipped_cols = False
        clipped_rows = False
        if out.shape[1] > max_cols:
            out = out.iloc[:, :max_cols].copy()
            clipped_cols = True
        if len(out) > max_rows:
            out = out.head(max_rows).copy()
            clipped_rows = True
        if clipped_rows or clipped_cols:
            details = []
            if clipped_rows:
                details.append(f"rows {len(frame):,}->{len(out):,}")
            if clipped_cols:
                details.append(f"cols {frame.shape[1]:,}->{out.shape[1]:,}")
            st.info(f"{label}: showing truncated table ({', '.join(details)}) to keep UI responsive.")
        return out

    status = build_data_status_result(export_dir)
    is_long_dip_mode = scanner_mode == "Long-Dip Watchlist"
    if is_long_dip_mode:
        config_snapshot: dict[str, Any] = dict(LONG_DIP_DEFAULTS)
        config_snapshot["top_n"] = int(top_n)
        config_snapshot["fast_scope_days_override"] = int(fast_scope_days)
        config_snapshot["bullish_score_profile"] = str(bullish_score_profile)
        watchlist_cfg = LongDipConfig(top_n=int(top_n))
        screen_result = st.session_state.get("dvs_watchlist_result")
        snapshot_history = _load_watchlist_snapshot_history(Path(export_dir).expanduser())
        if isinstance(screen_result, dict):
            screen_result = _augment_watchlist_result_with_intraday_context(screen_result, snapshot_history)
            st.session_state["dvs_watchlist_result"] = screen_result
    else:
        bullish_cfg = replace(build_default_bullish_quality_config(score_profile=str(bullish_score_profile)), top_n=int(top_n))
        config_snapshot = {**bullish_cfg.__dict__, "fast_scope_days_override": int(fast_scope_days)}
        screen_result = st.session_state.get("dvs_bullish_quality_result")
        snapshot_history = pd.DataFrame()

    today = datetime.now(US_EASTERN_TZ).date()
    berlin_tz = resolve_display_timezone("Europe/Berlin")

    focus_0800_berlin = datetime.combine(today, time(8, 0), tzinfo=US_EASTERN_TZ).astimezone(berlin_tz)
    focus_0400_berlin = datetime.combine(today, time(4, 0), tzinfo=US_EASTERN_TZ).astimezone(berlin_tz)

    # For the 04:00 ET premarket anchor, recommend windows after anchor so data is expected to exist.
    run0400_start = (focus_0400_berlin + timedelta(minutes=1)).time()
    run0400_end = (focus_0400_berlin + timedelta(minutes=4)).time()
    ref0400_start = (focus_0400_berlin + timedelta(minutes=6)).time()
    ref0400_end = (focus_0400_berlin + timedelta(minutes=8)).time()
    run0800_start = (focus_0800_berlin - timedelta(minutes=7)).time()
    run0800_end = (focus_0800_berlin - timedelta(minutes=4)).time()
    ref0800_start = (focus_0800_berlin - timedelta(minutes=2)).time()
    ref0800_end = (focus_0800_berlin - timedelta(minutes=1)).time()

    run_start, _ = compute_market_relative_window(today, "Europe/Berlin", pre_open_minutes=7, post_open_minutes=0)
    _, run_end = compute_market_relative_window(today, "Europe/Berlin", pre_open_minutes=4, post_open_minutes=0)
    ref_start, _ = compute_market_relative_window(today, "Europe/Berlin", pre_open_minutes=2, post_open_minutes=0)
    _, ref_end = compute_market_relative_window(today, "Europe/Berlin", pre_open_minutes=1, post_open_minutes=0)
    st.info(
        "Recommended start windows (not runtime): "
        f"04:00 ET focus run {run0400_start.strftime('%H:%M')}-{run0400_end.strftime('%H:%M')} Europe/Berlin, "
        f"optional refresh {ref0400_start.strftime('%H:%M')}-{ref0400_end.strftime('%H:%M')} Europe/Berlin; "
        f"08:00 ET focus run {run0800_start.strftime('%H:%M')}-{run0800_end.strftime('%H:%M')} Europe/Berlin, "
        f"optional refresh {ref0800_start.strftime('%H:%M')}-{ref0800_end.strftime('%H:%M')} Europe/Berlin; "
        f"09:30 ET focus run {run_start.strftime('%H:%M')}-{run_end.strftime('%H:%M')} Europe/Berlin, "
        f"optional refresh {ref_start.strftime('%H:%M')}-{ref_end.strftime('%H:%M')} Europe/Berlin."
    )

    status_cols = st.columns(4)
    status_cols[0].metric("Dataset", status.dataset or dataset)
    status_cols[1].metric("Freshness", "stale" if status.is_stale else "fresh")
    status_cols[2].metric("Export generated", status.export_generated_at or "n/a")
    status_cols[3].metric("Premarket fetched", status.premarket_fetched_at or "n/a")

    runtime_cols = st.columns(2)
    runtime_cols[0].metric(
        "Last Refresh Runtime",
        f"{st.session_state['dvs_last_refresh_seconds']:.1f}s" if st.session_state["dvs_last_refresh_seconds"] is not None else "n/a",
    )
    runtime_cols[1].metric(
        "Last Ranking Runtime",
        f"{st.session_state['dvs_last_watchlist_seconds']:.1f}s" if st.session_state["dvs_last_watchlist_seconds"] is not None else "n/a",
    )

    latest_fast_manifest = None
    if status.manifest_path:
        try:
            latest_fast_manifest = json.loads(Path(status.manifest_path).read_text(encoding="utf-8"))
        except Exception:
            logger.warning("Failed to parse manifest JSON: %s", status.manifest_path, exc_info=True)
            latest_fast_manifest = None

    if latest_fast_manifest and latest_fast_manifest.get("mode") == "preopen_fast_reduced_scope":
        fast_scope_cols = st.columns(3)
        fast_scope_cols[0].metric("Fast Scope Days", latest_fast_manifest.get("scope_days", "n/a"))
        fast_scope_cols[1].metric("Fast Scope Symbols", latest_fast_manifest.get("scope_symbol_count", "n/a"))
        fast_scope_cols[2].metric("Fast Scope Mode", latest_fast_manifest.get("scope_days_mode", "n/a"))

    action_cols = st.columns(4)
    fast_refresh = _streamlit_button_compat(action_cols[0].button, "Fast Pre-Open Refresh")
    full_refresh = _streamlit_button_compat(action_cols[1].button, "Full History Refresh")
    generate_watchlist = _streamlit_button_compat(action_cols[2].button, "Generate Watchlist" if is_long_dip_mode else "Generate Scanner")
    fast_pipeline = _streamlit_button_compat(action_cols[3].button, "Fast Pre-Open Pipeline", type="primary")
    if st.session_state["dvs_last_action_message"]:
        st.info(st.session_state["dvs_last_action_message"])
    st.caption(
        "Operational model: run Full History outside the pre-open window to rebuild the 30-day full-universe baseline and historical selected_top20pct symbol-days. "
        "Run Fast Pre-Open Refresh near the open to reuse that baseline with a reduced current premarket scope. Full History refresh does not require an immediate watchlist rebuild."
    )

    pipeline_refresh_ok = False
    if fast_refresh or fast_pipeline or full_refresh:
        selected_action = "Fast pre-open refresh" if fast_refresh else "Fast pre-open pipeline" if fast_pipeline else "Full history refresh"
        add_log(f"Action triggered: {selected_action}.")
        st.session_state["dvs_last_action_message"] = f"Running: {selected_action}..."
        if not databento_api_key:
            st.error("Databento API key is required for the data pipeline.")
            st.session_state["dvs_last_action_message"] = "Blocked: Databento API key missing."
        else:
            try:
                refresh_started = time_module.perf_counter()
                if fast_refresh or fast_pipeline:
                    _fast_status_container = st.status("Fast pre-open refresh: starting...", expanded=True)
                    _fast_progress_bar = st.progress(0, text="Fast refresh 0/5: waiting...")
                    _fast_progress_pct = 0
                    _fast_progress_step = 0
                    _fast_progress_total = 5
                    _fast_eta_smooth_seconds = None
                    def _fast_pipeline_progress(msg: str) -> None:
                        nonlocal _fast_progress_pct
                        nonlocal _fast_progress_step
                        nonlocal _fast_progress_total
                        nonlocal _fast_eta_smooth_seconds
                        _msg = str(msg)
                        _step_match = re.search(r"Fast refresh\s+(\d+)\s*/\s*(\d+)", _msg, flags=re.IGNORECASE)
                        if _step_match:
                            _step = max(0, int(_step_match.group(1)))
                            _step_total = max(1, int(_step_match.group(2)))
                            _fast_progress_step = _step
                            _fast_progress_total = _step_total
                            _fast_progress_pct = max(_fast_progress_pct, min(100, int(round((_step / _step_total) * 100))))
                        if "done" in _msg.lower():
                            _fast_progress_pct = 100
                        _elapsed = max(0.0, time_module.perf_counter() - refresh_started)
                        _eta_text = ""
                        if 0 < _fast_progress_step < _fast_progress_total:
                            _remaining_steps = _fast_progress_total - _fast_progress_step
                            _raw_eta_seconds = (_elapsed / max(1, _fast_progress_step)) * _remaining_steps
                            if _fast_eta_smooth_seconds is None:
                                _fast_eta_smooth_seconds = _raw_eta_seconds
                            else:
                                _fast_eta_smooth_seconds = (_fast_eta_smooth_seconds * 0.75) + (_raw_eta_seconds * 0.25)
                            _eta_seconds = int(round(max(0.0, _fast_eta_smooth_seconds)))
                            _eta_text = f" | ETA ~{_eta_seconds}s"
                        _progress_text = f"Step {_fast_progress_step}/{_fast_progress_total} ({_fast_progress_pct}%)" + _eta_text
                        _fast_progress_bar.progress(_fast_progress_pct, text=f"{_progress_text} - {_msg}")
                        _fast_status_container.update(label=msg)
                        _fast_status_container.write(msg)
                    _fast_result = None
                    try:
                        _fast_result = run_preopen_fast_refresh(
                            databento_api_key=databento_api_key,
                            dataset=dataset,
                            export_dir=Path(export_dir).expanduser(),
                            bundle=Path(export_dir).expanduser(),
                            scope_days=None if int(fast_scope_days) <= 0 else int(fast_scope_days),
                            bullish_score_profile=str(bullish_score_profile),
                            progress_callback=_fast_pipeline_progress,
                        )
                    except Exception:
                        _fast_status_container.update(label="Fast pre-open refresh: failed.", state="error", expanded=True)
                        raise
                    else:
                        _fast_progress_bar.progress(100, text="Step 5/5 (100%) - Fast refresh complete.")
                        _fast_status_container.update(label="Fast pre-open refresh: complete.", state="complete", expanded=True)
                    for _uw in (_fast_result or {}).get("user_warnings") or []:
                        st.warning(_uw)
                        add_log(f"WARNING: {_uw}")
                    refresh_label = "Fast pre-open"
                else:
                    _status_container = st.status("Full history refresh: starting pipeline...", expanded=True)
                    def _pipeline_progress(msg: str) -> None:
                        _status_container.update(label=msg)
                        _status_container.write(msg)
                    def _run_full_pipeline() -> None:
                        run_production_export_pipeline(
                            databento_api_key=databento_api_key,
                            fmp_api_key=fmp_api_key,
                            dataset=dataset,
                            lookback_days=int(lookback_days),
                            cache_dir=Path(__file__).resolve().parent / "artifacts" / "databento_volatility_cache",
                            export_dir=Path(export_dir).expanduser(),
                            use_file_cache=True,
                            force_refresh=force_refresh,
                            second_detail_scope="full_universe",
                            bullish_score_profile=str(bullish_score_profile),
                            progress_callback=_pipeline_progress,
                        )
                    _run_full_history_refresh_with_status(status_container=_status_container, run_pipeline=_run_full_pipeline)
                    refresh_label = "Full history"
            except Exception as exc:
                add_log(f"Pipeline refresh failed: {type(exc).__name__}: {exc}")
                st.error(f"Pipeline refresh failed: {type(exc).__name__}: {exc}")
                st.session_state["dvs_last_action_message"] = f"Failed: {selected_action} ({type(exc).__name__})"
            else:
                refresh_seconds = time_module.perf_counter() - refresh_started
                st.session_state["dvs_last_refresh_seconds"] = refresh_seconds
                add_log(f"Data basis refreshed in mode={refresh_label} for dataset={dataset}, score_profile={bullish_score_profile} in {refresh_seconds:.1f}s.")
                st.session_state["dvs_last_action_message"] = f"Completed: {refresh_label} refresh in {refresh_seconds:.1f}s for {dataset} with profile {bullish_score_profile}."
                _persist_ui_runtime_state(
                    Path(export_dir).expanduser(),
                    refresh_seconds=refresh_seconds,
                    action_message=st.session_state["dvs_last_action_message"],
                )
                status = build_data_status_result(export_dir)
                pipeline_refresh_ok = True
                if full_refresh and not fast_pipeline:
                    st.success(f"Data basis refreshed in {refresh_label} mode in {refresh_seconds:.1f}s.")
                    st.rerun()

    if generate_watchlist or ((fast_pipeline or fast_refresh) and pipeline_refresh_ok):
        try:
            watchlist_started = time_module.perf_counter()
            if is_long_dip_mode:
                st.session_state["dvs_watchlist_result"] = None
            else:
                st.session_state["dvs_bullish_quality_result"] = None
            with st.spinner("Generating watchlist from the latest exported data..." if is_long_dip_mode else "Generating bullish-quality rankings from the latest exported data..."):
                if is_long_dip_mode:
                    screen_result = generate_watchlist_result(export_dir=Path(export_dir).expanduser(), cfg=watchlist_cfg)
                    trigger = _resolve_watchlist_snapshot_trigger(
                        generate_watchlist=generate_watchlist,
                        fast_pipeline=fast_pipeline,
                        fast_refresh=fast_refresh,
                    )
                    snapshot_history = _persist_watchlist_snapshot(
                        Path(export_dir).expanduser(),
                        screen_result,
                        trigger=trigger,
                    )
                    screen_result = _augment_watchlist_result_with_intraday_context(screen_result, snapshot_history)
                else:
                    bullish_result = generate_bullish_quality_scanner_result(export_dir=Path(export_dir).expanduser(), cfg=bullish_cfg)
                    screen_result = {
                        "generated_at": bullish_result.generated_at,
                        "trade_date": bullish_result.trade_date.isoformat() if bullish_result.trade_date else None,
                        "source_data_fetched_at": bullish_result.source_data_fetched_at,
                        "config_snapshot": bullish_result.config_snapshot,
                        "rankings_table": bullish_result.rankings_table,
                        "latest_window_table": bullish_result.latest_window_table,
                        "filter_diagnostics_table": bullish_result.filter_diagnostics_table,
                        "window_feature_table": bullish_result.window_feature_table,
                        "warnings": bullish_result.warnings,
                    }
            watchlist_seconds = time_module.perf_counter() - watchlist_started
            st.session_state["dvs_last_watchlist_seconds"] = watchlist_seconds
            if is_long_dip_mode:
                st.session_state["dvs_watchlist_result"] = screen_result
                txt_exports = _write_streamlit_watchlist_txt_exports(Path(export_dir).expanduser(), screen_result)
                if txt_exports:
                    add_log("TradingView watchlist TXT exported: " + ", ".join(str(path) for path in txt_exports.values()))
                if snapshot_history is not None and not snapshot_history.empty:
                    add_log(f"Watchlist snapshot updated: {_watchlist_snapshot_path(Path(export_dir).expanduser())}")
                add_log(f"Watchlist generated with top_n={top_n} in {watchlist_seconds:.1f}s.")
                st.session_state["dvs_last_action_message"] = f"Completed: watchlist generation in {watchlist_seconds:.1f}s."
                st.success(f"Watchlist updated in {watchlist_seconds:.1f}s.")
            else:
                st.session_state["dvs_bullish_quality_result"] = screen_result
                add_log(f"Bullish-quality scanner generated with top_n={top_n}, score_profile={bullish_score_profile} in {watchlist_seconds:.1f}s.")
                st.session_state["dvs_last_action_message"] = f"Completed: bullish-quality ranking in {watchlist_seconds:.1f}s."
                st.success(f"Bullish-quality scanner updated in {watchlist_seconds:.1f}s using profile {bullish_score_profile}.")
            _persist_ui_runtime_state(
                Path(export_dir).expanduser(),
                watchlist_seconds=watchlist_seconds,
                action_message=st.session_state["dvs_last_action_message"],
            )
        except Exception as exc:
            if is_long_dip_mode:
                st.session_state["dvs_watchlist_result"] = None
            else:
                st.session_state["dvs_bullish_quality_result"] = None
            screen_result = None
            add_log(f"Ranking generation failed: {type(exc).__name__}: {exc}")
            st.error(f"Ranking generation failed: {type(exc).__name__}: {exc}")

    autoload_ready = (Path(export_dir).expanduser() / "daily_symbol_features_full_universe.parquet").exists()
    if not is_long_dip_mode:
        autoload_ready = autoload_ready and (Path(export_dir).expanduser() / "premarket_window_features_full_universe.parquet").exists()
    if screen_result is None and autoload_ready:
        try:
            if is_long_dip_mode:
                screen_result = generate_watchlist_result(export_dir=Path(export_dir).expanduser(), cfg=watchlist_cfg)
                snapshot_history = _load_watchlist_snapshot_history(Path(export_dir).expanduser())
                screen_result = _augment_watchlist_result_with_intraday_context(screen_result, snapshot_history)
                st.session_state["dvs_watchlist_result"] = screen_result
            else:
                bullish_result = generate_bullish_quality_scanner_result(export_dir=Path(export_dir).expanduser(), cfg=bullish_cfg)
                screen_result = {
                    "generated_at": bullish_result.generated_at,
                    "trade_date": bullish_result.trade_date.isoformat() if bullish_result.trade_date else None,
                    "source_data_fetched_at": bullish_result.source_data_fetched_at,
                    "config_snapshot": bullish_result.config_snapshot,
                    "rankings_table": bullish_result.rankings_table,
                    "latest_window_table": bullish_result.latest_window_table,
                    "filter_diagnostics_table": bullish_result.filter_diagnostics_table,
                    "window_feature_table": bullish_result.window_feature_table,
                    "warnings": bullish_result.warnings,
                }
                st.session_state["dvs_bullish_quality_result"] = screen_result
        except Exception as exc:
            add_log(f"Initial ranking load failed: {type(exc).__name__}: {exc}")
            st.warning(f"Existing export data could not be loaded automatically: {type(exc).__name__}: {exc}")

    status_area, config_area = st.columns([1.1, 1.0])
    with status_area:
        st.subheader("Data Status")
        _streamlit_dataframe_compat(st.dataframe, build_status_table(status), hide_index=True)
    with config_area:
        st.subheader("Active Config")
        _streamlit_dataframe_compat(st.dataframe, build_config_table(config_snapshot), hide_index=True)

    if not screen_result:
        st.warning("No ranking available yet. Refresh the data basis or generate rankings from existing exports.")
    elif not is_long_dip_mode:
        st.subheader("Bullish-Quality Rankings")
        trade_date_label = screen_result.get("trade_date") or "n/a"
        fetched_at_label = screen_result.get("source_data_fetched_at") or "n/a"
        st.caption(f"Trade date: {trade_date_label} | Source data fetched at: {fetched_at_label}")
        trade_date_ts = pd.to_datetime(trade_date_label, errors="coerce")
        if pd.notna(trade_date_ts):
            today_et = datetime.now(US_EASTERN_TZ).date()
            if trade_date_ts.date() < today_et:
                st.warning(
                    "Bullish ranking uses previous trade date. Run Fast Pre-Open Refresh to update today's premarket windows, "
                    "or use Fast Pre-Open Pipeline to refresh + regenerate in one click."
                )

        for warning in screen_result.get("warnings", []):
            st.warning(warning)

        latest_window_table = screen_result.get("latest_window_table")
        rankings_table = screen_result.get("rankings_table")
        diagnostics_table = screen_result.get("filter_diagnostics_table")
        window_feature_table = screen_result.get("window_feature_table")

        if isinstance(latest_window_table, pd.DataFrame) and not latest_window_table.empty:
            st.markdown("**Latest Window Top-N**")
            _streamlit_dataframe_compat(
                st.dataframe,
                latest_window_table,
                hide_index=True,
                height=360,
                column_config={
                    "quality_rank_within_window": st.column_config.NumberColumn("Rank", width="small", format="%.0f"),
                    "symbol": st.column_config.TextColumn("Symbol", width="small"),
                    "window_tag": st.column_config.TextColumn("Window", width="small"),
                    "window_quality_score": st.column_config.NumberColumn("Score", width="small", format="%.2f"),
                    "window_structure_bias_score": st.column_config.NumberColumn("Structure Bias", width="small", format="%.1f"),
                    "window_structure_alignment_score": st.column_config.NumberColumn("Alignment", width="small", format="%.1f"),
                    "window_structure_last_event": st.column_config.TextColumn("Last Event", width="small"),
                    "quality_reason": st.column_config.TextColumn("Reason", width="medium"),
                },
            )
        else:
            st.info("No bullish-quality candidates matched the configured filters for the latest window.")

        if isinstance(diagnostics_table, pd.DataFrame) and not diagnostics_table.empty:
            with st.expander("Window Filter Diagnostics", expanded=True):
                _streamlit_dataframe_compat(st.dataframe, diagnostics_table, hide_index=True, height=260)

        if isinstance(rankings_table, pd.DataFrame) and not rankings_table.empty:
            with st.expander("All Window Rankings", expanded=False):
                rankings_table_display = _truncate_table_for_ui(
                    rankings_table,
                    label="All Window Rankings",
                    max_rows=1500,
                    max_cols=70,
                )
                _streamlit_dataframe_compat(st.dataframe, rankings_table_display, hide_index=True, height=420)

        if isinstance(window_feature_table, pd.DataFrame) and not window_feature_table.empty:
            with st.expander("Window Feature Detail", expanded=False):
                window_feature_table_display = _truncate_table_for_ui(
                    window_feature_table,
                    label="Window Feature Detail",
                    max_rows=1500,
                    max_cols=70,
                )
                _streamlit_dataframe_compat(st.dataframe, window_feature_table_display, hide_index=True, height=420)
    else:
        st.subheader("Top-N Watchlist")
        view_mode = st.radio(
            "Watchlist view",
            options=["Latest trade date", "Full history"],
            index=0,
            horizontal=True,
            key="dvs_watchlist_view_mode",
        )
        watchlist_table, watchlist_caption = resolve_watchlist_display_table(
            watchlist_result=screen_result,
            view_mode=view_mode,
        )
        st.caption(watchlist_caption)
        filter_profile = screen_result.get("filter_profile") or {}
        if filter_profile:
            st.info(
                "Active filter profile: "
                f"{filter_profile.get('profile_name', 'standard')} | "
                f"reason={filter_profile.get('profile_reason', 'n/a')} | "
                f"premarket_symbols={filter_profile.get('premarket_symbols', 'n/a')}"
            )
            if filter_profile.get("profile_name") == "liquidity_relaxed":
                relaxed_from_dollar_volume = float(filter_profile.get("relaxed_from_dollar_volume", 0.0) or 0.0)
                relaxed_from_volume_value = _safe_number(filter_profile.get("relaxed_from_volume"))
                relaxed_from_trade_count_value = _safe_number(filter_profile.get("relaxed_from_trade_count"))
                relaxed_from_active_seconds_value = _safe_number(filter_profile.get("relaxed_from_active_seconds"))
                relaxed_from_volume = f"{int(relaxed_from_volume_value):,}" if relaxed_from_volume_value is not None else "n/a"
                relaxed_from_trade_count = f"{int(relaxed_from_trade_count_value):,}" if relaxed_from_trade_count_value is not None else "n/a"
                relaxed_from_active_seconds = f"{int(relaxed_from_active_seconds_value):,}" if relaxed_from_active_seconds_value is not None else "n/a"
                relaxed_dollar_volume_to = float(screen_result["config_snapshot"].get("min_premarket_dollar_volume", 0.0) or 0.0)
                relaxed_volume_to = int(screen_result["config_snapshot"].get("min_premarket_volume", 0))
                relaxed_trade_count_to = int(screen_result["config_snapshot"].get("min_premarket_trade_count", 0))
                relaxed_active_seconds_to = int(screen_result["config_snapshot"].get("min_premarket_active_seconds", 0))
                st.warning(
                    "Liquidity fallback active: "
                    f"base_profile={filter_profile.get('base_profile_name', 'n/a')} | "
                    f"bottleneck={filter_profile.get('relaxed_bottleneck', 'n/a')} | "
                    f"premarket_dollar_volume {relaxed_from_dollar_volume:,.0f} -> {relaxed_dollar_volume_to:,.0f} | "
                    f"premarket_volume {relaxed_from_volume} -> {relaxed_volume_to:,} | "
                    f"premarket_trade_count {relaxed_from_trade_count} -> {relaxed_trade_count_to:,} | "
                    f"premarket_active_seconds {relaxed_from_active_seconds} -> {relaxed_active_seconds_to:,}"
                )
        for warning in screen_result.get("warnings", []):
            st.warning(warning)

        filter_funnel = screen_result.get("filter_funnel", [])
        if filter_funnel:
            with st.expander("Filter Funnel Diagnostic (latest trade date)", expanded=True):
                funnel_df = pd.DataFrame(filter_funnel)
                _streamlit_dataframe_compat(st.dataframe, funnel_df, hide_index=True)
                bottleneck = next((s for idx, s in enumerate(filter_funnel) if s["remaining"] == 0 and idx > 0), None)
                if bottleneck:
                    st.error(f"Bottleneck: **{bottleneck['filter']}** (threshold {bottleneck['threshold']}) eliminated all remaining candidates.")

        if watchlist_table.empty:
            st.info("No symbols matched the current filters.")
        else:
            focus_view = st.radio(
                "Open-window focus view",
                options=["All (04:00 + 08:00 + 09:30)", "09:30 only", "08:00 only", "04:00 only"],
                horizontal=True,
                key="open_window_focus_view",
            )

            preferred_columns = [
                "watchlist_rank",
                "watchlist_rank_change",
                "watchlist_rank_delta",
                "intraday_watchlist_rank_change",
                "intraday_watchlist_rank_delta",
                "previous_watchlist_rank",
                "previous_intraday_watchlist_rank",
                "intraday_rank_reference_display",
                "symbol",
                "focus_window_coverage",
                "structure_trend_state",
                "structure_last_event",
                "structure_bias_score",
                "structure_alignment_score",
                "structure_break_quality_score",
                "structure_pressure_score",
                "structure_reclaim_flag",
                "structure_failed_break_flag",
                "premarket_trade_count",
                "premarket_trade_count_age",
                "prev_close_to_premarket_pct",
                "premarket_dollar_volume",
                "premarket_volume",
                "reclaimed_start_price_within_30s",
                "reclaim_second_30s",
                "open_30s_volume",
                "focus_0930_open_30s_volume",
                "focus_0800_open_30s_volume",
                "focus_0400_open_30s_volume",
                "early_dip_second",
                "trade_date",
                "premarket_last",
                "early_dip_pct_10s",
                "focus_0930_early_dip_pct_10s",
                "focus_0800_early_dip_pct_10s",
                "focus_0400_early_dip_pct_10s",
                "focus_0930_reclaim_second_30s",
                "focus_0800_reclaim_second_30s",
                "focus_0400_reclaim_second_30s",
                "open_pattern_status",
                "l1_limit_buy",
                "l1_take_profit",
                "l1_stop_loss",
                "l2_limit_buy",
                "l2_take_profit",
                "l2_stop_loss",
                "l3_limit_buy",
                "l3_take_profit",
                "l3_stop_loss",
                "position_budget_usd",
            ]
            display_watchlist_table = watchlist_table.copy()
            if focus_view == "09:30 only":
                if "focus_0930_open_30s_volume" in display_watchlist_table.columns:
                    display_watchlist_table["open_30s_volume"] = display_watchlist_table["focus_0930_open_30s_volume"]
                if "focus_0930_early_dip_pct_10s" in display_watchlist_table.columns:
                    display_watchlist_table["early_dip_pct_10s"] = display_watchlist_table["focus_0930_early_dip_pct_10s"]
                if "focus_0930_reclaim_second_30s" in display_watchlist_table.columns:
                    display_watchlist_table["reclaim_second_30s"] = display_watchlist_table["focus_0930_reclaim_second_30s"]
            elif focus_view == "08:00 only":
                if "focus_0800_open_30s_volume" in display_watchlist_table.columns:
                    display_watchlist_table["open_30s_volume"] = display_watchlist_table["focus_0800_open_30s_volume"]
                if "focus_0800_early_dip_pct_10s" in display_watchlist_table.columns:
                    display_watchlist_table["early_dip_pct_10s"] = display_watchlist_table["focus_0800_early_dip_pct_10s"]
                if "focus_0800_reclaim_second_30s" in display_watchlist_table.columns:
                    display_watchlist_table["reclaim_second_30s"] = display_watchlist_table["focus_0800_reclaim_second_30s"]
            elif focus_view == "04:00 only":
                if "focus_0400_open_30s_volume" in display_watchlist_table.columns:
                    display_watchlist_table["open_30s_volume"] = display_watchlist_table["focus_0400_open_30s_volume"]
                if "focus_0400_early_dip_pct_10s" in display_watchlist_table.columns:
                    display_watchlist_table["early_dip_pct_10s"] = display_watchlist_table["focus_0400_early_dip_pct_10s"]
                if "focus_0400_reclaim_second_30s" in display_watchlist_table.columns:
                    display_watchlist_table["reclaim_second_30s"] = display_watchlist_table["focus_0400_reclaim_second_30s"]

            display_watchlist_table["open_pattern_status"] = _build_open_pattern_status_series(
                watchlist_table,
                focus_view,
            )
            display_watchlist_table["focus_window_coverage"] = _build_focus_window_coverage_series(watchlist_table)
            if "watchlist_rank_change" in display_watchlist_table.columns:
                display_watchlist_table["watchlist_rank_change"] = [
                    _highlight_rank_change_label(label, delta)
                    for label, delta in zip(
                        display_watchlist_table["watchlist_rank_change"],
                        display_watchlist_table.get("watchlist_rank_delta", pd.Series(index=display_watchlist_table.index, dtype=float)),
                        strict=False,
                    )
                ]
            if "intraday_watchlist_rank_change" in display_watchlist_table.columns:
                display_watchlist_table["intraday_watchlist_rank_change"] = [
                    _highlight_rank_change_label(label, delta)
                    for label, delta in zip(
                        display_watchlist_table["intraday_watchlist_rank_change"],
                        display_watchlist_table.get("intraday_watchlist_rank_delta", pd.Series(index=display_watchlist_table.index, dtype=float)),
                        strict=False,
                    )
                ]
            if "intraday_rank_reference_at" in display_watchlist_table.columns:
                display_watchlist_table["intraday_rank_reference_display"] = display_watchlist_table[
                    "intraday_rank_reference_at"
                ].map(lambda value: _format_intraday_reference_time(value, display_timezone=DEFAULT_DISPLAY_TZ))
            visible_columns = [column for column in preferred_columns if column in display_watchlist_table.columns]
            focus_0930_cols = ["focus_0930_open_30s_volume", "focus_0930_early_dip_pct_10s", "focus_0930_reclaim_second_30s"]
            focus_0800_cols = ["focus_0800_open_30s_volume", "focus_0800_early_dip_pct_10s", "focus_0800_reclaim_second_30s"]
            focus_0400_cols = ["focus_0400_open_30s_volume", "focus_0400_early_dip_pct_10s", "focus_0400_reclaim_second_30s"]
            if focus_view == "09:30 only":
                visible_columns = [col for col in visible_columns if col not in focus_0800_cols and col not in focus_0930_cols and col not in focus_0400_cols]
            elif focus_view == "08:00 only":
                visible_columns = [col for col in visible_columns if col not in focus_0800_cols and col not in focus_0930_cols and col not in focus_0400_cols]
            elif focus_view == "04:00 only":
                visible_columns = [col for col in visible_columns if col not in focus_0800_cols and col not in focus_0930_cols and col not in focus_0400_cols]
            if "open_pattern_status" not in visible_columns and "open_pattern_status" in display_watchlist_table.columns:
                visible_columns = ["open_pattern_status", *visible_columns]
            if "focus_window_coverage" not in visible_columns and "focus_window_coverage" in display_watchlist_table.columns:
                visible_columns = ["focus_window_coverage", *visible_columns]
            table_frame = display_watchlist_table[visible_columns].copy()
            if "reclaimed_start_price_within_30s" in table_frame.columns:
                table_frame["reclaimed_start_price_within_30s"] = _format_reclaim_status_series(display_watchlist_table)
            table_frame = _truncate_table_for_ui(
                table_frame,
                label="Top-N Watchlist",
                max_rows=1200,
                max_cols=70,
            )
            st.caption(
                "`All` means at least one focus window has usable detail. "
                "Single-window views check only that exact 04:00 / 08:00 / 09:30 slice for the dip/reclaim metrics. "
                "`Focus Window Coverage` lists the windows that actually have usable second-detail rows for that symbol-day. "
                "`Rank Change` compares against the last available watchlist day for the same symbol, and positive deltas mean the symbol moved up the list. "
                "`Intraday Rank Change` compares against the previous saved watchlist snapshot on the same trade date; `Intraday Ref Time` shows that earlier snapshot in Europe/Berlin, and `FIRST` means no earlier snapshot has been captured yet. "
                "`PM Trade Count Age` measures the lag between the latest premarket trade used in the count and the source-data timestamp. "
                "`n/a` means the selected slice has no usable regular-open second-detail rows."
            )
            column_config: dict[str, Any] = {}
            if "watchlist_rank" in table_frame.columns:
                column_config["watchlist_rank"] = st.column_config.NumberColumn("Rank", width="small")
            if "watchlist_rank_change" in table_frame.columns:
                column_config["watchlist_rank_change"] = st.column_config.TextColumn("Day Rank\nChange", width="small")
            if "watchlist_rank_delta" in table_frame.columns:
                column_config["watchlist_rank_delta"] = st.column_config.NumberColumn("Day Δ", width="small", format="%+.0f")
            if "intraday_watchlist_rank_change" in table_frame.columns:
                column_config["intraday_watchlist_rank_change"] = st.column_config.TextColumn("Intraday Rank\nChange", width="small")
            if "intraday_watchlist_rank_delta" in table_frame.columns:
                column_config["intraday_watchlist_rank_delta"] = st.column_config.NumberColumn("Intraday Δ", width="small", format="%+.0f")
            if "previous_watchlist_rank" in table_frame.columns:
                column_config["previous_watchlist_rank"] = st.column_config.NumberColumn("Prev Day\nRank", width="small")
            if "previous_intraday_watchlist_rank" in table_frame.columns:
                column_config["previous_intraday_watchlist_rank"] = st.column_config.NumberColumn("Prev Intraday\nRank", width="small")
            if "intraday_rank_reference_display" in table_frame.columns:
                column_config["intraday_rank_reference_display"] = st.column_config.TextColumn("Intraday Ref\nTime", width="medium")
            if "symbol" in table_frame.columns:
                column_config["symbol"] = st.column_config.TextColumn("Symbol", width="small")
            if "focus_window_coverage" in table_frame.columns:
                column_config["focus_window_coverage"] = st.column_config.TextColumn("Focus Window\nCoverage", width="medium")
            if "structure_trend_state" in table_frame.columns:
                column_config["structure_trend_state"] = st.column_config.NumberColumn("Trend", width="small", format="%.0f")
            if "structure_last_event" in table_frame.columns:
                column_config["structure_last_event"] = st.column_config.TextColumn("Last Event", width="small")
            if "structure_bias_score" in table_frame.columns:
                column_config["structure_bias_score"] = st.column_config.NumberColumn("Structure Bias", width="small", format="%.1f")
            if "structure_alignment_score" in table_frame.columns:
                column_config["structure_alignment_score"] = st.column_config.NumberColumn("Alignment", width="small", format="%.1f")
            if "structure_break_quality_score" in table_frame.columns:
                column_config["structure_break_quality_score"] = st.column_config.NumberColumn("Break Quality", width="small", format="%.1f")
            if "structure_pressure_score" in table_frame.columns:
                column_config["structure_pressure_score"] = st.column_config.NumberColumn("Pressure", width="small", format="%.1f")
            if "structure_reclaim_flag" in table_frame.columns:
                column_config["structure_reclaim_flag"] = st.column_config.CheckboxColumn("Reclaim", width="small")
            if "structure_failed_break_flag" in table_frame.columns:
                column_config["structure_failed_break_flag"] = st.column_config.CheckboxColumn("Failed Break", width="small")
            if "premarket_trade_count_age" in table_frame.columns:
                column_config["premarket_trade_count_age"] = st.column_config.TextColumn("PM Trade\nCount Age", width="small")
            if "reclaimed_start_price_within_30s" in table_frame.columns:
                column_config["reclaimed_start_price_within_30s"] = st.column_config.TextColumn("Reclaimed\nStart Price\n30s", width="small")
            if "reclaim_second_30s" in table_frame.columns:
                column_config["reclaim_second_30s"] = st.column_config.TextColumn("Reclaim\nSecond\n30s", width="small")
            if "open_30s_volume" in table_frame.columns:
                column_config["open_30s_volume"] = st.column_config.TextColumn("Open\n30s\nVolume", width="small")
            if "focus_0930_open_30s_volume" in table_frame.columns:
                column_config["focus_0930_open_30s_volume"] = st.column_config.TextColumn("09:30\nOpen 30s\nVolume", width="small")
            if "focus_0800_open_30s_volume" in table_frame.columns:
                column_config["focus_0800_open_30s_volume"] = st.column_config.TextColumn("08:00\nOpen 30s\nVolume", width="small")
            if "focus_0930_early_dip_pct_10s" in table_frame.columns:
                column_config["focus_0930_early_dip_pct_10s"] = st.column_config.NumberColumn("09:30\nEarly Dip %\n10s", width="small", format="%.2f")
            if "focus_0800_early_dip_pct_10s" in table_frame.columns:
                column_config["focus_0800_early_dip_pct_10s"] = st.column_config.NumberColumn("08:00\nEarly Dip %\n10s", width="small", format="%.2f")
            if "focus_0930_reclaim_second_30s" in table_frame.columns:
                column_config["focus_0930_reclaim_second_30s"] = st.column_config.TextColumn("09:30\nReclaim Sec\n30s", width="small")
            if "focus_0800_reclaim_second_30s" in table_frame.columns:
                column_config["focus_0800_reclaim_second_30s"] = st.column_config.TextColumn("08:00\nReclaim Sec\n30s", width="small")
            if "prev_close_to_premarket_pct" in table_frame.columns:
                column_config["prev_close_to_premarket_pct"] = st.column_config.NumberColumn("Prev Close ->\nPremarket %", width="small", format="%.2f")
            if "premarket_dollar_volume" in table_frame.columns:
                column_config["premarket_dollar_volume"] = st.column_config.NumberColumn("Premarket\n$ Volume", width="small", format="%.0f")
            if "premarket_volume" in table_frame.columns:
                column_config["premarket_volume"] = st.column_config.NumberColumn("Premarket\nVolume", width="small", format="%.0f")
            if "open_pattern_status" in table_frame.columns:
                column_config["open_pattern_status"] = st.column_config.TextColumn("Open Pattern\nStatus", width="small")

            if len(table_frame) <= 400 and table_frame.shape[1] <= 60:
                table_styles = _build_watchlist_table_style_frame(table_frame)
                styled_table_frame = table_frame.style.apply(lambda _: table_styles, axis=None)
                _streamlit_dataframe_compat(
                    st.dataframe,
                    styled_table_frame,
                    hide_index=True,
                    height=420,
                    column_config=column_config,
                )
            else:
                st.info("Styling disabled for large watchlist table to avoid browser payload limits.")
                _streamlit_dataframe_compat(
                    st.dataframe,
                    table_frame,
                    hide_index=True,
                    height=420,
                    column_config=column_config,
                )

            active_symbols_source = screen_result.get("active_watchlist_table")
            if not isinstance(active_symbols_source, pd.DataFrame) or active_symbols_source.empty:
                active_symbols_source = watchlist_table
            active_symbols = []
            if isinstance(active_symbols_source, pd.DataFrame) and "symbol" in active_symbols_source.columns:
                active_symbols = active_symbols_source["symbol"].astype(str).head(int(top_n)).tolist()
            snapshot_summary_frame, snapshot_trail_frame = _build_watchlist_snapshot_panel_frames(
                snapshot_history,
                trade_date=screen_result.get("trade_date"),
                active_symbols=active_symbols,
                display_timezone=DEFAULT_DISPLAY_TZ,
            )
            with st.expander("Intraday Snapshot History", expanded=False):
                st.caption(
                    "Saved same-day watchlist snapshots from "
                    f"{screen_result.get('trade_date') or 'n/a'} in {DEFAULT_DISPLAY_TZ}. "
                    f"Source file: {WATCHLIST_SNAPSHOT_FILE}."
                )
                if snapshot_summary_frame.empty:
                    st.info("No intraday snapshots have been captured for the active trade date yet.")
                else:
                    snapshot_cols = st.columns([1.0, 1.4])
                    with snapshot_cols[0]:
                        st.markdown("**Snapshot Summary**")
                        _streamlit_dataframe_compat(
                            st.dataframe,
                            snapshot_summary_frame,
                            hide_index=True,
                            height=240,
                            column_config={
                                "snapshot_time": st.column_config.TextColumn("Snapshot Time", width="medium"),
                                "trigger": st.column_config.TextColumn("Trigger", width="small"),
                                "symbols": st.column_config.NumberColumn("Symbols", width="small", format="%.0f"),
                                "leader": st.column_config.TextColumn("Leader", width="small"),
                                "top3": st.column_config.TextColumn("Top 3", width="medium"),
                            },
                        )
                    with snapshot_cols[1]:
                        st.markdown("**Rank Trail (Current Top-N Symbols)**")
                        _streamlit_dataframe_compat(
                            st.dataframe,
                            snapshot_trail_frame,
                            hide_index=True,
                            height=240,
                        )

            st.subheader("Detail View (all Top-N entries)")
            detail_slice = watchlist_table.head(int(top_n)).reset_index(drop=True)

            def _style_checklist_row(row: pd.Series) -> list[str]:
                status_label = str(row.get("status") or "")
                if status_label == "erfuellt":
                    bg = "#dcfce7"
                elif status_label == "grenzwertig":
                    bg = "#fef3c7"
                else:
                    bg = "#fee2e2"
                return [f"background-color: {bg}; color: #000000" for _ in row.index]

            for detail_idx, selected_row in detail_slice.iterrows():
                rank_value = selected_row.get("watchlist_rank")
                rank_label = int(rank_value) if pd.notna(rank_value) else detail_idx + 1
                symbol_label = str(selected_row.get("symbol") or "n/a")
                trade_date_label = str(selected_row.get("trade_date") or "n/a")
                with st.expander(f"{trade_date_label} | #{rank_label} | {symbol_label}", expanded=(detail_idx == 0)):
                    detail_frame = pd.DataFrame(
                        {
                            "field": selected_row.index.astype(str).tolist(),
                            "value": [str(value) for value in selected_row.tolist()],
                        }
                    )
                    _streamlit_dataframe_compat(st.dataframe, detail_frame, hide_index=True, height=260)

                    checklist_frame, checklist_note, checklist_score = build_entry_checklist_table(
                        status=status,
                        selected_row=selected_row,
                        watchlist_table=watchlist_table,
                        watchlist_config=screen_result.get("requested_config_snapshot") or screen_result.get("config_snapshot"),
                    )
                    checklist_styler = checklist_frame.style.apply(_style_checklist_row, axis=1)
                    checklist_styler = checklist_styler.format(
                        {
                            "erfuellt": lambda value: "TRUE" if bool(value) else "FALSE",
                            "status": lambda value: {
                                "erfuellt": "Gruen",
                                "grenzwertig": "Gelb",
                                "nicht_erfuellt": "Rot",
                            }.get(str(value), str(value)),
                        }
                    )

                    st.markdown(
                        (
                            "<div style='margin-bottom: 0.25rem;'>"
                            "<div style='font-size: 0.875rem; font-weight: 600; color: #000000;'>Checklist Score</div>"
                            f"<div style='font-size: 2rem; line-height: 1.1; font-weight: 700; color: #000000;'>{checklist_score}/5</div>"
                            "</div>"
                        ),
                        unsafe_allow_html=True,
                    )
                    st.caption(checklist_note)
                    _streamlit_dataframe_compat(st.dataframe, checklist_styler, hide_index=True)

    with st.expander("Run Logs", expanded=False):
        if st.session_state["dvs_run_logs"]:
            st.text("\n".join(st.session_state["dvs_run_logs"]))
        else:
            st.caption("No actions executed in this session yet.")


if __name__ == "__main__" and sys.argv and sys.argv[0].endswith("databento_volatility_screener.py"):
    run_streamlit_app()